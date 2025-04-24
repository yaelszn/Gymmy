# -*- coding: utf-8 -*-
import queue
import subprocess
import sys
import tkinter as tk
import webbrowser
from datetime import datetime
from statistics import mean, stdev
from tkinter import ttk
import random
from datetime import datetime

import numpy as np
import openpyxl
import matplotlib
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtWidgets import QVBoxLayout, QLabel
from matplotlib import pyplot as plt


from PyZedWrapper import PyZedWrapper


matplotlib.use('TkAgg')  # Use the TkAgg backend
import cv2
import pandas as pd
import pygame
from PIL import Image, ImageTk, ImageSequence
from email_validator import validate_email, EmailNotValidError
import Settings as s
from Audio import say, get_wav_duration, ContinuousAudio
from gtts import gTTS
import os
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
import Excel
import re
import cv2
import time
import os
import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
import pandas as pd


s.exercise_name_repeated_explanation = None

class Screen(tk.Tk):
    def __init__(self):
        print("screen start")
        tk.Tk.__init__(self, className='Gymmy')
        self._frame = None
        self["bg"]="#F3FCFB"
        #self.after(1000, lambda: self.pause_execution())

        # # Create a top bar
        # self.bar = tk.Frame(self, bg="#4CAF50", height=50)
        # self.bar.place(x=0, y=0, width=self.winfo_screenwidth(), height=50)
        #
        # # Add a scale to the bar
        # self.scale = tk.Scale(
        #     self.bar,
        #     from_=0, to=100,
        #     orient="horizontal",
        #     bg="#4CAF50",
        #     fg="white",
        #     highlightthickness=0,
        #     font=("Arial", 10),
        #     length=200
        # )
        # self.scale.pack(side="left", padx=20)

        shut_down_button_img = Image.open("Pictures//shut_down.jpg")  # Change path to your image file
        shut_down_button_photo = ImageTk.PhotoImage(shut_down_button_img)

        shut_down_button = tk.Button(self, image=shut_down_button_photo, command=lambda: self.on_click_shut_down(),
                                              width=shut_down_button_img.width, height=shut_down_button_img.height, bd=0, highlightthickness=0)  # Set border width to 0 to remove button border
        shut_down_button.image = shut_down_button_photo  # Store reference to image to prevent garbage collection
        shut_down_button.place(x=60, y=60)

    def pause_execution(self):
        # Create a dummy function that does nothing
        pass

    def switch_frame(self, frame_class, **kwargs):
        """Destroys all existing frames and creates a new one safely from any thread."""

        def switch():
            # Destroy all existing frames
            for widget in self.winfo_children():
                widget.destroy()

            # Create a new frame
            new_frame = frame_class(self, **kwargs)
            self._frame = new_frame
            self._frame.pack()

        # Use after() to ensure it runs in the main thread
        self.after(0, switch)

    def quit(self):
        self.destroy()



class EntrancePage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)

        image = Image.open('Pictures//background.jpg')
        self.photo_image = ImageTk.PhotoImage(image) #self. - for keeping the photo in memory so it will be shown
        tk.Label(self, image = self.photo_image).pack()


        # Set initial value
        s.chosen_patient_ID = None

        # Load images for buttons
        therapist_image = Image.open("Pictures//therapist_entrance_button.jpg")
        therapist_photo = ImageTk.PhotoImage(therapist_image)
        patient_image = Image.open("Pictures//patient_entrance_button.jpg")
        patient_photo = ImageTk.PhotoImage(patient_image)

        # Store references to prevent garbage collection
        self.therapist_photo = therapist_photo
        self.patient_photo = patient_photo

        # Create buttons with images
        enter_as_therapist_button = tk.Button(self, image=therapist_photo,
                                              command=self.on_click_therapist_chosen,
                                              width=therapist_image.width, height=therapist_image.height,
                                              bg='#50a6ad', bd=0, highlightthickness=0)
        enter_as_therapist_button.place(x=160, y=130)

        enter_as_patient_button = tk.Button(self, image=patient_photo,
                                            command=self.on_click_patient_chosen,
                                            width=patient_image.width, height=patient_image.height,
                                            bg='#50a6ad', bd=0, highlightthickness=0)
        enter_as_patient_button.place(x=540, y=130)

        shut_down_button_img = Image.open("Pictures//shut_down.jpg")  # Change path to your image file
        shut_down_button_photo = ImageTk.PhotoImage(shut_down_button_img)

        shut_down_button = tk.Button(self, image=shut_down_button_photo, command=lambda: self.on_click_shut_down(),
                                              width=shut_down_button_img.width, height=shut_down_button_img.height, bd=0, highlightthickness=0)  # Set border width to 0 to remove button border
        shut_down_button.image = shut_down_button_photo  # Store reference to image to prevent garbage collection
        shut_down_button.place(x=30, y=30)




        s.ex_in_training = []

    def on_click_shut_down(self):
        s.finish_program = True  # Signal all threads to stop
        s.req_exercise=""
        s.camera.join()
        s.training.join()
        s.robot.join()
        s.continuous_audio.join()
        # s.zed_camera.join()
        print("All threads have stopped.")
        self.quit()

    def on_click_therapist_chosen(self):
        s.screen.switch_frame(ID_therapist_fill_page)

    def on_click_patient_chosen(self):
        s.screen.switch_frame(ID_patient_fill_page)

#enterence page of patient
class ID_patient_fill_page(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//IDFillPage.jpg')
        self.photo_image = ImageTk.PhotoImage(image)  # self. - for keeping the photo in memory so it will be shown
        tk.Label(self, image=self.photo_image).pack()
        self.user_id_entry_patient = tk.Entry(self, font=('Thaoma', 14), width=20)
        self.user_id_entry_patient.place(x=400, y=260)
        self.labels = []

        to_previous_button_img = Image.open("Pictures//previous.jpg")
        to_previous_button_photo = ImageTk.PhotoImage(to_previous_button_img)
        to_previous_button = tk.Button(self, image=to_previous_button_photo, command=lambda: self.to_previous_button_click(),
                                   width=to_previous_button_img.width, height=to_previous_button_img.height, bd=0,
                                   highlightthickness=0)  # Set border width to 0 to remove button border
        to_previous_button.image = to_previous_button_photo  # Store reference to image to prevent garbage collection
        to_previous_button.place(x=30, y=30)

        # Load image for the submit button
        submit_image = Image.open("Pictures//enter.jpg")
        submit_photo = ImageTk.PhotoImage(submit_image)

        # Create button with image
        submit_button = tk.Button(self, image=submit_photo,
                                  command=lambda: self.on_submit_patient(),
                                  width=submit_image.width, height=submit_image.height,
                                  bg='#50a6ad', bd=0, highlightthickness=0)  # Set highlightthickness to 0
        submit_button.image = submit_photo  # Store reference to image to prevent garbage collection
        submit_button.place(x=460, y=300)  # Adjust x and y coordinates as needed


    def to_previous_button_click(self):
        s.screen.switch_frame(EntrancePage)

    def on_submit_patient(self):
        self.delete_all_labels()
        user_id = self.user_id_entry_patient.get()
        print(f"User ID entered: {user_id}")
        # Add your logic here to handle the user ID, e.g., store it in a variable or perform an action.

        if user_id != "":
            excel_file_path = "Patients.xlsx"
            df = pd.read_excel(excel_file_path, sheet_name="patients_details")

            # Convert the first column to string for comparison
            df.iloc[:, 0] = df.iloc[:, 0].astype(str)

            # Convert the user_id to string and remove leading/trailing spaces for comparison
            user_id_cleaned = str(user_id).strip()

            # Filter rows based on the condition
            row_of_patient = df[df.iloc[:, 0] == user_id_cleaned]

            if row_of_patient.empty:
                # empty image as background for label
                back = Image.open('Pictures//id_not_in_system.jpg')
                background_img = ImageTk.PhotoImage(back)

                self.label = tk.Label(self, image=background_img, compound=tk.CENTER, highlightthickness=0)
                self.label.place(x=320, y=360)
                self.label.image = background_img
                self.labels.append(self.label)

            else:
                ezer=row_of_patient.iloc[0]["number of exercises"]

                if ezer==0:
                    # empty image as background for label
                    back = Image.open('Pictures//no_exercsies_for_patient.jpg')
                    background_img = ImageTk.PhotoImage(back)

                    self.label = tk.Label(self, image=background_img, compound=tk.CENTER, highlightthickness=0)
                    self.label.place(x=260, y=360)
                    self.label.image = background_img
                    self.labels.append(self.label)



                else:
                    patient_workbook_path= "Patients.xlsx"
                    s.chosen_patient_ID=user_id
                    # s.email_of_patient= Excel.find_value_by_colName_and_userID(patient_workbook_path, "patients_details", user_id, "email")
                    # s.current_level_of_patient= Excel.find_value_by_colName_and_userID(patient_workbook_path, "patients_details", user_id, "level")
                    # s.points_in_current_level_before_training= Excel.find_value_by_colName_and_userID(patient_workbook_path, "patients_details", user_id, "points in current level")
                    s.rep= int(Excel.find_value_by_colName_and_userID(patient_workbook_path, "patients_details", user_id, "number of repetitions in each exercise"))
                    s.gender = Excel.find_value_by_colName_and_userID(patient_workbook_path, "patients_details", user_id, "gender")
                    s.rate = Excel.find_value_by_colName_and_userID(patient_workbook_path, "patients_details", user_id, "rate")
                    # s.dist_between_shoulders = Excel.find_value_by_colName_and_userID(patient_workbook_path, "patients_details", user_id, "dist between shoulders")
                    # s.full_name = Excel.find_value_by_colName_and_userID(patient_workbook_path, "patients_details", user_id, "first name") + " " + Excel.find_value_by_colName_and_userID(patient_workbook_path, "patients_details", user_id, "last name")
                    s.audio_path = 'audio files/Hebrew/' + s.gender + '/'

                    df1 = pd.read_excel(excel_file_path, sheet_name="patients_exercises")

                    # Convert the first column to string for comparison
                    df1.iloc[:, 0] = df1.iloc[:, 0].astype(str)


                    # Filter rows based on the condition
                    row_of_patient = df1[df1.iloc[:, 0] == user_id_cleaned]


                    s.ex_in_training=[]
                    num_columns1= df1.shape[1]


                    for i in range (0,num_columns1): #including 0
                        if row_of_patient.iloc[0,i]==True:
                            s.ex_in_training.append(df1.columns[i])

                    print(s.ex_in_training)
                    s.screen.switch_frame(StartOfTraining)


        else:
            back = Image.open('Pictures//no_id.jpg')
            background_img = ImageTk.PhotoImage(back)

            self.label = tk.Label(self, image=background_img, compound=tk.CENTER, highlightthickness=0)
            self.label.place(x=320, y=360)
            self.label.image = background_img
            self.labels.append(self.label)

    def delete_all_labels(self):
        for label in self.labels:
            label.destroy()

        self.labels=[]

#enterence page of therapist
class ID_therapist_fill_page(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//IDFillPage.jpg')
        self.photo_image = ImageTk.PhotoImage(image) #self. - for keeping the photo in memory so it will be shown
        tk.Label(self, image = self.photo_image).pack()
        self.user_id_entry_physio= tk.Entry(self, font=('Thaoma', 14), width=20)
        self.user_id_entry_physio.place(x=400, y=260)
        self.labels=[]


        to_previous_button_img = Image.open("Pictures//previous.jpg")
        to_previous_button_photo = ImageTk.PhotoImage(to_previous_button_img)
        to_previous_button = tk.Button(self, image=to_previous_button_photo, command=lambda: self.to_previous_button_click(),
                                   width=to_previous_button_img.width, height=to_previous_button_img.height, bd=0,
                                   highlightthickness=0)  # Set border width to 0 to remove button border
        to_previous_button.image = to_previous_button_photo  # Store reference to image to prevent garbage collection
        to_previous_button.place(x=30, y=30)

        # Load image for the submit button
        submit_image = Image.open("Pictures//enter.jpg")  # Change path to your image file
        submit_photo = ImageTk.PhotoImage(submit_image)

        # Create button with image
        submit_button = tk.Button(self, image=submit_photo,
                                  command=lambda: self.on_submit_physio(),
                                  width=submit_image.width, height=submit_image.height,
                                  bg='#50a6ad', bd=0, highlightthickness=0)  # Set highlightthickness to 0
        submit_button.image = submit_photo  # Store reference to image to prevent garbage collection
        submit_button.place(x=460, y=300)  # Adjust x and y coordinates as needed

    def to_previous_button_click(self):
        s.screen.switch_frame(EntrancePage)

    def on_submit_physio(self):
        self.delete_all_labels()
        user_id = self.user_id_entry_physio.get()
        print(f"User ID entered: {user_id}")
        # Add your logic here to handle the user ID, e.g., store it in a variable or perform an action.

        if user_id != "":
            excel_file_path = "Physiotherapists.xlsx"
            df = pd.read_excel(excel_file_path, sheet_name="details")

            # Convert the first column to string for comparison
            df['ID_str'] = df.iloc[:, 0].astype(str).str.strip()

            # Convert the user_id to string and remove leading/trailing spaces for comparison
            user_id_cleaned = str(user_id).strip()

            # Filter rows based on the condition
            row_of_patient = df[df['ID_str'] == user_id_cleaned]

            if row_of_patient.empty:
                back = Image.open('Pictures//id_not_in_system.jpg')
                background_img = ImageTk.PhotoImage(back)

                self.label = tk.Label(self, image=background_img, compound=tk.CENTER, highlightthickness=0)
                self.label.place(x=250, y=360)
                self.label.image = background_img
                self.labels.append(self.label)



            else:
                s.screen.switch_frame(Choose_Action_Physio)


        else:
            back = Image.open('Pictures//no_id.jpg')
            background_img = ImageTk.PhotoImage(back)

            self.label = tk.Label(self, image=background_img, compound=tk.CENTER, highlightthickness=0)
            self.label.place(x=320, y=360)
            self.label.image = background_img
            self.labels.append(self.label)


    def delete_all_labels(self):
        for label in self.labels:
            label.destroy()

        self.labels = []


class Choose_Action_Physio(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)

        # Create the canvas for everything
        self.canvas = tk.Canvas(self, width=1024, height=576, bd=0, highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)

        # Load and display the background image on the canvas
        image = Image.open("Pictures//background_physical_therapist.jpg")
        self.photo_image = ImageTk.PhotoImage(image)
        self.canvas.create_image(0, 0, image=self.photo_image, anchor="nw")

        self.options = ["גבר", "אישה"]
        self.gender = 'Male'

        self.selected_option = tk.StringVar()
        self.selected_option.set(self.options[0])

        # Exit button
        exit_button_img = Image.open("Pictures//exit_system.jpg")
        exit_button_photo = ImageTk.PhotoImage(exit_button_img)
        exit_button = tk.Button(self, image=exit_button_photo, command=self.on_click_quit,
                                width=exit_button_img.width, height=exit_button_img.height,
                                bd=0, highlightthickness=0)
        exit_button.image = exit_button_photo
        exit_button_window = self.canvas.create_window(30, 30, anchor="nw", window=exit_button)

        # Buttons
        therapist_register_button_img = Image.open("Pictures//add_physio.jpg")
        therapist_register_button_photo = ImageTk.PhotoImage(therapist_register_button_img)

        patient_register_button_img = Image.open("Pictures//add_patient.jpg")
        patient_register_button_photo = ImageTk.PhotoImage(patient_register_button_img)

        go_to_training_sessions_page_button_img = Image.open("Pictures//to_patients_list.jpg")
        go_to_training_sessions_page_button_photo = ImageTk.PhotoImage(go_to_training_sessions_page_button_img)

        # Training session button
        go_button = tk.Button(self, image=go_to_training_sessions_page_button_photo,
                              command=self.on_go_to_training_sessions_page_click,
                              width=go_to_training_sessions_page_button_img.width,
                              height=go_to_training_sessions_page_button_img.height,
                              bg='#50a6ad', bd=0, highlightthickness=0)
        go_button.image = go_to_training_sessions_page_button_photo
        self.canvas.create_window(225, 90, anchor="nw", window=go_button)

        # Entry for ID
        self.id_entry = tk.Entry(self, font=('Thaoma', 14), width=10, justify='right')
        self.canvas.create_window(430, 415, anchor="nw", window=self.id_entry)

        # Dropdown for gender
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Custom.TMenubutton', font=('Arial', 12, 'bold'), background='lightgray', foreground='black', anchor='e')

        self.option_menu = ttk.OptionMenu(self, self.selected_option, self.selected_option.get(), *self.options, command=self.on_select_gender)
        self.option_menu['style'] = 'Custom.TMenubutton'
        self.option_menu.config(width=6)
        self.canvas.create_window(240, 408, anchor="nw", window=self.option_menu)

        add_patient_button_img = Image.open("Pictures//add.jpg")
        add_patient_button_photo = ImageTk.PhotoImage(add_patient_button_img)

        # Add patient button (use the same photo for both display and sizing)
        add_patient_button = tk.Button(self, image=add_patient_button_photo,
                                       command=self.on_click_patient_registration,
                                       width=add_patient_button_img.width,
                                       height=add_patient_button_img.height,
                                       bd=0, highlightthickness=0)
        add_patient_button.image = add_patient_button_photo  # Store reference
        self.canvas.create_window(70, 400, anchor="nw", window=add_patient_button)

        self.labels = []  # canvas text/shape item IDs

    def on_select_gender(self, option):
        self.gender = 'Female' if option == 'אישה' else 'Male'

    def on_go_to_training_sessions_page_click(self):
        s.screen.switch_frame(PatientDisplaying)

    def on_click_quit(self):
        s.screen.switch_frame(ID_therapist_fill_page)


    def on_click_patient_registration(self):
        self.delete_all_labels()
        excel_file_path = "Patients.xlsx"
        workbook = openpyxl.load_workbook(excel_file_path)
        df = pd.read_excel(excel_file_path, sheet_name="patients_details")
        ID_entered = self.id_entry.get()
        is_in_ID = ID_entered in df['ID'].astype(str).values

        font = ("Helvetica", 20, "bold")

        if ID_entered == "":
            text = "לא הוזן מספר מזהה"  # No ID entered

            # Display message on canvas
            text_id = self.canvas.create_text(512, 500, text=text, font=font, fill="black", anchor="center")
            bbox = self.canvas.bbox(text_id)
            rect_id = self.canvas.create_rectangle(bbox, outline="red", width=3)

            self.labels.extend([text_id, rect_id])

        elif not ID_entered.isdigit() or ID_entered.startswith("0"):
            text = "המספר המזהה חייב להכיל רק ספרות ולא להתחיל בספרה אפס"  # ID must be digits and not start with 0

            # Display message on canvas
            text_id = self.canvas.create_text(512, 500, text=text, font=font, fill="black", anchor="center")
            bbox = self.canvas.bbox(text_id)
            rect_id = self.canvas.create_rectangle(bbox, outline="red", width=3)

            self.labels.extend([text_id, rect_id])

        elif is_in_ID:

            text = "המספר המזהה כבר שמור במערכת, נסה מספר אחר"  # ID must be digits and not start with 0

            # Display message on canvas
            text_id = self.canvas.create_text(512, 500, text=text, font=font, fill="black", anchor="center")
            bbox = self.canvas.bbox(text_id)
            rect_id = self.canvas.create_rectangle(bbox, outline="red", width=3)

            self.labels.extend([text_id, rect_id])

        else:
            s.chosen_patient_ID=ID_entered
            #modifying the columns that has other value than false
            new_row_data_details={
                'ID': str(ID_entered),
                'gender': self.gender,
                'number of exercises': 25,
                'number of repetitions in each exercise' : 10,
                'rate': "moderate",
            }

            s.average_dist = None
            sheet1=workbook["patients_details"]
            columns = list(new_row_data_details.keys())
            sheet1.append([new_row_data_details[column] for column in columns])
            #workbook.save("Patients.xlsx")

            #insert a row to the excel of history of training
            new_row_hystory_of_training = {'ID': ID_entered}
            sheet2 = workbook["patients_history_of_trainings"]
            columns = list(new_row_hystory_of_training.keys())
            sheet2.append([new_row_hystory_of_training[column] for column in columns])
            #workbook.save("Patients.xlsx")

            #insert a row to the excel of exercises
            df2 = pd.read_excel(excel_file_path, sheet_name="patients_exercises")
            sheet3 = workbook["patients_exercises"]
            new_row_data_exercises = {column: True for column in df2.columns}
            new_row_data_exercises.update({'ID': ID_entered})

            columns = list(new_row_data_exercises.keys())
            sheet3.append([new_row_data_exercises[column] for column in columns])
            workbook.save("Patients.xlsx")

            self.create_folders_when_insert_patient()

            # Success message: patient added
            font = ("Helvetica", 20, "bold")
            success_text = "!המשתמש נוסף בהצלחה"  # Hebrew: "Patient successfully added"

            text_id = self.canvas.create_text(512, 500, text=success_text, font=font, fill="black", anchor="center")
            bbox = self.canvas.bbox(text_id)
            rect_id = self.canvas.create_rectangle(bbox, outline="green", width=3)

            # Reset fields
            self.id_entry.delete(0, tk.END)
            self.selected_option.set(self.options[0])
            self.gender = 'Male'

            # Track for deletion
            self.labels.extend([text_id, rect_id])

    def create_folders_when_insert_patient(self):
        Excel.create_and_open_folder(f"Patients/{s.chosen_patient_ID}")  # open folder for patient
        Excel.create_and_open_folder(f"Patients/{s.chosen_patient_ID}/Graphs")  # open graphs folder
        Excel.create_and_open_folder(f"Patients/{s.chosen_patient_ID}/Tables")  # open graphs folder

        df = pd.read_excel("Patients.xlsx", sheet_name="patients_exercises", header=None)

        # Assuming the headers are in the first row of the DataFrame
        headers = df.iloc[0]  # Extract headers from the first row

        # Exclude a specific header
        header_to_exclude = 'ID'
        selected_headers = [header for header in headers if header != header_to_exclude]

        # Iterate over each selected header
        for header in selected_headers:
            # create a file for each execise
            Excel.create_and_open_folder(f"Patients/{s.chosen_patient_ID}/Graphs/{header}")  # open graphs folder
            Excel.create_and_open_folder(f"Patients/{s.chosen_patient_ID}/Tables/{header}")

    def delete_all_labels(self):
        for label in self.labels:
            self.canvas.delete(label)  # Proper way to remove canvas text/rectangles
        self.labels = []


class PatientDisplaying(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        excel_file_path = "Patients.xlsx"
        df = pd.read_excel(excel_file_path, sheet_name="patients_details", usecols=['ID'])

        # Rename column
        new_headers = {'ID': 'מספרים מזהים'}
        df.rename(columns=new_headers, inplace=True)

        # Shared variables (you can set them elsewhere too)
        s.chosen_patient_ID = None
        s.excel_file_path_Patient = None

        # Load background image
        image = Image.open('Pictures//Patient_list.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        self.background_label = tk.Label(self, image=self.photo_image)
        self.background_label.pack(fill="both", expand=True)

        # Treeview setup
        self.treeview = ttk.Treeview(self, style="Treeview", show="headings")
        self.treeview["columns"] = tuple(df.columns)

        # Style
        style = ttk.Style(self)
        style.configure("Treeview", font=("Thaoma", 14), rowheight=30)
        style.configure("Treeview.Heading", font=("Thaoma", 16, 'bold'), rowheight=30)

        # Column settings
        col_width = 150
        for col in df.columns:
            self.treeview.column(col, anchor="e", width=col_width)
            self.treeview.heading(col, text=col, anchor="e")

        # Insert data
        for index, row in df.iterrows():
            values = tuple(row)
            self.treeview.insert("", index, values=values, tags=(index,))

        self.treeview.tag_configure("selected", background="lightblue")
        self.treeview.bind("<ButtonRelease-1>", self.on_treeview_click)
        self.treeview.config(height=10)

        # Dynamic placement
        tree_width = len(df.columns) * col_width
        tree_height = 310
        screen_width = self.winfo_screenwidth()
        tree_x = (1024 // 2) - (tree_width // 2)
        tree_y = 180

        # Place treeview and scrollbar
        self.treeview.place(x=tree_x, y=tree_y)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.treeview.yview)
        self.treeview.configure(yscrollcommand=scrollbar.set)
        scrollbar.place(x=tree_x + tree_width + 10, y=tree_y, height=tree_height + 15)

        # Back button
        back_button_img = Image.open("Pictures//back_to_menu.jpg")
        back_button_photo = ImageTk.PhotoImage(back_button_img)
        back_button = tk.Button(self, image=back_button_photo, command=self.on_click_to_physio_menu,
                                width=back_button_img.width, height=back_button_img.height, bd=0,
                                highlightthickness=0)
        back_button.image = back_button_photo
        back_button.place(x=30, y=30)

    def on_click_to_physio_menu(self): #go back to the physio menu page
        s.screen.switch_frame(Choose_Action_Physio)

    def on_treeview_click(self, event):
        print("Clicked on a row!")
        # Get the selected item and tag (row index)
        item_id = self.treeview.identify_row(event.y)
        if item_id:
            # Update the color for the selected row
            self.treeview.tag_configure("selected", background="lightblue")

            # Deselect the previously selected row
            for item in self.treeview.get_children():
                self.treeview.tag_configure(item, background="white")

            # Get the selected row data
            selected_row = self.treeview.item(item_id, "values")
            print("Selected Row:", selected_row)

            s.chosen_patient_ID=selected_row[0]
            #s.screen.switch_frame(ChooseBallExercisesPage)
            s.screen.switch_frame(ChooseTrainingOrExerciseInformation)




def play_video(cap, label, exercise, previous, scale_factor=0.22, slow_factor=5):
    # Define the on_click function
    if previous is not None:
        def on_click(event):
            s.screen.switch_frame(TablesPage, exercise=exercise, previous=previous)
            print("video clicked!")
    else:
        def on_click(event):
            print()

    # Shared variable to control playback
    label.playing = False
    label.after_id = None  # To store the ID of the scheduled `after` callback

    def show_static_frame():
        """Display a static frame from the video when the mouse is not hovering."""
        cap.set(cv2.CAP_PROP_POS_FRAMES, 0)  # Reset to the first frame
        ret, frame = cap.read()
        if ret:
            # Detect and crop black margins
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            _, thresh = cv2.threshold(gray_frame, 10, 255, cv2.THRESH_BINARY)
            x, y, w, h = cv2.boundingRect(thresh)
            cropped_frame = frame[y:y + h, x:x + w]

            # Resize the cropped frame
            frame = cv2.resize(cropped_frame, (0, 0), fx=scale_factor + 0.02, fy=scale_factor)

            # Convert BGR to RGB
            image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

            # Convert to PhotoImage
            photo = ImageTk.PhotoImage(image=Image.fromarray(image))

            # Update the label
            label.config(image=photo)
            label.image = photo

    def update_frame():
        """Play video frames when label is hovered."""
        if label.playing:
            ret, frame = cap.read()
            if ret:
                # Detect and crop black margins
                gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                _, thresh = cv2.threshold(gray_frame, 10, 255, cv2.THRESH_BINARY)
                x, y, w, h = cv2.boundingRect(thresh)
                cropped_frame = frame[y:y + h, x:x + w]

                # Resize the cropped frame
                frame = cv2.resize(cropped_frame, (0, 0), fx=scale_factor + 0.02, fy=scale_factor)

                # Convert BGR to RGB
                image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

                # Convert to PhotoImage
                photo = ImageTk.PhotoImage(image=Image.fromarray(image))

                # Update the label
                label.config(image=photo)
                label.image = photo

                # Adjust the frame rate
                frame_rate = cap.get(cv2.CAP_PROP_FPS)
                adjusted_frame_rate = frame_rate * slow_factor
                delay = int(1000 / adjusted_frame_rate)

                # Schedule the next frame update
                label.after_id = label.after(delay, update_frame)
            else:
                # Reset to the first frame if the video ends
                cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
                update_frame()

    def on_enter(event):
        """Start video playback on hover."""
        label.playing = True
        cap.set(cv2.CAP_PROP_POS_FRAMES, 0)  # Reset video to the first frame
        if label.after_id:  # Cancel any existing scheduled updates
            label.after_cancel(label.after_id)
        update_frame()

    def on_leave(event):
        """Stop video playback and show a static frame."""
        label.playing = False
        if label.after_id:  # Cancel any scheduled updates
            label.after_cancel(label.after_id)
            label.after_id = None
        show_static_frame()

    # Bind hover and click events to the label
    label.bind("<Enter>", on_enter)
    label.bind("<Leave>", on_leave)
    label.bind("<Button-1>", on_click)

    # Show the initial static frame
    show_static_frame()



def play_video_explanation(cap, label, video_path, scale_factor_x = 0.65, scale_factor_y = 0.65, slow_factor = 1):
    """
    Plays a video frame by frame in a Tkinter label widget at the original video rate in real-time.
    Crops out black parts dynamically and plays continuously. When the video ends, it switches to a second video
    with '_2' appended to the original file name.

    Parameters:
    - cap: OpenCV video capture object.
    - label: Tkinter label widget to display the video.
    - video_path: Path to the current video file.
    - scale_factor: Scale factor for resizing the video frame.
    """
    start_time = time.time()

    def play_frame():
        nonlocal start_time, cap

        ret, frame = cap.read()

        if ret:
            # Detect and crop black margins
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            _, thresh = cv2.threshold(gray_frame, 10, 255, cv2.THRESH_BINARY)
            x, y, w, h = cv2.boundingRect(thresh)  # Get bounding box of non-black areas
            cropped_frame = frame[y:y + h, x:x + w] if w > 0 and h > 0 else frame

            # Resize the cropped frame
            frame_resized = cv2.resize(cropped_frame, (0, 0), fx=scale_factor_x, fy=scale_factor_y)

            # Convert BGR to RGB
            image = cv2.cvtColor(frame_resized, cv2.COLOR_BGR2RGB)

            # Convert to PhotoImage
            photo = ImageTk.PhotoImage(image=Image.fromarray(image))

            # Update the label with the current video frame
            label.config(image=photo)
            label.image = photo

            # Adjust timing for real-time playback
            frame_rate = cap.get(cv2.CAP_PROP_FPS)  # Get the original video frame rate
            interval = (1 / frame_rate) / slow_factor  # Interval in seconds
            elapsed_time = time.time() - start_time

            if elapsed_time < interval:
                time.sleep(interval - elapsed_time)  # Ensure the frame timing matches the video rate

            start_time = time.time()  # Reset start time for the next frame

            # Schedule the next frame
            label.after(1, play_frame)
        else:
            # End of the video, load the next video if available
            print("Video playback finished.")
            next_video_path = video_path.replace('.mp4', '_2.mp4')  # Append '_2' to the video file name
            cap.release()  # Release the current video
            try:
                cap = cv2.VideoCapture(next_video_path)  # Try loading the next video
                if not cap.isOpened():
                    print(f"Next video '{next_video_path}' not found.")
                    label.image = None
                else:
                    print(f"Playing next video: {next_video_path}")
                    play_frame()  # Start playing the next video
            except Exception as e:
                print(f"Error loading next video: {e}")
                label.image = None

    # Start playing frames
    play_frame()





def ex_in_training_or_not(data_row, exercise):
    if data_row.iloc[0, data_row.columns.get_loc(exercise)] == True:
        return True
    else:
        return False

def get_row_of_exercises_patient():
    excel_file_path = "Patients.xlsx"
    df = pd.read_excel(excel_file_path, sheet_name="patients_exercises")

    print("Current dtype:", df.iloc[:, 0].dtype)
    # Convert the first column to string for comparison
    df.iloc[:, 0] = df.iloc[:, 0].astype(str)

    # Convert the user_id to string and remove leading/trailing spaces for comparison
    user_id_cleaned = str(s.chosen_patient_ID).strip()

    # Filter rows based on the condition
    row_of_patient = df[df.iloc[:, 0] == user_id_cleaned]

    return row_of_patient

def which_image_to_put (row_of_patient, ex):

    if ex_in_training_or_not(row_of_patient, ex)==True:
        return "vi"
    else:
        return "empty"

    # A function that check which image to put as a checkbox on each checkbox click


def which_checkbox(button1, ex):
    row_of_patient = get_row_of_exercises_patient()

    # Retrieve the current checkbox value
    current_value = ex_in_training_or_not(row_of_patient, ex)

    # Toggle the value in the underlying data
    new_value_ex_patient = {ex: (not current_value)}
    Excel.find_and_change_values_exercises(new_value_ex_patient)

    # Retrieve the new current value after the update to ensure it has been toggled
    updated_value = not current_value

    # Choose the new image based on the updated value
    if updated_value:
        image1 = Image.open('Pictures//vi.png')  # True = checked (vi)
    else:
        image1 = Image.open('Pictures//empty.png')  # False = empty

    # Convert to PhotoImage
    button1_photo = ImageTk.PhotoImage(image1)

    # Update the button's image
    button1.config(image=button1_photo)

    # Store a reference to prevent garbage collection
    button1.image = button1_photo  # Keep a reference to the image


class ChooseBallExercisesPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        # Load background image
        background_image = Image.open('Pictures//background_empty.jpg')
        self.background_photo = ImageTk.PhotoImage(background_image)
        self.background_label = tk.Label(self, image=self.background_photo)
        self.background_label.pack()

        add_to_exercise_page(self, "תרגילים עם כדור")


        forward_arrow_button_img = Image.open("Pictures//forward_arrow.jpg")
        forward_arrow_button_photo = ImageTk.PhotoImage(forward_arrow_button_img)
        forward_arrow_button = tk.Button(self, image=forward_arrow_button_photo, command=lambda: self.on_arrow_click(),
                                width=forward_arrow_button_img.width, height=forward_arrow_button_img.height, bd=0,
                                highlightthickness=0)
        forward_arrow_button.image = forward_arrow_button_photo
        forward_arrow_button.place(x=50, y=480)


        to_patients_list_button_img = Image.open("Pictures//previous.jpg")
        to_patients_list_button_photo = ImageTk.PhotoImage(to_patients_list_button_img)
        to_patients_list_button = tk.Button(self, image=to_patients_list_button_photo, command=lambda: self.to_previous_button_click(),
                                   width=to_patients_list_button_img.width, height=to_patients_list_button_img.height, bd=0,
                                   highlightthickness=0)  # Set border width to 0 to remove button border
        to_patients_list_button.image = to_patients_list_button_photo  # Store reference to image to prevent garbage collection
        to_patients_list_button.place(x=30, y=30)

        row_of_patient = get_row_of_exercises_patient()

        count=1
        self.background_color = "#deeaf7"  # Set the background color to light blue

        ex_1_name="ball_bend_elbows"
        ex_2_name= "ball_raise_arms_above_head"
        ex_3_name= "ball_switch"
        ex_4_name= "ball_open_arms_and_forward"
        formatted_ex_1_name = Excel.get_name_by_exercise(ex_1_name)
        formatted_ex_2_name = Excel.get_name_by_exercise(ex_2_name)
        formatted_ex_3_name = Excel.get_name_by_exercise(ex_3_name)
        formatted_ex_4_name = Excel.get_name_by_exercise(ex_4_name)

        # Create labels for videos
        self.label1 = tk.Label(self)
        self.label1.place(x=320, y=125)  # Adjust x and y coordinates for the first video
        button1_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_1_name)}.png')
        button1_photo = ImageTk.PhotoImage(button1_image)
        button1 = tk.Button(self, image=button1_photo, command=lambda: which_checkbox(button1, ex_1_name),
                            width=button1_photo.width(), height=button1_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button1.image = button1_photo  # Store reference to image to prevent garbage collection
        button1.place(x=410, y=290)
        self.label_text1 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_1_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text1.place(x=330, y=85)
        count += 1

        self.label2 = tk.Label(self)
        self.label2.place(x=545, y=125)  # Adjust x and y coordinates for the second video
        button2_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_2_name)}.png')
        button2_photo = ImageTk.PhotoImage(button2_image)
        button2 = tk.Button(self, image=button2_photo,
                            command=lambda: which_checkbox(button2, ex_2_name),
                            width=button2_photo.width(), height=button2_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button2.image = button2_photo  # Store reference to image to prevent garbage collection
        button2.place(x=635, y=290)
        self.label_text2 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_2_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text2.place(x=555, y=85)
        count += 1

        self.label3 = tk.Label(self)
        self.label3.place(x=320, y=365)  # Adjust x and y coordinates for the third video
        button3_image = Image.open(
            f'Pictures//{which_image_to_put(row_of_patient, ex_3_name)}.png')
        button3_photo = ImageTk.PhotoImage(button3_image)
        button3 = tk.Button(self, image=button3_photo,
                            command=lambda: which_checkbox(button3, ex_3_name),
                            width=button3_photo.width(), height=button3_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button3.image = button3_photo  # Store reference to image to prevent garbage collection
        button3.place(x=410, y=530)
        self.label_text3 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_3_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text3.place(x=320, y=325)
        count += 1

        self.label4 = tk.Label(self)
        self.label4.place(x=545, y=365)  # Adjust x and y coordinates for the third video
        button4_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_4_name)}.png')
        button4_photo = ImageTk.PhotoImage(button4_image)
        button4 = tk.Button(self, image=button4_photo,
                            command=lambda: which_checkbox(button4, ex_4_name),
                            width=button4_photo.width(), height=button4_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button4.image = button4_photo  # Store reference to image to prevent garbage collection
        button4.place(x=635, y=530)
        self.label_text4 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_4_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text4.place(x=555, y=325)
        count += 1



              # Video paths
        video_file1 = f'Videos//{ex_1_name}_2.mp4'
        video_path1 = os.path.join(os.getcwd(), video_file1)
        self.cap1 = cv2.VideoCapture(video_path1)

        video_file2 = f'Videos//{ex_2_name}_2.mp4'
        video_path2 = os.path.join(os.getcwd(), video_file2)
        self.cap2 = cv2.VideoCapture(video_path2)

        video_file3 = f'Videos//{ex_3_name}_2.mp4'
        video_path3 = os.path.join(os.getcwd(), video_file3)
        self.cap3 = cv2.VideoCapture(video_path3)

        video_file4 = f'Videos//{ex_4_name}_2.mp4'
        video_path4 = os.path.join(os.getcwd(), video_file4)
        self.cap4 = cv2.VideoCapture(video_path4)



        # Check if videos are opened successfully
        if not (
                self.cap1.isOpened() and self.cap2.isOpened() and self.cap3.isOpened() and self.cap4.isOpened()):
            print("Error opening video streams or files")

        else:
            # Play videos
            play_video(self.cap1, self.label1, ex_1_name, "ball")
            play_video(self.cap2, self.label2, ex_2_name, "ball")
            play_video(self.cap3, self.label3, ex_3_name, "ball")
            play_video(self.cap4, self.label4, ex_4_name, "ball")


    def on_arrow_click(self):
        Excel.find_and_change_values_patients({"number of repetitions in each exercise": self.selected_option_rep.get(),
                                               "rate": self.selected_option_rate.get()})
        s.screen.switch_frame(ChooseRubberBandExercisesPage)


    def to_previous_button_click(self):
        num_of_exercises_in_training=Excel.count_number_of_exercises_in_training_by_ID()
        if num_of_exercises_in_training<5:
            back = Image.open('Pictures//not_enough_exercises_chosen.jpg')
            background_img = ImageTk.PhotoImage(back)

            self.label = tk.Label(self, image=background_img, compound=tk.CENTER, highlightthickness=0)
            self.label.place(x=170, y=20)
            self.label.image = background_img

        else:
            Excel.find_and_change_values_patients({"number of exercises": num_of_exercises_in_training, "number of repetitions in each exercise": self.selected_option_rep.get(), "rate": self.selected_option_rate.get()})
            s.screen.switch_frame(ChooseTrainingOrExerciseInformation)




class ChooseRubberBandExercisesPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)

        # Load background image
        background_image = Image.open('Pictures//background_empty.jpg')
        self.background_photo = ImageTk.PhotoImage(background_image)
        self.background_label = tk.Label(self, image=self.background_photo)
        self.background_label.pack()

        add_to_exercise_page(self, "תרגילים עם גומייה")


        forward_arrow_button_img = Image.open("Pictures//forward_arrow.jpg")
        forward_arrow_button_photo = ImageTk.PhotoImage(forward_arrow_button_img)
        forward_arrow_button = tk.Button(self, image=forward_arrow_button_photo, command=lambda: self.on_arrow_click_forward(),
                                   width=forward_arrow_button_img.width, height=forward_arrow_button_img.height, bd=0,
                                   highlightthickness=0)
        forward_arrow_button.image = forward_arrow_button_photo
        forward_arrow_button.place(x=50, y=480)

        backward_arrow_button_img = Image.open("Pictures//previous_arrow.jpg")
        backward_arrow_button_photo = ImageTk.PhotoImage(backward_arrow_button_img)
        backward_arrow_button = tk.Button(self, image=backward_arrow_button_photo, command=lambda: self.on_arrow_click_back(),
                                   width=backward_arrow_button_img.width, height=backward_arrow_button_img.height, bd=0,
                                   highlightthickness=0)
        backward_arrow_button.image = backward_arrow_button_photo
        backward_arrow_button.place(x=913, y=480)

        to_patients_list_button_img = Image.open("Pictures//previous.jpg")
        to_patients_list_button_photo = ImageTk.PhotoImage(to_patients_list_button_img)
        to_patients_list_button = tk.Button(self, image=to_patients_list_button_photo,
                                   command=lambda: self.to_previous_button_click(),
                                   width=to_patients_list_button_img.width, height=to_patients_list_button_img.height,
                                   bd=0,
                                   highlightthickness=0)  # Set border width to 0 to remove button border
        to_patients_list_button.image = to_patients_list_button_photo  # Store reference to image to prevent garbage collection
        to_patients_list_button.place(x=30, y=30)

        row_of_patient = get_row_of_exercises_patient()

        count= 1+s.ball_exercises_number
        self.background_color = "#deeaf7"  # Set the background color to light blue

        ex_1_name = "band_open_arms"
        ex_2_name = "band_open_arms_and_up"
        ex_3_name = "band_up_and_lean"
        ex_4_name = "band_straighten_left_arm_elbows_bend_to_sides"
        ex_5_name = "band_straighten_right_arm_elbows_bend_to_sides"
        formatted_ex_1_name = Excel.get_name_by_exercise(ex_1_name)
        formatted_ex_2_name = Excel.get_name_by_exercise(ex_2_name)
        formatted_ex_3_name = Excel.get_name_by_exercise(ex_3_name)
        formatted_ex_4_name = Excel.get_name_by_exercise(ex_4_name)
        formatted_ex_5_name = Excel.get_name_by_exercise(ex_5_name)


        # Create labels for videos
        self.label1 = tk.Label(self)
        self.label1.place(x=220, y=125)  # Adjust x and y coordinates for the first video
        button1_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_1_name)}.png')
        button1_photo = ImageTk.PhotoImage(button1_image)
        button1 = tk.Button(self, image=button1_photo, command=lambda: which_checkbox(button1, ex_1_name),
                            width=button1_photo.width(), height=button1_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button1.image = button1_photo  # Store reference to image to prevent garbage collection
        button1.place(x=310, y=290)
        self.label_text1 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_1_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text1.place(x=230, y=85)
        count += 1

        self.label2 = tk.Label(self)
        self.label2.place(x=445, y=125)  # Adjust x and y coordinates for the second video
        button2_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_2_name)}.png')
        button2_photo = ImageTk.PhotoImage(button2_image)
        button2 = tk.Button(self, image=button2_photo,
                            command=lambda: which_checkbox(button2, ex_2_name),
                            width=button2_photo.width(), height=button2_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button2.image = button2_photo  # Store reference to image to prevent garbage collection
        button2.place(x=535, y=290)
        self.label_text2 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_2_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text2.place(x=455, y=85)
        count += 1

        self.label3 = tk.Label(self)
        self.label3.place(x=670, y=125)  # Adjust x and y coordinates for the third video
        button3_image = Image.open(
            f'Pictures//{which_image_to_put(row_of_patient, ex_3_name)}.png')
        button3_photo = ImageTk.PhotoImage(button3_image)
        button3 = tk.Button(self, image=button3_photo,
                            command=lambda: which_checkbox(button3, ex_3_name),
                            width=button3_photo.width(), height=button3_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button3.image = button3_photo  # Store reference to image to prevent garbage collection
        button3.place(x=760, y=290)
        self.label_text3 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_3_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text3.place(x=680, y=85)
        count += 1

        self.label4 = tk.Label(self)
        self.label4.place(x=330, y=365)  # Adjust x and y coordinates for the third video
        button4_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_4_name)}.png')
        button4_photo = ImageTk.PhotoImage(button4_image)
        button4 = tk.Button(self, image=button4_photo,
                            command=lambda: which_checkbox(button4, ex_4_name),
                            width=button4_photo.width(), height=button4_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button4.image = button4_photo  # Store reference to image to prevent garbage collection
        button4.place(x=410, y=530)
        self.label_text4 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_4_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text4.place(x=340, y=325)
        count += 1

        self.label5 = tk.Label(self)
        self.label5.place(x=555, y=365)  # Adjust x and y coordinates for the third video
        button5_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_5_name)}.png')
        button5_photo = ImageTk.PhotoImage(button5_image)
        button5 = tk.Button(self, image=button5_photo,
                            command=lambda: which_checkbox(button5, ex_5_name),
                            width=button5_photo.width(), height=button5_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button5.image = button5_photo  # Store reference to image to prevent garbage collection
        button5.place(x=645, y=530)
        self.label_text5 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_5_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text5.place(x=565, y=325)
        count += 1

        # Video paths
        video_file1 = f'Videos//{ex_1_name}_2.mp4'
        video_path1 = os.path.join(os.getcwd(), video_file1)
        self.cap1 = cv2.VideoCapture(video_path1)

        video_file2 = f'Videos//{ex_2_name}_2.mp4'
        video_path2 = os.path.join(os.getcwd(), video_file2)
        self.cap2 = cv2.VideoCapture(video_path2)

        video_file3 = f'Videos//{ex_3_name}_2.mp4'
        video_path3 = os.path.join(os.getcwd(), video_file3)
        self.cap3 = cv2.VideoCapture(video_path3)

        video_file4 = f'Videos//{ex_4_name}_2.mp4'
        video_path4 = os.path.join(os.getcwd(), video_file4)
        self.cap4 = cv2.VideoCapture(video_path4)

        video_file5 = f'Videos//{ex_5_name}_2.mp4'
        video_path5 = os.path.join(os.getcwd(), video_file5)
        self.cap5 = cv2.VideoCapture(video_path5)

        # Check if videos are opened successfully
        if not (
                self.cap1.isOpened() and self.cap2.isOpened() and self.cap3.isOpened() and self.cap4.isOpened()  and self.cap5.isOpened()):
            print("Error opening video streams or files")

        else:
            # Play videos
            play_video(self.cap1, self.label1, ex_1_name, "band")
            play_video(self.cap2, self.label2, ex_2_name, "band")
            play_video(self.cap3, self.label3, ex_3_name, "band")
            play_video(self.cap4, self.label4, ex_4_name, "band")
            play_video(self.cap5, self.label5, ex_5_name, "band")

    def on_arrow_click_forward(self):
        Excel.find_and_change_values_patients({"number of repetitions in each exercise": self.selected_option_rep.get(),
                                               "rate": self.selected_option_rate.get()})
        s.screen.switch_frame(ChooseStickExercisesPage)

    def on_arrow_click_back(self):
        Excel.find_and_change_values_patients({"number of repetitions in each exercise": self.selected_option_rep.get(),
                                               "rate": self.selected_option_rate.get()})
        s.screen.switch_frame(ChooseBallExercisesPage)

    def to_previous_button_click(self):
        num_of_exercises_in_training = Excel.count_number_of_exercises_in_training_by_ID()
        if num_of_exercises_in_training < 5:
            back = Image.open('Pictures//not_enough_exercises_chosen.jpg')
            background_img = ImageTk.PhotoImage(back)

            self.label = tk.Label(self, image=background_img, compound=tk.CENTER, highlightthickness=0)
            self.label.place(x=170, y=20)
            self.label.image = background_img

        else:
            Excel.find_and_change_values_patients({"number of exercises": num_of_exercises_in_training, "number of repetitions in each exercise": self.selected_option_rep.get(), "rate": self.selected_option_rate.get()})
            s.screen.switch_frame(ChooseTrainingOrExerciseInformation)



class ChooseStickExercisesPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)

        # Load background image
        background_image = Image.open('Pictures//background_empty.jpg')
        self.background_photo = ImageTk.PhotoImage(background_image)
        self.background_label = tk.Label(self, image=self.background_photo)
        self.background_label.pack()

        add_to_exercise_page(self, "תרגילים עם מקל/בר")


        forward_arrow_button_img = Image.open("Pictures//forward_arrow.jpg")
        forward_arrow_button_photo = ImageTk.PhotoImage(forward_arrow_button_img)
        forward_arrow_button = tk.Button(self, image=forward_arrow_button_photo,
                                         command=lambda: self.on_arrow_click_forward(),
                                         width=forward_arrow_button_img.width, height=forward_arrow_button_img.height,
                                         bd=0,
                                         highlightthickness=0)
        forward_arrow_button.image = forward_arrow_button_photo
        forward_arrow_button.place(x=50, y=480)

        backward_arrow_button_img = Image.open("Pictures//previous_arrow.jpg")
        backward_arrow_button_photo = ImageTk.PhotoImage(backward_arrow_button_img)
        backward_arrow_button = tk.Button(self, image=backward_arrow_button_photo,
                                          command=lambda: self.on_arrow_click_back(),
                                          width=backward_arrow_button_img.width,
                                          height=backward_arrow_button_img.height, bd=0,
                                          highlightthickness=0)
        backward_arrow_button.image = backward_arrow_button_photo
        backward_arrow_button.place(x=913, y=480)

        to_patients_list_button_img = Image.open("Pictures//previous.jpg")
        to_patients_list_button_photo = ImageTk.PhotoImage(to_patients_list_button_img)
        to_patients_list_button = tk.Button(self, image=to_patients_list_button_photo,
                                            command=lambda: self.to_previous_button_click(),
                                            width=to_patients_list_button_img.width,
                                            height=to_patients_list_button_img.height,
                                            bd=0,
                                            highlightthickness=0)  # Set border width to 0 to remove button border
        to_patients_list_button.image = to_patients_list_button_photo  # Store reference to image to prevent garbage collection
        to_patients_list_button.place(x=30, y=30)

        row_of_patient=get_row_of_exercises_patient()

        count= 1 + s.ball_exercises_number + s.band_exercises_number
        self.background_color = "#deeaf7"  # Set the background color to light blue

        ex_1_name = "stick_bend_elbows"
        ex_2_name = "stick_bend_elbows_and_up"
        ex_3_name = "stick_raise_arms_above_head"
        ex_4_name =  "stick_switch"
        ex_5_name =  "stick_up_and_lean"
        formatted_ex_1_name = Excel.get_name_by_exercise(ex_1_name)
        formatted_ex_2_name = Excel.get_name_by_exercise(ex_2_name)
        formatted_ex_3_name = Excel.get_name_by_exercise(ex_3_name)
        formatted_ex_4_name = Excel.get_name_by_exercise(ex_4_name)
        formatted_ex_5_name = Excel.get_name_by_exercise(ex_5_name)


        # Create labels for videos
        self.label1 = tk.Label(self)
        self.label1.place(x=220, y=125)  # Adjust x and y coordinates for the first video
        button1_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_1_name)}.png')
        button1_photo = ImageTk.PhotoImage(button1_image)
        button1 = tk.Button(self, image=button1_photo, command=lambda: which_checkbox(button1, ex_1_name),
                            width=button1_photo.width(), height=button1_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button1.image = button1_photo  # Store reference to image to prevent garbage collection
        button1.place(x=310, y=290)
        self.label_text1 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_1_name}', font=("Thaoma", 9, 'bold'), bg=self.background_color,
            justify='center', width=25, wraplength=170, anchor='center')
        self.label_text1.place(x=230, y=85)
        count += 1


        self.label2 = tk.Label(self)
        self.label2.place(x=445, y=125)  # Adjust x and y coordinates for the second video
        button2_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_2_name)}.png')
        button2_photo = ImageTk.PhotoImage(button2_image)
        button2 = tk.Button(self, image=button2_photo,
                            command=lambda: which_checkbox(button2, ex_2_name),
                            width=button2_photo.width(), height=button2_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button2.image = button2_photo  # Store reference to image to prevent garbage collection
        button2.place(x=535, y=290)
        self.label_text2 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_2_name}', font=("Thaoma", 9, 'bold'), bg=self.background_color,
            justify='center', width=25, wraplength=170, anchor='center')
        self.label_text2.place(x=455, y=85)
        count += 1


        self.label3 = tk.Label(self)
        self.label3.place(x=670, y=125)  # Adjust x and y coordinates for the third video
        button3_image = Image.open(
            f'Pictures//{which_image_to_put(row_of_patient, ex_3_name)}.png')
        button3_photo = ImageTk.PhotoImage(button3_image)
        button3 = tk.Button(self, image=button3_photo,
                            command=lambda: which_checkbox(button3, ex_3_name),
                            width=button3_photo.width(), height=button3_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button3.image = button3_photo  # Store reference to image to prevent garbage collection
        button3.place(x=760, y=290)
        self.label_text3 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_3_name}', font=("Thaoma", 9, 'bold'), bg=self.background_color,
            justify='center', width=25, wraplength=170, anchor='center')
        self.label_text3.place(x=680, y=85)
        count += 1


        self.label4 = tk.Label(self)
        self.label4.place(x=330, y=365)  # Adjust x and y coordinates for the third video
        button4_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_4_name)}.png')
        button4_photo = ImageTk.PhotoImage(button4_image)
        button4 = tk.Button(self, image=button4_photo,
                            command=lambda: which_checkbox(button4, ex_4_name),
                            width=button4_photo.width(), height=button4_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button4.image = button4_photo  # Store reference to image to prevent garbage collection
        button4.place(x=410, y=530)
        self.label_text4 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_4_name}', font=("Thaoma", 9, 'bold'), bg=self.background_color,
            justify='center', width=25, wraplength=170, anchor='center')
        self.label_text4.place(x=340, y=325)
        count += 1

        self.label5 = tk.Label(self)
        self.label5.place(x=555, y=365)  # Adjust x and y coordinates for the third video
        button5_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_5_name)}.png')
        button5_photo = ImageTk.PhotoImage(button5_image)
        button5 = tk.Button(self, image=button5_photo,
                            command=lambda: which_checkbox(button5, ex_5_name),
                            width=button5_photo.width(), height=button5_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button5.image = button5_photo  # Store reference to image to prevent garbage collection
        button5.place(x=645, y=530)
        self.label_text5 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_5_name}', font=("Thaoma", 9, 'bold'), bg=self.background_color,
            justify='center', width=25, wraplength=170, anchor='center')
        self.label_text5.place(x=565, y=325)
        count += 1


        # Video paths
        video_file1 = f'Videos//{ex_1_name}_2.mp4'
        video_path1 = os.path.join(os.getcwd(), video_file1)
        self.cap1 = cv2.VideoCapture(video_path1)

        video_file2 = f'Videos//{ex_2_name}_2.mp4'
        video_path2 = os.path.join(os.getcwd(), video_file2)
        self.cap2 = cv2.VideoCapture(video_path2)

        video_file3 = f'Videos//{ex_3_name}_2.mp4'
        video_path3 = os.path.join(os.getcwd(), video_file3)
        self.cap3 = cv2.VideoCapture(video_path3)

        video_file4 = f'Videos//{ex_4_name}_2.mp4'
        video_path4 = os.path.join(os.getcwd(), video_file4)
        self.cap4 = cv2.VideoCapture(video_path4)

        video_file5 = f'Videos//{ex_5_name}_2.mp4'
        video_path5 = os.path.join(os.getcwd(), video_file5)
        self.cap5 = cv2.VideoCapture(video_path5)


        # Check if videos are opened successfully
        if not (
                self.cap1.isOpened() and self.cap2.isOpened() and self.cap3.isOpened() and self.cap4.isOpened()  and self.cap5.isOpened()):
            print("Error opening video streams or files")

        else:
            # Play videos
            play_video(self.cap1, self.label1, ex_1_name, "stick")
            play_video(self.cap2, self.label2,ex_2_name, "stick")
            play_video(self.cap3, self.label3,ex_3_name, "stick")
            play_video(self.cap4, self.label4,ex_4_name, "stick")
            play_video(self.cap5, self.label5,ex_5_name, "stick")


    def on_arrow_click_forward(self):
        Excel.find_and_change_values_patients({"number of repetitions in each exercise": self.selected_option_rep.get(),
                                               "rate": self.selected_option_rate.get()})
        s.screen.switch_frame(ChooseWeightsExercisesPage)

    def on_arrow_click_back(self):
        Excel.find_and_change_values_patients({"number of repetitions in each exercise": self.selected_option_rep.get(),
                                               "rate": self.selected_option_rate.get()})
        s.screen.switch_frame(ChooseRubberBandExercisesPage)

    def to_previous_button_click(self):
        num_of_exercises_in_training = Excel.count_number_of_exercises_in_training_by_ID()
        if num_of_exercises_in_training < 5:
            back = Image.open('Pictures//not_enough_exercises_chosen.jpg')
            background_img = ImageTk.PhotoImage(back)

            self.label = tk.Label(self, image=background_img, compound=tk.CENTER, highlightthickness=0)
            self.label.place(x=170, y= 20)
            self.label.image = background_img

        else:
            Excel.find_and_change_values_patients({"number of exercises": num_of_exercises_in_training, "number of repetitions in each exercise": self.selected_option_rep.get(), "rate": self.selected_option_rate.get()})
            s.screen.switch_frame(ChooseTrainingOrExerciseInformation)




class ChooseWeightsExercisesPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)

        # Load background image
        background_image = Image.open('Pictures//background_empty.jpg')
        self.background_photo = ImageTk.PhotoImage(background_image)
        self.background_label = tk.Label(self, image=self.background_photo)
        self.background_label.pack()

        add_to_exercise_page(self, "תרגילים עם משקולות")


        forward_arrow_button_img = Image.open("Pictures//forward_arrow.jpg")
        forward_arrow_button_photo = ImageTk.PhotoImage(forward_arrow_button_img)
        forward_arrow_button = tk.Button(self, image=forward_arrow_button_photo, command=lambda: self.on_arrow_click_forward(),
                                   width=forward_arrow_button_img.width, height=forward_arrow_button_img.height, bd=0,
                                   highlightthickness=0)
        forward_arrow_button.image = forward_arrow_button_photo
        forward_arrow_button.place(x=50, y=480)

        backward_arrow_button_img = Image.open("Pictures//previous_arrow.jpg")
        backward_arrow_button_photo = ImageTk.PhotoImage(backward_arrow_button_img)
        backward_arrow_button = tk.Button(self, image=backward_arrow_button_photo, command=lambda: self.on_arrow_click_back(),
                                   width=backward_arrow_button_img.width, height=backward_arrow_button_img.height, bd=0,
                                   highlightthickness=0)
        backward_arrow_button.image = backward_arrow_button_photo
        backward_arrow_button.place(x=913, y=480)

        to_patients_list_button_img = Image.open("Pictures//previous.jpg")
        to_patients_list_button_photo = ImageTk.PhotoImage(to_patients_list_button_img)
        to_patients_list_button = tk.Button(self, image=to_patients_list_button_photo,
                                   command=lambda: self.to_previous_button_click(),
                                   width=to_patients_list_button_img.width, height=to_patients_list_button_img.height,
                                   bd=0,
                                   highlightthickness=0)  # Set border width to 0 to remove button border
        to_patients_list_button.image = to_patients_list_button_photo  # Store reference to image to prevent garbage collection
        to_patients_list_button.place(x=30, y=30)

        row_of_patient = get_row_of_exercises_patient()
        count= 1 + s.ball_exercises_number + s.band_exercises_number + s.stick_exercises_number
        self.background_color = "#deeaf7"  # Set the background color to light blue


        ex_1_name = "weights_open_arms_and_forward"
        ex_2_name = "weights_abduction"
        formatted_ex_1_name = Excel.get_name_by_exercise(ex_1_name)
        formatted_ex_2_name = Excel.get_name_by_exercise(ex_2_name)



        self.label1 = tk.Label(self)
        self.label1.place(x=320, y=220)  # Adjust x and y coordinates for the third video
        button1_image = Image.open(
            f'Pictures//{which_image_to_put(row_of_patient, ex_1_name)}.png')
        button1_photo = ImageTk.PhotoImage(button1_image)
        button1 = tk.Button(self, image=button1_photo,
                            command=lambda: which_checkbox(button1, ex_1_name),
                            width=button1_photo.width(), height=button1_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button1.image = button1_photo  # Store reference to image to prevent garbage collection
        button1.place(x=410, y=385)
        self.label_text1 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_1_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=30, wraplength=200, anchor='center')
        self.label_text1.place(x=310, y=180)
        count += 1

        self.label2 = tk.Label(self)
        self.label2.place(x=545, y=220)  # Adjust x and y coordinates for the third video
        button2_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_2_name)}.png')
        button2_photo = ImageTk.PhotoImage(button2_image)
        button2 = tk.Button(self, image=button2_photo,
                            command=lambda: which_checkbox(button2, ex_2_name),
                            width=button2_photo.width(), height=button2_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button2.image = button2_photo  # Store reference to image to prevent garbage collection
        button2.place(x=635, y=385)
        self.label_text2 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_2_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text2.place(x=555, y=180)
        count += 1



        video_file1 = f'Videos//{ex_1_name}_2.mp4'
        video_path1 = os.path.join(os.getcwd(), video_file1)
        self.cap1 = cv2.VideoCapture(video_path1)

        video_file2 = f'Videos//{ex_2_name}_2.mp4'
        video_path2 = os.path.join(os.getcwd(), video_file2)
        self.cap2 = cv2.VideoCapture(video_path2)


        # Check if videos are opened successfully
        if not (
                self.cap1.isOpened() and self.cap2.isOpened()):
            print("Error opening video streams or files")

        else:
            # Play videos
            # play_video(self.cap1, self.label1, ex_1_name, "weights")
            # play_video(self.cap2, self.label2, ex_2_name, "weights")
            play_video(self.cap1, self.label1, ex_1_name, "weights")
            play_video(self.cap2, self.label2, ex_2_name, "weights")

    def on_arrow_click_forward(self):
        Excel.find_and_change_values_patients({"number of repetitions in each exercise": self.selected_option_rep.get(),
                                               "rate": self.selected_option_rate.get()})
        s.screen.switch_frame(ChooseNoToolExercisesPage)

    def on_arrow_click_back(self):
        Excel.find_and_change_values_patients({"number of repetitions in each exercise": self.selected_option_rep.get(),
                                               "rate": self.selected_option_rate.get()})
        s.screen.switch_frame(ChooseStickExercisesPage)

    def to_previous_button_click(self):
        num_of_exercises_in_training = Excel.count_number_of_exercises_in_training_by_ID()
        if num_of_exercises_in_training < 5:
            back = Image.open('Pictures//not_enough_exercises_chosen.jpg')
            background_img = ImageTk.PhotoImage(back)

            self.label = tk.Label(self, image=background_img, compound=tk.CENTER, highlightthickness=0)
            self.label.place(x=170, y=20)
            self.label.image = background_img

        else:
            Excel.find_and_change_values_patients({"number of exercises": num_of_exercises_in_training, "number of repetitions in each exercise": self.selected_option_rep.get(), "rate": self.selected_option_rate.get()})
            s.screen.switch_frame(ChooseTrainingOrExerciseInformation)


class ChooseNoToolExercisesPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)

        # Load background image
        background_image = Image.open('Pictures//background_empty.jpg')
        self.background_photo = ImageTk.PhotoImage(background_image)
        self.background_label = tk.Label(self, image=self.background_photo)
        self.background_label.pack()

        add_to_exercise_page(self, "תרגילים ללא אביזר")

        end_button_img = Image.open("Pictures//end_button.jpg")
        end_button_photo = ImageTk.PhotoImage(end_button_img)
        end_button = tk.Button(self, image=end_button_photo,
                                         command=lambda: self.on_end_click(),
                                         width=end_button_img.width, height=end_button_img.height,
                                         bd=0,
                                         highlightthickness=0)
        end_button.image = end_button_photo
        end_button.place(x=50, y=490)

        backward_arrow_button_img = Image.open("Pictures//previous_arrow.jpg")
        backward_arrow_button_photo = ImageTk.PhotoImage(backward_arrow_button_img)
        backward_arrow_button = tk.Button(self, image=backward_arrow_button_photo,
                                          command=lambda: self.on_arrow_click_back(),
                                          width=backward_arrow_button_img.width,
                                          height=backward_arrow_button_img.height, bd=0,
                                          highlightthickness=0)
        backward_arrow_button.image = backward_arrow_button_photo
        backward_arrow_button.place(x=913, y=480)

        row_of_patient=get_row_of_exercises_patient()

        count= 1 + s.ball_exercises_number + s.band_exercises_number + s.stick_exercises_number + s.weights_exercises_number
        self.background_color = "#deeaf7"  # Set the background color to light blue

        ex_1_name = "notool_hands_behind_and_lean"
        ex_2_name = "notool_right_hand_up_and_bend"
        ex_3_name = "notool_left_hand_up_and_bend"
        ex_4_name = "notool_raising_hands_diagonally"
        ex_5_name = "notool_right_bend_left_up_from_side"
        ex_6_name = "notool_left_bend_right_up_from_side"
        formatted_ex_1_name = Excel.get_name_by_exercise(ex_1_name)
        formatted_ex_2_name = Excel.get_name_by_exercise(ex_2_name)
        formatted_ex_3_name = Excel.get_name_by_exercise(ex_3_name)
        formatted_ex_4_name = Excel.get_name_by_exercise(ex_4_name)
        formatted_ex_5_name = Excel.get_name_by_exercise(ex_5_name)
        formatted_ex_6_name = Excel.get_name_by_exercise(ex_6_name)

        # Create labels for videos
        self.label1 = tk.Label(self)
        self.label1.place(x=220, y=125)  # Adjust x and y coordinates for the first video
        button1_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_1_name)}.png')
        button1_photo = ImageTk.PhotoImage(button1_image)
        button1 = tk.Button(self, image=button1_photo, command=lambda: which_checkbox(button1, ex_1_name),
                            width=button1_photo.width(), height=button1_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button1.image = button1_photo  # Store reference to image to prevent garbage collection
        button1.place(x=310, y=290)
        self.label_text1 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_1_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text1.place(x=230, y=85)
        count += 1

        self.label2 = tk.Label(self)
        self.label2.place(x=445, y=125)  # Adjust x and y coordinates for the second video
        button2_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_2_name)}.png')
        button2_photo = ImageTk.PhotoImage(button2_image)
        button2 = tk.Button(self, image=button2_photo,
                            command=lambda: which_checkbox(button2, ex_2_name),
                            width=button2_photo.width(), height=button2_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button2.image = button2_photo  # Store reference to image to prevent garbage collection
        button2.place(x=535, y=290)
        self.label_text2 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_2_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text2.place(x=455, y=85)
        count += 1

        self.label3 = tk.Label(self)
        self.label3.place(x=670, y=125)  # Adjust x and y coordinates for the third video
        button3_image = Image.open(
            f'Pictures//{which_image_to_put(row_of_patient, ex_3_name)}.png')
        button3_photo = ImageTk.PhotoImage(button3_image)
        button3 = tk.Button(self, image=button3_photo,
                            command=lambda: which_checkbox(button3, ex_3_name),
                            width=button3_photo.width(), height=button3_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button3.image = button3_photo  # Store reference to image to prevent garbage collection
        button3.place(x=760, y=290)
        self.label_text3 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_3_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text3.place(x=680, y=85)
        count += 1

        self.label4 = tk.Label(self)
        self.label4.place(x=220, y=365)  # Adjust x and y coordinates for the third video
        button4_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_4_name)}.png')
        button4_photo = ImageTk.PhotoImage(button4_image)
        button4 = tk.Button(self, image=button4_photo,
                            command=lambda: which_checkbox(button4, ex_4_name),
                            width=button4_photo.width(), height=button4_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button4.image = button4_photo  # Store reference to image to prevent garbage collection
        button4.place(x=310, y=530)
        self.label_text4 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_4_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=170, anchor='center')
        self.label_text4.place(x=230, y=325)
        count += 1

        self.label5 = tk.Label(self)
        self.label5.place(x=445, y=365)  # Adjust x and y coordinates for the third video
        button5_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_5_name)}.png')
        button5_photo = ImageTk.PhotoImage(button5_image)
        button5 = tk.Button(self, image=button5_photo,
                            command=lambda: which_checkbox(button5, ex_5_name),
                            width=button5_photo.width(), height=button5_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button5.image = button5_photo  # Store reference to image to prevent garbage collection
        button5.place(x=535, y=530)
        self.label_text5 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_5_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=25, wraplength=200, anchor='center')
        self.label_text5.place(x=447, y=325)
        count += 1

        self.label6 = tk.Label(self)
        self.label6.place(x=670, y=365)  # Adjust x and y coordinates for the third video
        button6_image = Image.open(f'Pictures//{which_image_to_put(row_of_patient, ex_6_name)}.png')
        button6_photo = ImageTk.PhotoImage(button6_image)
        button6 = tk.Button(self, image=button6_photo,
                            command=lambda: which_checkbox(button6, ex_6_name),
                            width=button6_photo.width(), height=button6_photo.height(), bd=0,
                            highlightthickness=0)  # Set border width to 0 to remove button border
        button6.image = button6_photo  # Store reference to image to prevent garbage collection
        button6.place(x=760, y=530)
        self.label_text6 = tk.Label(self, text=f'Exercise {count}\n{formatted_ex_6_name}', font=("Thaoma", 9, 'bold'),
                                    bg=self.background_color,
                                    justify='center', width=27, wraplength=200, anchor='center')
        self.label_text6.place(x=663, y=325)
        count += 1

        # Video paths
        video_file1 = f'Videos//{ex_1_name}_2.mp4'
        video_path1 = os.path.join(os.getcwd(), video_file1)
        self.cap1 = cv2.VideoCapture(video_path1)

        video_file2 = f'Videos//{ex_2_name}_2.mp4'
        video_path2 = os.path.join(os.getcwd(), video_file2)
        self.cap2 = cv2.VideoCapture(video_path2)

        video_file3 = f'Videos//{ex_3_name}_2.mp4'
        video_path3 = os.path.join(os.getcwd(), video_file3)
        self.cap3 = cv2.VideoCapture(video_path3)

        video_file4 = f'Videos//{ex_4_name}_2.mp4'
        video_path4 = os.path.join(os.getcwd(), video_file4)
        self.cap4 = cv2.VideoCapture(video_path4)

        video_file5 = f'Videos//{ex_5_name}_2.mp4'
        video_path5 = os.path.join(os.getcwd(), video_file5)
        self.cap5 = cv2.VideoCapture(video_path5)

        video_file6 = f'Videos//{ex_6_name}_2.mp4'
        video_path6 = os.path.join(os.getcwd(), video_file6)
        self.cap6 = cv2.VideoCapture(video_path6)

        # Check if videos are opened successfully
        if not (
                self.cap1.isOpened() and self.cap2.isOpened() and self.cap3.isOpened() and self.cap4.isOpened() and self.cap5.isOpened() and self.cap6.isOpened()):
            print("Error opening video streams or files")

        else:
            # Play videos
            play_video(self.cap1, self.label1, ex_1_name, "notool")
            play_video(self.cap2, self.label2, ex_2_name, "notool")
            play_video(self.cap3, self.label3, ex_3_name, "notool")
            play_video(self.cap4, self.label4, ex_4_name, "notool")
            play_video(self.cap5, self.label5, ex_5_name, "notool")
            play_video(self.cap6, self.label6, ex_6_name, "notool")


    def on_end_click(self):
        num_of_exercises_in_training = Excel.count_number_of_exercises_in_training_by_ID()
        if num_of_exercises_in_training < 5:
            back = Image.open('Pictures//not_enough_exercises_chosen.jpg')
            background_img = ImageTk.PhotoImage(back)

            self.label = tk.Label(self, image=background_img, compound=tk.CENTER, highlightthickness=0)
            self.label.place(x=190, y=480)
            self.label.image = background_img

        else:
            Excel.find_and_change_values_patients({"number of exercises": num_of_exercises_in_training, "number of repetitions in each exercise": self.selected_option_rep.get(), "rate": self.selected_option_rate.get()})
            s.screen.switch_frame(ChooseTrainingOrExerciseInformation)

    def on_arrow_click_back(self):
        Excel.find_and_change_values_patients({"number of repetitions in each exercise": self.selected_option_rep.get(),
                                               "rate": self.selected_option_rate.get()})
        s.screen.switch_frame(ChooseWeightsExercisesPage)





def get_sorted_folders(directory):
    # Get the list of folder names in the specified directory
    folder_names = [f for f in os.listdir(directory) if os.path.isdir(os.path.join(directory, f))]

    # Convert folder names to datetime objects
    folder_datetimes = []
    for folder in folder_names:
        try:
            folder_datetime = datetime.strptime(folder, "%d-%m-%Y %H-%M-%S")
            folder_datetimes.append((folder, folder_datetime))
        except ValueError:
            # Skip folder names that don't match the expected format
            continue

    # Sort the list of tuples by datetime in descending order
    sorted_folders = sorted(folder_datetimes, key=lambda x: x[1], reverse=True)

    # Extract just the folder names from the sorted list
    sorted_folder_names = [folder for folder, _ in sorted_folders]

    return sorted_folder_names


def convert_white_to_transparent(image_path, tolerance=100):
    img = Image.open(image_path)
    img = img.convert("RGBA")
    datas = img.getdata()

    newData = []
    for item in datas:
        # Check if the pixel is near-white by allowing a tolerance range
        if item[0] >= 255 - tolerance and item[1] >= 255 - tolerance and item[2] >= 255 - tolerance:
            # Replace near-white with transparent
            newData.append((255, 255, 255, 0))  # Make it fully transparent
        else:
            newData.append(item)

    img.putdata(newData)
    return img


#Page that appears when there is an exercise
class ExercisePage(tk.Frame):
    def __init__(self, master, exercise_type, reverse_color, reverse_bar, min_distance, max_distance, which_side= None, min_distance_x = None, max_distance_x =None):
        super().__init__(master)
        self.master = master
        self.start_of_750_ms_bar_len = None
        self.start_of_time_count_all_rules_not_limit = None
        self.start_of_time_count_hands_not_good = None
        self.previous_all_rules_ok = False
        self.time_of_exercise_start = time.time()
        self.end_of_comment_recording = None
        self.comments_audio = {}
        self.last_loop_time = time.time()


        # The subjects from which the algorithm can choose
        subjects = ["garden", "galaxy", "butterflies"]
        random.shuffle(subjects)
        self.chosen_subject = subjects[0]  # Instance variable for later use
        self.percent_of_bar = 0.3
        self.comment = None
        self.can_comment = False

        # Create a canvas for layering background and transparent image
        self.canvas = tk.Canvas(self, width=1024, height=576)
        self.canvas.pack(fill="both", expand=True)

        # Load background image
        self.background_image = Image.open(f'Pictures/{self.chosen_subject}/background.jpg')
        self.background_photo = ImageTk.PhotoImage(self.background_image)

        self.stop_training_button_img = Image.open("Pictures/stop_training_button.jpg")
        self.stop_training_button_photo = ImageTk.PhotoImage(self.stop_training_button_img)
        self.stop_training_button = tk.Button(self, image=self.stop_training_button_photo,
                                         command=self.stop_training_button_click,
                                         bd=0, highlightthickness=0)  # Set border width to 0 to remove button border
        self.stop_training_button.image = self.stop_training_button_photo  # Prevent garbage collection
        self.stop_training_button.place(x=15, y=10)
        self.stop = False


        # self.stop = False

        self.pause_training_button_img = Image.open("Pictures/pause_training_button.jpg")
        self.pause_training_button_photo = ImageTk.PhotoImage(self.pause_training_button_img)
        self.pause_training_button = tk.Button(self, image=self.pause_training_button_photo,
                                         command=self.pause_training_button_click,
                                         bd=0, highlightthickness=0)  # Set border width to 0 to remove button border
        self.pause_training_button.image = self.pause_training_button_photo  # Prevent garbage collection
        self.pause_training_button.place(x=95, y=10)

        # self.another_calibration_button_img = Image.open("Pictures/calibration_button.jpg")
        # self.another_calibration_button_photo = ImageTk.PhotoImage(self.another_calibration_button_img)
        # self.another_calibration_button = tk.Button(self, image=self.another_calibration_button_photo,
        #                                       command=self.another_calibration_button_click,
        #                                       bd=0,
        #                                       highlightthickness=0)  # Set border width to 0 to remove button border
        # self.another_calibration_button.image = self.another_calibration_button_photo  # Prevent garbage collection
        # self.another_calibration_button.place(x=175, y=10)

        # Place background image on canvas
        self.canvas.create_image(0, 0, image=self.background_photo, anchor="nw")

        # List to hold references to the images (to avoid garbage collection)
        self.image_references = []

        # Predetermined positions (you can adjust these to your liking)
        self.positions = [
            (200, 250), (370, 250), (540, 250), (710, 250), (880, 250),
            (200, 420), (370, 420), (540, 420), (710, 420), (880, 420)
        ]


        # Track count for patient repetitions
        self.count = 1


        # Add a scale widget
        self.scale_value_label = tk.Label(self, text=f"Volume \n {s.volume*100}", font=("Arial", 16, "bold"), bg="white", fg="black")
        self.scale_value_label.place(x=5, y=100)

        self.scale = tk.Scale(self, from_=100, to=0, orient="vertical", length=300,
                              command=self.on_scale_moved, bg="#50a6ad", fg="black",
                              troughcolor="#83c2c6", highlightthickness=0, showvalue=0)
        self.scale.set(s.volume*100)  # Set initial value
        self.scale.place(x=35, y=160)


        # Time of the last update
        self.last_update_time = 0

        # Label for repetition count
        self.repetition_label_robot = tk.Label(self,
                                         text=f":רובוט\n{s.robot_counter}/{s.rep}",
                                         font=("Arial", 35, "bold"),
                                         bg="white", fg="black",
                                         justify="center", anchor="center")

        self.repetition_label_robot.place(x=870, y=10)  # Adjust x, y for correct positioning

        # Label for repetition count
        self.repetition_label_patient = tk.Label(self,
                                         text=f"{s.patient_repetitions_counting_in_exercise}",
                                         font=("Arial", 70, "bold"),
                                         bg="white", fg="black",
                                         justify="center", anchor="center")

        self.repetition_label_patient.place(x=480, y=10)  # Adjust x, y for correct positioning

        ###########################################################################
        self.exercise_type = exercise_type
        # self.real_distance = real_distance  # ✅ Manually set real distance
        self.reverse_color = reverse_color
        self.reverse_bar = reverse_bar
        self.min_distance = min_distance
        self.max_distance = max_distance
        self.min_distance_x = min_distance_x
        self.max_distance_x = max_distance_x
        self.which_side = which_side
        self.time_of_comment = 0

        # # Create a canvas for images
        # self.canvas = tk.Canvas(self, width=background_image.width, height=background_image.height)
        # self.canvas.pack(fill="both", expand=True)
        # self.bg_image = self.canvas.create_image(0, 0, image=self.background_photo, anchor="nw")

        # Bar properties
        self.bar_start_x = self.background_image.width // 2  # Center
        self.bar_start_y = self.background_image.height - 30  # Bottom
        self.bar_width = 30  # Thickness
        self.bar_length = 0  # Initial length
        self.bar_color = "red"  # Default color

        # Create the bar
        self.bar = self.canvas.create_rectangle(
            self.bar_start_x, self.bar_start_y - self.bar_width // 2,
            self.bar_start_x, self.bar_start_y + self.bar_width // 2,
            fill=self.bar_color, outline=""
        )

        # Comment Label (Updates dynamically)
        self.comment_label = tk.Label(self, text="", font=("Arial", 50, "bold"), fg="red", bg="white")
        self.comment_label.place(x=0, y=20)  # Temporary position

        # if s.req_exercise == "ball_open_arms_and_forward":
        #     s.was_in_opposite_limit = False
        # else:
        #     s.was_in_opposite_limit = True

        self.image_direction = None

        self.direction_image = None  # Store image reference to prevent garbage collection


        # Start GUI update loop
        self.after(1, self.update_exercise())

    def center_comment_label(self):
        """Centers the comment label dynamically based on its width."""
        self.update_idletasks()  # Ensure Tkinter updates the label size
        canvas_width = self.canvas.winfo_width()
        label_width = self.comment_label.winfo_reqwidth()  # Get the current width of the label

        # Set new position to center the label
        self.comment_label.place(x=(canvas_width - label_width) // 2, y=250)


    def update_bar(self, keypoints):
        """Updates the bar based on the selected exercise type."""
        if self.exercise_type == "wrist_distance_x":  # פתיחת ידיים לצדדים
            self.update_wrist_distance_x(keypoints)
        elif self.exercise_type == "shoulders_distance_x":  # כל ה-switch
            self.update_shoulders_distance_x(keypoints)
        elif self.exercise_type == "single_wrist_x":
            self.update_single_wrist_x(keypoints)  # Right wrist & shoulder
        elif self.exercise_type == "wrist_height_y":
            self.update_wrist_height_y(keypoints)
        elif self.exercise_type == "wrist_height_y_distance_x":
            self.update_wrist_height_y_distance_x(keypoints)



    def on_scale_moved(self, value):
        current_time = time.time()  # Get the current time
        if current_time - self.last_update_time >= 0.01:  # Check if 0.1 seconds have passed
            self.update_scale_value(value)
            self.last_update_time = current_time  # Update the last update time

    def update_scale_value(self, value):
        # Update the volume and label
        self.scale_value_label.config(text=f"Volume \n {value}")
        s.volume = round(int(value) * 0.01, 2)
        print(f"Updated volume to {s.volume}")


        # elif s.repeat_explanation:
        #     if self.name_of_exercise_repeated_explanation is None:
        #         s.req_exercise = ""
        #         self.name_of_exercise_repeated_explanation = s.req_exercise
        #         say("repeat_explanation")
        #         time.sleep(get_wav_duration("repeat_explanation"))

    def update_exercise(self):
        # Non-blocking exercise loop using after()

        if (not s.gymmy_done) or (not s.camera_done):  # While the exercise is still running
            if s.suggest_repeat_explanation and (s.exercise_name_repeated_explanation is None or not s.exercise_name_repeated_explanation == s.req_exercise):
                s.exercise_name_repeated_explanation = s.req_exercise
                self.show_repeat_explanation_overlay()

            else:
                if s.suggest_repeat_explanation and s.exercise_name_repeated_explanation and s.exercise_name_repeated_explanation == s.req_exercise:
                    s.suggest_repeat_explanation = False
                    s.repeat_explanation = False


                if not self.previous_all_rules_ok and s.all_rules_ok:
                    self.previous_all_rules_ok = True
                    self.comment_label.config(text="", fg="red", bg=self.cget("bg"))  # Hide it completely
                    self.comment_label.place_forget()  # Remove from layout
                    self.comment = None  # No comments, reset
                    self.time_of_comment = 0


                elif self.previous_all_rules_ok and not s.all_rules_ok:
                    self.previous_all_rules_ok = False
                    self.comment_label.config(text="", fg="red", bg=self.cget("bg"))  # Hide it completely
                    self.comment_label.place_forget()  # Remove from layout
                    self.comment = None  # No comments, reset
                    self.time_of_comment = 0

                self.repetition_label_robot.config(text=f":רובוט\n{s.robot_counter}/{s.rep}")
                # Label for repetition count
                self.repetition_label_patient.config(text=f"{s.patient_repetitions_counting_in_exercise}")


                keypoints = s.latest_keypoints

                if keypoints:
                    self.update_bar(keypoints)

                if self.percent_of_bar < 0.2 and self.start_of_750_ms_bar_len is None and s.all_rules_ok:
                    self.start_of_750_ms_bar_len = time.time()

                elif self.percent_of_bar < 0.2 and self.start_of_750_ms_bar_len is not None and s.all_rules_ok:
                    if time.time() - self.start_of_750_ms_bar_len >= 0.75:  # 50 milliseconds = 0.05 seconds
                        self.can_comment = True

                elif (self.percent_of_bar > 0.2 and not s.reached_max_limit) or (self.percent_of_bar <= 0.2 and not s.all_rules_ok):
                    self.start_of_750_ms_bar_len = None
                    self.can_comment = False

                else:
                    if s.reached_max_limit and self.start_of_750_ms_bar_len is None:
                        self.start_of_750_ms_bar_len = time.time()

                    elif s.reached_max_limit and self.start_of_750_ms_bar_len is not None:
                        if time.time() - self.start_of_750_ms_bar_len >= 0.75:  # 50 milliseconds = 0.05 seconds
                            self.can_comment = True

                    elif not s.reached_max_limit:
                        self.start_of_750_ms_bar_len = None
                        self.can_comment = False


                if s.not_reached_max_limit_rest_rules_ok and self.start_of_time_count_all_rules_not_limit is None:
                    self.start_of_time_count_all_rules_not_limit = time.time()

                elif not s.not_reached_max_limit_rest_rules_ok:
                    self.start_of_time_count_all_rules_not_limit = None

                if s.hand_not_good and self.start_of_time_count_hands_not_good is None:
                    self.start_of_time_count_hands_not_good = time.time()

                elif not s.hand_not_good:
                    self.start_of_time_count_hands_not_good = None


                if s.rate == "slow":
                    time_limit_change_position = 5
                    time_limit_all_rules_not_limit = 1.25

                elif s.rate == "moderate":
                    time_limit_change_position = 3
                    time_limit_all_rules_not_limit = 1

                else:
                    time_limit_change_position = 3
                    time_limit_all_rules_not_limit = 0.75


                if s.req_exercise == "notool_raising_hands_diagonally":
                    time_limit_hands_not_good = 7

                else:
                    time_limit_hands_not_good = 1

                if not s.did_training_paused:

                    if s.req_exercise != "" and s.can_comment_robot and\
                        ((s.reached_max_limit and not s.all_rules_ok  and self.can_comment and s.was_in_first_condition) or \
                         ((s.all_rules_ok and self.percent_of_bar < 0.2 and self.can_comment) or (not s.all_rules_ok and self.percent_of_bar < 0.2 and not s.was_in_first_condition) and\
                         ((s.req_exercise in ["notool_right_hand_up_and_bend", "notool_left_hand_up_and_bend"]) or not self.exercise_type == "shoulders_distance_x")) or \
                         (s.time_of_change_position and time.time() - s.time_of_change_position >= time_limit_change_position) or \
                         (s.not_reached_max_limit_rest_rules_ok and time.time() - self.start_of_time_count_all_rules_not_limit >= time_limit_all_rules_not_limit) or \
                         (s.hand_not_good and time.time() - self.start_of_time_count_hands_not_good >= time_limit_hands_not_good) or \
                            not s.reached_max_limit and (s.time_of_change_position is None and time.time() - self.time_of_exercise_start >= 6 or time.time()- s.time_of_change_position >= 4)):
                            self.check_are_there_comments()

                    else:
                        self.comment_label.config(text="", fg="red", bg=self.cget("bg"))  # Hide it completely
                        self.comment_label.place_forget()  # Remove from layout
                        self.comment = None  # No comments, reset
                        self.time_of_comment = 0


                #
                # elif not s.all_rules_ok:
                #     self.start_of_sec = None

                # Update direction arrow on the Canvas
                if s.direction is None:
                    if hasattr(self, "direction_canvas_image"):
                        self.canvas.delete(self.direction_canvas_image)  # Remove previous image
                    self.image_direction = None

                elif s.direction == self.image_direction:
                    pass  # Skip update if the direction hasn't changed

                else:
                    image_path = f"Pictures/{self.chosen_subject}/{s.direction}_arrow.png"
                    try:
                        # Convert white/near-white to transparent if needed
                        transparent_image = convert_white_to_transparent(image_path, tolerance=2)

                        # Convert image for Tkinter
                        self.direction_image = ImageTk.PhotoImage(transparent_image)

                        # Remove previous arrow if it exists
                        if hasattr(self, "direction_canvas_image"):
                            self.canvas.delete(self.direction_canvas_image)

                        # Ensure Canvas size updates before centering
                        self.update_idletasks()


                        # Calculate correct x position to center at x=400
                        new_x = 350

                        # Add new arrow image to the Canvas (centered at x=400)
                        self.direction_canvas_image = self.canvas.create_image(
                            new_x,  # Center X
                            75,  # Y position (adjust as needed)
                            image=self.direction_image,
                            anchor="center"
                        )

                        self.image_direction = s.direction  # Store current direction to avoid redundant updates

                    except Exception as e:
                        print(f"Error loading image {image_path}: {e}")


                if self.count == s.patient_repetitions_counting_in_exercise:
                    # Get the center position
                    x_center, y_center = self.positions[self.count - 1]  # Adjust index

                    # Path to a random image in the chosen category
                    image_num = random.randint(1, 27)
                    image_path = f'Pictures/{self.chosen_subject}/{image_num}.png'

                    # Convert near-white background to transparent (adjust tolerance as needed)
                    transparent_image = convert_white_to_transparent(image_path, tolerance=30)

                    # Load the image with transparent background into Tkinter
                    exercise_photo = ImageTk.PhotoImage(transparent_image)

                    # Place the image at the exact center
                    self.canvas.create_image(
                        x_center,  # Exact center X
                        y_center,  # Exact center Y
                        image=exercise_photo,
                        anchor="center"  # Center the image properly
                    )

                    # Append each image reference to prevent garbage collection
                    self.image_references.append(exercise_photo)

                    # Increment the count to wait for the next successful repetition
                    self.count += 1

                # Schedule next iteration to keep updating the screen
                self.after(1, self.update_exercise)  # Call the function every 1ms

        else:
            print("Exercise complete")

    def hide_overlay(self):
        if hasattr(self, 'overlay_frame') and self.overlay_frame is not None:
            self.overlay_frame.destroy()
            self.overlay_frame = None

    def show_repeat_explanation_overlay(self):
        button_style = {
            "font": ("Arial", 25, "bold"),
            "bg": "#3996da",
            "fg": "white",  # a calm deep blue
            "activebackground": "#e1f5fe",
            "activeforeground": "#0d47a1",
            "width": 15,
            "height": 3,
            "relief": "raised",
            "bd": 3,
            "cursor": "hand2"
        }


        self.end_of_comment_recording = None
        self.comment_label.config(text="", fg="red", bg=self.cget("bg"))  # Hide it completely
        self.comment_label.place_forget()  # Remove from layout
        self.comment = None  # No comments, reset
        self.time_of_comment = 0

        if s.suggest_repeat_explanation:
            say("ask_if_repeat_explanation",is_popping_screen = True)
            s.did_training_paused = True

            # Create overlay
            self.overlay_frame = tk.Frame(self, width=1024, height=576, bg="#d3f1f1")
            self.overlay_frame.place(x=0, y=0)
            self.overlay_frame.tkraise()

            # Label
            label = tk.Label(self.overlay_frame, text="?תרצה לראות שוב את ההסבר על התרגיל",
                             font=("Arial", 35), bg="#d3f1f1")
            label.place(relx=0.5, rely=0.2, anchor="center")

            # YES Button
            yes_button = tk.Button(self.overlay_frame, text="כן",
                                   command=self.repeat_explanation_yes, **button_style)
            yes_button.place(relx=0.8, rely=0.5, anchor="center")

            # NO Button
            no_button = tk.Button(self.overlay_frame, text="לא",
                                  command=self.repeat_explanation_no, **button_style)
            no_button.place(relx=0.5, rely=0.5, anchor="center")

            # Skip Button
            skip_button = tk.Button(self.overlay_frame, text="דלג על התרגיל",
                                    command=self.skip_exercise, **button_style)
            skip_button.place(relx=0.2, rely=0.5, anchor="center")

            self.timer_canvas = tk.Canvas(self.overlay_frame, width=150, height=150, bg="#d3f1f1", highlightthickness=0)
            self.timer_canvas.place_forget()  # Hide initially

            self.timer_canvas.create_oval(10, 10, 140, 140, outline="gray", width=6)  # Static full circle
            self.timer_arc = self.timer_canvas.create_arc(10, 10, 140, 140, start=90, extent=360,
                                                          style="arc", outline="lime green", width=10)  # Thicker arc
            self.timer_text = self.timer_canvas.create_text(75, 75, text="", fill="black", font=("Arial", 60, "bold"))

            # Start countdown after 10 seconds
            self.after(int(get_wav_duration("ask_if_repeat_explanation") * 1000 + 100) , self.start_circle_countdown)


        else:

            # Create overlay
            self.overlay_frame = tk.Frame(self, width=1024, height=576, bg="#d3f1f1")
            self.overlay_frame.place(x=0, y=0)
            self.overlay_frame.tkraise()


            yes_button = tk.Button(self.overlay_frame, text="חזור על ההסבר",
                                   command=self.repeat_explanation_yes, **button_style)
            yes_button.place(relx=0.7, rely=0.2, anchor="center")

            no_button = tk.Button(self.overlay_frame, text="המשך באימון",
                                  command=self.repeat_explanation_no, **button_style)
            no_button.place(relx=0.3, rely=0.2, anchor="center")

            skip_button = tk.Button(self.overlay_frame, text="דלג על התרגיל",
                                    command=self.skip_exercise, **button_style)
            skip_button.place(relx=0.7, rely=0.5, anchor="center")

            calibration_button = tk.Button(self.overlay_frame, text="בצע כיול חוזר",
                                           command=self.another_calibration_button_click, **button_style)
            calibration_button.place(relx=0.3, rely=0.5, anchor="center")


            label = tk.Label(self.overlay_frame, text="אי לחיצה על אף כפתור תמשיך את השהיית האימון *",
                             font=("Arial", 30, "bold"), bg="#d3f1f1", fg= "red")
            label.place(relx=0.5, rely=0.8, anchor="center")


    def start_circle_countdown(self):
        if not hasattr(self, 'timer_canvas') or not self.timer_canvas.winfo_exists():
            print("❗ Timer canvas doesn't exist — skipping countdown.")
            return

        self.timer_canvas.place(relx=0.5, rely=0.8, anchor="center")
        self.countdown_total_time = 5
        self.countdown_current = self.countdown_total_time
        self.circle_countdown_step()


    def circle_countdown_step(self):
        if self.countdown_current >= 0:
            angle = (self.countdown_current / self.countdown_total_time) * 360
            self.timer_canvas.itemconfig(self.timer_arc, extent=-angle)

            # Bold and large countdown text
            self.timer_canvas.itemconfig(self.timer_text,
                                         text=str(self.countdown_current),
                                         font=("Arial", 60, "bold"))

            self.countdown_current -= 1
            self.countdown_timer = self.after(1000, self.circle_countdown_step)
        else:
            self.repeat_explanation_no()

    def skip_exercise(self):
        self.cancel_repeat_overlay_timers()
        s.skipped_exercise = True
        s.did_training_paused = False
        s.suggest_repeat_explanation = False
        self.hide_overlay()

    def repeat_explanation_yes(self):
        self.cancel_repeat_overlay_timers()
        s.repeat_explanation = True
        s.did_training_paused = False
        s.suggest_repeat_explanation = False
        self.hide_overlay()

    def repeat_explanation_no(self):
        self.cancel_repeat_overlay_timers()
        print("User skipped explanation.")
        self.hide_overlay()
        s.did_training_paused = False
        s.suggest_repeat_explanation = False
        self.after(1, self.update_exercise)

    def cancel_repeat_overlay_timers(self):
        if hasattr(self, 'start_circle_countdown_timer'):
            self.after_cancel(self.start_circle_countdown_timer)
            del self.start_circle_countdown_timer

        if hasattr(self, 'countdown_timer'):
            self.after_cancel(self.countdown_timer)
            del self.countdown_timer

        if hasattr(self, 'overlay_frame'):
            self.overlay_frame.destroy()
            del self.overlay_frame

    def stop_training_button_click(self):
        s.req_exercise = ""
        s.stop_requested=True
        # s.finish_workout= True
        self.after_cancel(self.after_id)  # Cancel any pending after() calls

        print("Stop training button clicked")

    def another_calibration_button_click(self):
        # s.stop_requested = True
        # s.finish_workout= True
        # self.after_cancel(self.after_id)  # Cancel any pending after() calls

        s.try_again_calibration = True

        print("Stop training button clicked")

    def pause_training_button_click(self):
        s.starts_and_ends_of_stops.append(time.time())

        self.show_repeat_explanation_overlay()
        s.did_training_paused = True

        # if s.did_training_paused:
        #     # Define what happens when the button is clicked
        #     print("Pause training button clicked")
        #     s.did_training_paused= False
        #
        #     if hasattr(self, 'pause_training_button'):
        #         self.pause_training_button.destroy()
        #
        #         # Create a new button
        #     self.pause_training_button_img = Image.open("Pictures//pause_training_button.jpg")
        #     self.pause_training_button_photo = ImageTk.PhotoImage(self.pause_training_button_img)
        #     self.pause_training_button = tk.Button(self, image=self.pause_training_button_photo,
        #                                            command=self.pause_training_button_click,
        #                                            bd=0, highlightthickness=0)
        #     self.pause_training_button.image = self.pause_training_button_photo  # Prevent garbage collection
        #     self.pause_training_button.place(x=95, y=10)
        #     s.number_of_pauses += 1 # when continuing the training after pause, we add the pause (if we add it before there is a chance that we stopped it after the pause).
        #
        #
        # else:
        #     if hasattr(self, 'pause_training_button'):
        #         self.pause_training_button.destroy()
        #
        #         # Create a new button
        #     self.pause_training_button_img = Image.open("Pictures//continue_training_button.jpg")
        #     self.pause_training_button_photo = ImageTk.PhotoImage(self.pause_training_button_img)
        #     self.pause_training_button = tk.Button(self, image=self.pause_training_button_photo,
        #                                            command=self.pause_training_button_click,
        #                                            bd=0, highlightthickness=0)
        #     self.pause_training_button.image = self.pause_training_button_photo  # Prevent garbage collection
        #     self.pause_training_button.place(x=95, y=10)
        #
        #     s.did_training_paused = True



    #########################################
    def update_wrist_height_y(self, keypoints):

        # Get wrist and shoulder positions
        right_wrist = keypoints["R_wrist"]  # Right wrist
        left_wrist = keypoints["L_wrist"]  # Left wrist
        right_shoulder = keypoints["R_shoulder"]  # Right shoulder
        left_shoulder = keypoints["L_shoulder"]  # Left shoulder

        # Ensure all keypoints are detected
        if any(np.all(kp == -1) for kp in [right_wrist, left_wrist, right_shoulder, left_shoulder]):
            return  # Skip if any keypoint is missing

        if self.which_side == "both":
            # Compute average heights
            wrist_heights = (right_wrist.y + left_wrist.y) / 2
            shoulder_heights = (right_shoulder.y + left_shoulder.y) / 2

        elif self.which_side == "left":
            # Compute average heights
            wrist_heights = left_wrist.y
            shoulder_heights = left_shoulder.y

        else:
            # Compute average heights
            wrist_heights = right_wrist.y
            shoulder_heights = right_shoulder.y


        # Compute wrist-to-shoulder height difference (preserving sign)
        wrist_height_avg = wrist_heights - shoulder_heights

        # Maximum bar length
        max_bar_length = self.background_image.width // 2

        # Scale the bar length
        if wrist_height_avg >= self.max_distance:
            scaled_length = 0 if not self.reverse_bar else max_bar_length  # Default: Smallest bar when far below shoulders
        elif wrist_height_avg <= self.min_distance:
            scaled_length = max_bar_length if not self.reverse_bar else 0  # Default: Full bar when above shoulders
        else:
            # Interpolate smoothly between min and max distances
            scaled_length = max_bar_length * (
                        1 - (wrist_height_avg - self.min_distance) / (self.max_distance - self.min_distance))

            # If reversed, flip the scaling effect
            if self.reverse_bar:
                scaled_length = max_bar_length - scaled_length

        # Set color: Green when close, transitioning to red when far
        self.bar_color = self.get_color_gradient(wrist_height_avg, self.min_distance, self.max_distance,
                                                 reverse=self.reverse_color)

        # print("Height Difference:", wrist_height_avg, "| Scaled Length:", scaled_length)

        # Update the bar display
        self.update_bar_display(scaled_length)

    import numpy as np

    def update_wrist_height_y_distance_x(self, keypoints):
        # Get wrist and shoulder positions
        right_wrist = keypoints["R_wrist"]
        left_wrist = keypoints["L_wrist"]
        right_shoulder = keypoints["R_shoulder"]
        left_shoulder = keypoints["L_shoulder"]

        # Ensure all keypoints are detected
        if any(np.all(kp == -1) for kp in [right_wrist, left_wrist, right_shoulder, left_shoulder]):
            return  # Skip if any keypoint is missing

        # Compute wrist-to-wrist distance (x-axis)
        wrist_distance_x = abs(right_wrist.x - left_wrist.x)

        # Compute wrist-to-shoulder height difference (y-axis)
        wrist_heights = (right_wrist.y + left_wrist.y) / 2
        shoulder_heights = (right_shoulder.y + left_shoulder.y) / 2
        wrist_height_y = wrist_heights - shoulder_heights  # Preserve sign

        # Maximum bar length
        max_bar_length = self.background_image.width // 2
        half_bar_length = max_bar_length / 2  # First 50% comes from x-distance

        if s.req_exercise == "band_open_arms_and_up":
            # Normalize x-distance to the first half of the bar
            if wrist_distance_x <= self.min_distance_x:
                scaled_length_x = 0
            elif wrist_distance_x >= self.max_distance_x:
                scaled_length_x = half_bar_length
            else:
                scaled_length_x = ((wrist_distance_x - self.min_distance_x) / (
                            self.max_distance_x - self.min_distance_x)) * half_bar_length

            if scaled_length_x < half_bar_length:
                # If x-distance hasn't filled the first half, only use x-distance
                scaled_length = scaled_length_x
            else:
                # Normalize y-distance to the second half of the bar
                if wrist_height_y <= self.min_distance:
                    scaled_length_y = half_bar_length
                elif wrist_height_y >= self.max_distance:
                    scaled_length_y = 0
                else:
                    scaled_length_y = ((self.max_distance - wrist_height_y) / (
                                self.max_distance - self.min_distance)) * half_bar_length

                # Final bar length: X contribution + Y contribution
                scaled_length = half_bar_length + max(0, scaled_length_y)

        # If reversed, flip the scaling effect
        if self.reverse_bar:
            scaled_length = max_bar_length - scaled_length

        # Set color using the bar length instead of wrist height
        self.bar_color = self.get_color_gradient(scaled_length, 0, max_bar_length, reverse=self.reverse_color)

        # Check exercise conditions
        if s.req_exercise == "band_open_arms_and_up":
            wrist_distance = abs(right_wrist.x - left_wrist.x)

            if s.direction == "out" and wrist_distance >= s.dist_between_wrists / 2:
                s.direction = "up"
                s.time_of_change_position = time.time()

            elif s.direction == "down" and abs(wrist_heights - shoulder_heights) < 50:
                s.direction = "in"
                s.time_of_change_position = time.time()


        # Update the bar display
        self.update_bar_display(scaled_length)


    def update_wrist_distance_x(self, keypoints):
        right_wrist = keypoints["R_wrist"]  # Right wrist
        left_wrist = keypoints["L_wrist"]  # Left wrist

        if np.all(right_wrist == -1) or np.all(left_wrist == -1):
            return  # Skip if any wrist is missing

        # Compute hand distance
        distance = abs(right_wrist.x - left_wrist.x)

        # # Define key distances
        # min_distance = self.real_distance   # The closest distance (smallest bar, most red)
        # max_distance = self.real_distance + 1000  # The farthest distance (largest bar, greenest)

        # if distance <= self.min_distance + 100:
        #     s.was_in_opposite_limit = True

        # Maximum bar length
        max_bar_length = self.background_image.width // 2

        # Scale the bar length (opposite logic from shoulders)
        if distance <= self.min_distance:
            scaled_length = 0  # Smallest bar when hands are too close
        elif distance >= self.max_distance:
            scaled_length = max_bar_length  # Largest bar when hands are far apart
        else:
            # Interpolate smoothly
            scaled_length = max_bar_length * ((distance - self.min_distance) / (self.max_distance - self.min_distance))

        # Set color: Green when far, transitioning to red when close
        self.bar_color = self.get_color_gradient(distance, self.min_distance, self.max_distance,
                                                 reverse=self.reverse_color)

        # Update the bar display
        self.update_bar_display(scaled_length)

    def update_shoulders_distance_x(self, keypoints):
        right_shoulder = keypoints["R_shoulder"]  # Right shoulder
        left_shoulder = keypoints["L_shoulder"]  # Left shoulder

        if np.all(right_shoulder == -1) or np.all(left_shoulder == -1):
            return  # Skip if any shoulder is missing

        # Compute shoulder distance
        distance = abs(right_shoulder.x - left_shoulder.x)


        if distance >= s.dist_between_shoulders - 30 and abs(right_shoulder.y - left_shoulder.y) <= 10 and\
                s.req_exercise not in ["notool_right_hand_up_and_bend", "notool_left_hand_up_and_bend"] and s.all_rules_ok:
            s.all_rules_ok = False
            s.time_of_change_position = time.time()

        # Scale the bar length
        if distance >= self.max_distance:
            scaled_length = 0  # Smallest bar when far apart
        elif distance <= self.min_distance:
            scaled_length = self.background_image.width // 2  # Full bar at best proximity
        else:
            # Interpolate smoothly between the two distances
            max_bar_length = self.background_image.width // 2
            scaled_length = max_bar_length * (
                        1 - (distance - self.min_distance) / (self.max_distance - self.min_distance))

        # Set color: Green when close, transitioning to red when far
        self.bar_color = self.get_color_gradient(distance, self.min_distance, self.max_distance, self.reverse_color)

        # Update the bar display
        self.update_bar_display(scaled_length)

        # reverse_color=False, reverse_bar=True

    # def update_shoulders_distance_y(self, keypoints):
    #     # Get shoulder positions
    #     right_shoulder = keypoints["R_shoulder"]  # Right shoulder
    #     left_shoulder = keypoints["L_shoulder"]  # Left shoulder
    #
    #     # Ensure all keypoints are detected
    #     if np.all(right_shoulder == -1) or np.all(left_shoulder == -1):
    #         return  # Skip if any keypoint is missing
    #
    #     # Compute absolute **vertical** distance between shoulders (ignoring direction)
    #     distance = abs(right_shoulder.y - left_shoulder.y)
    #
    #     if distance <= 30 and abs(right_shoulder.y - left_shoulder.y) <= 10:
    #         s.all_rules_ok = False
    #         s.was_in_opposite_limit = True
    #
    #     # Maximum bar length
    #     max_bar_length = self.background_image.width // 2
    #
    #     # Scale the bar length
    #     if distance >= self.max_distance:
    #         scaled_length = 0 if not self.reverse_bar else max_bar_length  # Default: Shortest bar when far apart
    #     elif distance <= self.min_distance:
    #         scaled_length = max_bar_length if not self.reverse_bar else 0  # Default: Longest bar when shoulders aligned
    #     else:
    #         # Interpolate smoothly
    #         scaled_length = max_bar_length * (
    #                     1 - (distance - self.min_distance) / (self.max_distance - self.min_distance))
    #
    #         # If reversed, flip the scaling effect
    #         if self.reverse_bar:
    #             scaled_length = max_bar_length - scaled_length
    #
    #     # Set color based on `reverse_color`
    #     self.bar_color = self.get_color_gradient(distance, self.min_distance, self.max_distance,
    #                                              reverse=self.reverse_color)
    #
    #     # print("Shoulder Height Difference:", distance, "| Scaled Length:", scaled_length)
    #
    #     # Update the bar display
    #     self.update_bar_display(scaled_length)


    def get_color_gradient(self, distance, min_distance, max_distance, reverse=False):
        """Returns the color based on distance: red when far, green when close, but prevents fluctuations."""

        # Normalize distance to [0,1] range
        norm_factor = (distance - min_distance) / (max_distance - min_distance)
        norm_factor = max(0, min(1, norm_factor))  # Clamp to [0,1]

        # Convert to discrete 10-step levels (0-9)
        step = max(0, min(9, int(norm_factor * 10)))  # Ensure step is between 0-9

        # If all_rules_ok is False, prevent step beyond 8
        if not s.all_rules_ok:
            step = min(step, 8)  # Limit to step 8 (not fully green)

        max_green = 150  # Maximum green value (Never full 255)

        if reverse:
            r = max(0, min(255, int(255 * (step / 8))))  # Ensure within [0, 255]
            g = max(0, min(255, int(max_green * (1 - step / 8))))
        else:
            r = max(0, min(255, int(255 * (1 - step / 8))))
            g = max(0, min(255, int(max_green * (step / 8))))

            # If all_rules_ok is True and max limit is reached, return fully green
        if s.all_rules_ok and s.reached_max_limit:
            return "#00FF00"

        # Convert RGB to valid hex color format
        return f"#{r:02X}{g:02X}00"

    def update_bar_display(self, scaled_length):
        """Ensures the bar length correctly maps to increasing distances."""

        max_width = self.background_image.width // 2  # Default max width
        max_limited_width = int(max_width * 0.9)  # 80% limit

        # If all_rules_ok is False, limit bar length to 80% max
        if not s.all_rules_ok:
            max_width = max_limited_width  # Limit max width to 80%

        # Ensure scaled_length is increasing consistently with distance
        scaled_length = min(scaled_length, max_width)  # Ensure it does not exceed max width
        scaled_length = max(0, scaled_length)  # Ensure it never becomes negative

        self.percent_of_bar = scaled_length/max_width

        # Check if it reaches the 80% limit
        if scaled_length >= max_limited_width:
            s.reached_max_limit = True  # Set condition as True



        if scaled_length < max_limited_width:
            s.reached_max_limit = False


        # Update the bar's position on the canvas
        self.canvas.coords(self.bar,
                           self.bar_start_x - scaled_length // 2, self.bar_start_y - self.bar_width // 2,
                           self.bar_start_x + scaled_length // 2, self.bar_start_y + self.bar_width // 2)
        self.canvas.itemconfig(self.bar, fill=self.bar_color)


    def check_are_there_comments(self):

        angles = s.last_entry_angles
        self.comments = []

        # use_alternate_angles = s.last_entry_angles[-3]
        # left_right_differ = s.last_entry_angles[-2]
        # side = s.last_entry_angles[-1]

        if angles:
            # Iterate over first_part in steps of 3
            for i in range(0, len(angles)):
                angle = angles[i]
                information_angle = s.information[i]


                if angle:
                    if not information_angle[3] <= angle <= information_angle[4]:
                        if angle <= information_angle[3]:
                            self.what_to_comment(information_angle[0], information_angle[1], information_angle[2],
                                                           "smaller", s.direction)
                        else:
                            self.what_to_comment(information_angle[0], information_angle[1], information_angle[2],
                                                           "bigger", s.direction)

        # if s.req_exercise == "notool_raising_hands_diagonally":


        if s.hand_not_good and self.start_of_time_count_hands_not_good is not None and time.time()- self.start_of_time_count_hands_not_good >= 1 or\
            not s.reached_max_limit and (s.time_of_change_position is None and time.time() - self.time_of_exercise_start >= 6 or time.time()- s.time_of_change_position >= 4):

            if s.req_exercise == "notool_raising_hands_diagonally" and self.start_of_time_count_hands_not_good and time.time()- self.start_of_time_count_hands_not_good >= 7:
                if s.direction is not None:
                    if s.direction == "left_diagonal":
                        if s.gender == "Male":

                            self.comments.append("מתח את הידיים מעבר \n לכתף שמאל")
                        else:
                            self.comments.append("מתחי את הידיים מעבר \n לכתף שמאל")


                    else:
                        if s.gender == "Male":
                            self.comments.append("מתח את הידיים מעבר \n לכתף ימין")

                        else:
                            self.comments.append("מתחי את הידיים מעבר \n לכתף ימין")


            if s.req_exercise == "notool_right_hand_up_and_bend":
                if not s.all_rules_ok:
                    if s.direction is not None:
                        if s.direction == "left":
                            if s.gender == "Male":
                                self.comments.append("הישען יותר ושלח את יד ימין \n יותר לכיוון הרצפה")

                            else:
                                self.comments.append("הישעני יותר ושלחי את יד ימין \n יותר לכיוון הרצפה")

                else:
                    if s.direction is not None:
                        if s.direction == "right":
                            if s.gender == "Male":
                                self.comments.append("התיישר יותר לצד ימין והבא \n את יד ימין מעל כתף ימין")

                            else:
                                self.comments.append("התיישרי יותר לצד ימין והביאי \n את יד ימין מעל כתף ימין")


            if s.req_exercise == "notool_left_hand_up_and_bend":
                if not s.all_rules_ok:
                    if s.direction is not None:
                        if s.direction == "right":
                            if s.gender == "Male":
                                self.comments.append("הישען יותר ושלח את יד שמאל \n יותר לכיוון הרצפה")

                            else:
                                self.comments.append("הישעני יותר ושלחי את יד שמאל \n יותר לכיוון הרצפה")

                if s.all_rules_ok:
                    if s.direction is not None:
                        if s.direction == "left":
                            if s.gender == "Male":
                                self.comments.append("התיישר יותר לצד שמאל והבא \n את יד שמאל מעל כתף שמאל")

                            else:
                                self.comments.append("התיישרי יותר לצד שמאל והביאי \n את יד שמאל מעל כתף שמאל")


        if s.not_reached_max_limit_rest_rules_ok and self.start_of_time_count_all_rules_not_limit is not None and time.time() - self.start_of_time_count_all_rules_not_limit>= 1 or\
            not s.reached_max_limit and (s.time_of_change_position is None and time.time() - self.time_of_exercise_start >= 6 or time.time()- s.time_of_change_position >= 4):
            if s.req_exercise in ["ball_switch", "stick_switch"]:
                    if s.direction is not None:
                        if s.direction == "left":
                            if s.gender == "Male":
                                self.comments.append("סובב יותר את הגב לצד שמאל \n יחד עם הכתפיים")

                            else:
                                self.comments.append("סובבי יותר את הגב לצד שמאל \n יחד עם הכתפיים")

                        else:
                            if s.gender == "Male":
                                self.comments.append("סובב יותר את הגב לצד ימין \n יחד עם הכתפיים")

                            else:
                                self.comments.append("סובבי יותר את הגב לצד ימין \n יחד עם הכתפיים")


            if s.req_exercise in  ["band_up_and_lean", "stick_up_and_lean", "notool_hands_behind_and_lean", "notool_right_hand_up_and_bend", "notool_left_hand_up_and_bend"] and (not s.reached_max_limit or \
                    not s.reached_max_limit and (s.time_of_change_position is None and time.time() - self.time_of_exercise_start >= 6 or time.time()- s.time_of_change_position >= 4)):

                if s.direction is not None:
                    if s.direction == "left" and s.req_exercise != "notool_left_hand_up_and_bend":
                            if s.gender == "Male":
                                self.comments.append("הישען יותר לצד שמאל \n יחד עם הכתפיים")

                            else:
                                self.comments.append("הישעני יותר לצד שמאל \n יחד עם הכתפיים")

                    elif s.direction == "right" and s.req_exercise != "notool_right_hand_up_and_bend":
                        if s.gender == "Male":
                            self.comments.append("הישען יותר לצד ימין \n יחד עם הכתפיים")

                        else:
                            self.comments.append("הישעני יותר לצד ימין \n יחד עם הכתפיים")



        if (self.end_of_comment_recording is None or self.end_of_comment_recording < time.time()) and (self.time_of_comment == 0 or (time.time() - self.last_loop_time) >= 1) and s.robot_counter < s.rep -1:
            self.last_loop_time = time.time()
            if self.comments:
                print("comments: ", self.comments)

                if self.comment:
                    print("comment: " + self.comment)

                if self.comment in self.comments:
                    # if time.time() - self.time_of_comment >= 2:
                    if time.time() - self.time_of_comment >= 7:
                        # Choose a different comment if there are more options
                        new_comments = [c for c in self.comments if c != self.comment]
                        if new_comments:
                            self.comment = random.choice(new_comments)
                            self.time_of_comment = time.time()
                    else:
                        pass

                    comment_no_signs = self.comment.replace(' \n', '')

                    if comment_no_signs in self.comments_audio:
                        if time.time() - self.comments_audio[comment_no_signs] >= 15:
                            can_say_audio = True

                        else:
                            can_say_audio = False

                    else:
                        can_say_audio = True

                    # if len(self.comments_audio) >= 1:
                    #     if self.comments_audio[-1] == comment_no_signs and self.end_of_comment_recording and time.time() - self.end_of_comment_recording < 10:
                    #         can_say_audio = False
                    #
                    #     elif self.end_of_comment_recording and (self.comments_audio[-1] == comment_no_signs and time.time() - self.end_of_comment_recording >= 10) or self.comments_audio[-1] != comment_no_signs:
                    #         can_say_audio = True
                    #
                    #     else:
                    #         can_say_audio = False

                    random_num = random.randint(1,5)
                    audio_path = f"{s.audio_path}{comment_no_signs}.wav"  # adjust the path and extension as needed

                    if self.end_of_comment_recording and time.time() > self.end_of_comment_recording:
                        self.end_of_comment_recording= None


                    if (time.time() - s.last_saying_time > 2) and not self.end_of_comment_recording and can_say_audio and self.time_of_comment and (
                            time.time() - self.time_of_comment >= 2) and s.robot_counter < s.rep - 1 and s.patient_repetitions_counting_in_exercise < s.rep - 1:

                        if os.path.exists(audio_path):
                            if random_num == 1: #Add something about recognition issues
                                dont_recognize_files= Excel.get_files_names_by_start_word("dont_recognize_")
                                chosen_dont_recognize = random.choice(dont_recognize_files)
                                self.end_of_comment_recording = time.time() + get_wav_duration(comment_no_signs)


                                if chosen_dont_recognize:
                                    say(chosen_dont_recognize)
                                    self.after(int(get_wav_duration(chosen_dont_recognize)*1000), lambda: say(comment_no_signs))

                                    self.end_of_comment_recording += get_wav_duration(chosen_dont_recognize)

                                else:
                                    say(comment_no_signs)


                                s.last_saying_time = time.time()
                                self.comments_audio[comment_no_signs] = time.time()

                            else:
                                say(comment_no_signs)
                                self.end_of_comment_recording = time.time() + get_wav_duration(comment_no_signs)
                                s.last_saying_time = time.time()
                                self.comments_audio[comment_no_signs] = time.time()
                        else:
                            print(f"Audio file not found: {audio_path}")

                    else:
                        pass  # Keep the same comment

                else:
                    self.time_of_comment = time.time()
                    self.comment = self.comments[0]  # Pick the first comment if the current one is not in the list

            else:
                self.end_of_comment_recording = None
                self.comment_label.config(text="", fg="red", bg=self.cget("bg"))  # Hide it completely
                self.comment_label.place_forget()  # Remove from layout
                self.comment = None  # No comments, reset
                # self.time_of_comment = 0

            if self.comment is not None:
                print(self.comment)

                # Step 1: Temporarily hide the label (so it's not visible while being positioned)
                self.comment_label.place_forget()

                # Step 2: Set the comment text
                self.comment_label.config(
                    text=self.comment,
                    fg="red",
                    font=("Arial", 50, "bold"),
                    bg="white",
                    justify="center",
                    anchor="center"
                )

                # Step 3: Update size calculations
                self.update_idletasks()

                # Step 4: Reposition centered
                self.center_comment_label()




    def what_to_comment(self, joint1, joint2, joint3, biggerORsmaller="smaller", side="left"):
        cleaned_joint_1 = (joint1.split("_"))[-1]
        cleaned_joint_2 = (joint2.split("_"))[-1]
        cleaned_joint_3 = (joint3.split("_"))[-1]


        if not s.req_exercise in ["band_straighten_left_arm_elbows_bend_to_sides", "band_straighten_right_arm_elbows_bend_to_sides", "notool_right_hand_up_and_bend", "notool_left_hand_up_and_bend", "notool_right_bend_left_up_from_side", "notool_left_bend_right_up_from_side"]:
            if cleaned_joint_2 == "elbow":
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("ישר יותר את המרפקים")

                    else:
                        self.comments.append("ישרי יותר את המרפקים")
                else:
                    if s.gender == "Male":
                        self.comments.append("כופף יותר את המרפקים")

                    else:
                        self.comments.append("כופפי יותר את המרפקים")

        elif s.req_exercise == "notool_right_hand_up_and_bend":
            if cleaned_joint_2 == "elbow":
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("ישר יותר את יד ימין")

                    else:
                        self.comments.append("ישרי יותר את יד ימין")



        elif s.req_exercise == "notool_left_hand_up_and_bend":
            if cleaned_joint_2 == "elbow":
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("ישר יותר את יד שמאל")

                    else:
                        self.comments.append("ישרי יותר את יד שמאל")


        elif s.req_exercise == "band_straighten_left_arm_elbows_bend_to_sides":
                if joint2 == "L_elbow":
                    if biggerORsmaller == "smaller":
                        if s.gender == "Male":
                            self.comments.append("ישר יותר את יד שמאל")

                        else:
                            self.comments.append("ישרי יותר את יד שמאל")
                    else:
                        if s.gender == "Male":
                            self.comments.append("כופף יותר את יד שמאל")

                        else:
                            self.comments.append("כופפי יותר את יד שמאל")

                elif joint2 == "R_elbow":
                    if biggerORsmaller == "smaller":
                        if s.gender == "Male":
                            self.comments.append("כופף מעט פחות את יד ימין")

                        else:
                            self.comments.append("כופפי מעט פחות את יד ימין")

                    else:
                        if s.gender == "Male":
                            self.comments.append("כופף יותר את יד ימין")

                        else:
                            self.comments.append("כופפי יותר את יד ימין")


        elif s.req_exercise == "band_straighten_right_arm_elbows_bend_to_sides":
                if joint2 == "L_elbow":
                    if biggerORsmaller == "smaller":
                        if s.gender == "Male":
                            self.comments.append("כופף מעט פחות את יד שמאל")

                        else:
                            self.comments.append("כופפי מעט פחות את יד שמאל")
                    else:
                        if s.gender == "Male":
                            self.comments.append("כופף יותר את יד שמאל")

                        else:
                            self.comments.append("כופפי יותר את יד שמאל")

                elif joint2 == "R_elbow":
                    if biggerORsmaller == "smaller":
                        if s.gender == "Male":
                            self.comments.append("ישר יותר את יד ימין")

                        else:
                            self.comments.append("ישרי יותר את יד ימין")
                    else:
                        if s.gender == "Male":
                            self.comments.append("כופף יותר את יד ימין")

                        else:
                            self.comments.append("כופפי יותר את יד ימין")



        elif s.req_exercise == "notool_right_bend_left_up_from_side":
            if joint2 == "L_elbow":
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("ישר יותר את יד שמאל")

                    else:
                        self.comments.append("ישרי יותר את יד שמאל")


            if joint2 == "R_elbow":
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("ישר יותר את יד ימין")

                    else:
                        self.comments.append("ישרי יותר את יד ימין")
                else:
                    if s.gender == "Male":
                        self.comments.append("כופף יותר את יד ימין")

                    else:
                        self.comments.append("כופפי יותר את יד ימין")

            if joint2 == "L_shoulder":
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("הרם יותר את יד שמאל")

                    else:
                        self.comments.append("הרימי יותר את יד שמאל")
                else:
                    if s.gender == "Male":
                        self.comments.append("הורד יותר את יד שמאל")

                    else:
                        self.comments.append("הורידי יותר את יד שמאל")

            if joint2 == "R_shoulder":
                if biggerORsmaller == "bigger":
                    if s.gender == "Male":
                        self.comments.append("הצמד יותר את זרוע ימין לגוף")

                    else:
                        self.comments.append("הצמידי יותר את זרוע ימין לגוף")


        elif s.req_exercise == "notool_left_bend_right_up_from_side":
            if joint2 == "R_elbow":
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("ישר יותר את יד ימין")

                    else:
                        self.comments.append("ישרי יותר את יד ימין")


            if joint2 == "L_elbow":
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("ישר יותר את יד שמאל")

                    else:
                        self.comments.append("ישרי יותר את יד שמאל")
                else:
                    if s.gender == "Male":
                        self.comments.append("כופף יותר את יד שמאל")

                    else:
                        self.comments.append("כופפי יותר את יד שמאל")

            if joint2 == "R_shoulder":
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("הרם יותר את יד ימין")

                    else:
                        self.comments.append("הרימי יותר את יד ימין")
                else:
                    if s.gender == "Male":
                        self.comments.append("הורד יותר את יד ימין")

                    else:
                        self.comments.append("הורידי יותר את יד ימין")

            if joint2 == "L_shoulder":
                if biggerORsmaller == "bigger":
                    if s.gender == "Male":
                        self.comments.append("הצמד יותר את זרוע שמאל לגוף")

                    else:
                        self.comments.append("הצמידי יותר את זרוע שמאל לגוף")



        if ((cleaned_joint_1 == "hip" and cleaned_joint_3 == "elbow") or (
                cleaned_joint_3 == "hip" and cleaned_joint_1 == "elbow")) and cleaned_joint_2 == "shoulder":
            if s.req_exercise in ["ball_bend_elbows", "stick_bend_elbows"]:
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("הרם מעט את המרפקים")

                    else:
                        self.comments.append("הרימי מעט את המרפקים")
                else:
                    if s.gender == "Male":
                        self.comments.append("הצמד את המרפקים לגוף")

                    else:
                        self.comments.append("הצמידי את המרפקים לגוף")

            if s.req_exercise in ["ball_raise_arms_above_head", "ball_open_arms_and_forward", "band_open_arms_and_up",\
                                  "band_straighten_left_arm_elbows_bend_to_sides", "band_straighten_right_arm_elbows_bend_to_sides",\
                                  "stick_raise_arms_above_head", "weights_open_arms_and_forward", "weights_abduction"]:
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("הרם את הידיים יותר למעלה")

                    else:
                        self.comments.append("הרימי את הידיים יותר למעלה")

                else:
                    if s.gender == "Male":
                        self.comments.append("הורד יותר את הידיים")

                    else:
                        self.comments.append("הורידי יותר את הידיים")


            if s.req_exercise == "stick_bend_elbows_and_up":
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("הרם את הידיים יותר למעלה")

                    else:
                        self.comments.append("הרימי את הידיים יותר למעלה")
                else:
                    if s.gender == "Male":
                        self.comments.append("הצמד את המרפקים לגוף")

                    else:
                        self.comments.append("הצמידי את המרפקים לגוף")


        if joint1 == "L_wrist" and joint2 == "L_hip" and joint3 == "R_hip":
            if biggerORsmaller == "smaller":
                if side is not None:
                    if side == "left":
                        if s.gender == "Male":
                            self.comments.append("סובב יותר את הגב לצד שמאל \n יחד עם הכתפיים")

                        else:
                            self.comments.append("סובבי יותר את הגב לצד שמאל \n יחד עם הכתפיים")
                    else:
                        if s.gender == "Male":
                            self.comments.append("סובב פחות את הגב לצד ימין")

                        else:
                            self.comments.append("סובבי פחות את הגב לצד ימין")

            if biggerORsmaller == "bigger":
                if side is not None:
                    if side == "left":
                        if s.gender == "Male":
                            self.comments.append("סובב פחות את הגב לצד שמאל")

                        else:
                            self.comments.append("סובבי פחות את הגב לצד שמאל")
                    else:
                        if s.gender == "Male":
                            self.comments.append("סובב יותר את הגב לצד ימין \n יחד עם הכתפיים")

                        else:
                            self.comments.append("סובבי יותר את הגב לצד ימין \n יחד עם הכתפיים")


        if joint1 == "R_wrist" and joint2 == "R_hip" and joint3 == "L_hip":
            if biggerORsmaller == "smaller":
                if side is not None:
                    if side == "left":
                        if s.gender == "Male":
                            self.comments.append("סובב פחות את הגב לצד שמאל")

                        else:
                            self.comments.append("סובב פחות את הגב לצד שמאל")
                    else:
                        if s.gender == "Male":
                            self.comments.append("סובב יותר את הגב לצד ימין \n יחד עם הכתפיים")

                        else:
                            self.comments.append("סובבי יותר את הגב לצד ימין \n יחד עם הכתפיים")

            if biggerORsmaller == "bigger":
                if side is not None:
                    if side == "left":
                        if s.gender == "Male":
                            self.comments.append("סובב יותר את הגב לצד שמאל \n יחד עם הכתפיים")

                        else:
                            self.comments.append("סובבי יותר את הגב לצד שמאל \n יחד עם הכתפיים")
                    else:
                        if s.gender == "Male":
                            self.comments.append("סובב פחות את הגב לצד ימין")

                        else:
                            self.comments.append("סובבי פחות את הגב לצד ימין")



        if joint1 == "L_hip" and joint2 == "L_shoulder" and joint3 == "L_wrist" and s.req_exercise == "notool_left_hand_up_and_bend":
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("התיישר יותר לצד שמאל והבא \n את יד שמאל מעל כתף שמאל")

                    else:
                        self.comments.append("התיישרי יותר לצד שמאל והביאי \n את יד שמאל מעל כתף שמאל")
                else:
                    if s.gender == "Male":
                        self.comments.append("הישען יותר ושלח את יד שמאל \n יותר לכיוון הרצפה")

                    else:
                        self.comments.append("הישעני יותר ושלחי את יד שמאל \n יותר לכיוון הרצפה")


        if joint1 == "R_hip" and joint2 == "R_shoulder" and joint3 == "R_wrist" and s.req_exercise == "notool_right_hand_up_and_bend":
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("התיישר יותר לצד ימין והבא \n את יד ימין מעל כתף ימין")

                    else:
                        self.comments.append("התיישרי יותר לצד ימין והביאי \n את יד ימין מעל כתף ימין")
                else:
                    if s.gender == "Male":
                        self.comments.append("הישען יותר ושלח את יד ימין \n יותר לכיוון הרצפה")

                    else:
                        self.comments.append("הישעני יותר ושלחי את יד ימין \n יותר לכיוון הרצפה")


        if cleaned_joint_1 == "hip" and cleaned_joint_2 == "shoulder" and cleaned_joint_3 == "wrist" and s.req_exercise in ["band_open_arms" ,"notool_raising_hands_diagonally"]:
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("הרם את הידיים יותר למעלה")

                    else:
                        self.comments.append("הרימי את הידיים יותר למעלה")
                else:
                    if s.gender == "Male":
                        self.comments.append("הורד יותר את הידיים")

                    else:
                        self.comments.append("הורידי יותר את הידיים")



        if (cleaned_joint_1 == "wrist" or cleaned_joint_1 == "elbow") and cleaned_joint_2 == "shoulder" and cleaned_joint_3 == "shoulder":
            if s.req_exercise in ["band_open_arms", "band_open_arms_and_up"]:
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("מתח יותר את הגומייה")

                    else:
                        self.comments.append("מתחי יותר את הגומייה")

                else:
                    if s.gender == "Male":
                        self.comments.append("שחרר יותר את הגומייה")

                    else:
                        self.comments.append("שחררי יותר את הגומייה")

            else:
                if biggerORsmaller == "smaller":
                    if s.gender == "Male":
                        self.comments.append("פתח יותר את הידיים")

                    else:
                        self.comments.append("פתחי יותר את הידיים")

                else:
                    if s.gender == "Male":
                        self.comments.append("סגור יותר את הידיים")

                    else:
                        self.comments.append("סגרי יותר את הידיים")


        if joint1 == "L_elbow" and joint2 == "L_hip" and joint3 == "R_hip":
            if biggerORsmaller == "smaller":
                if side is not None:
                    if side == "left":
                        if s.gender == "Male":
                            self.comments.append("מתח את יד שמאל עוד למטה")

                        else:
                            self.comments.append("מתחי את יד שמאל עוד למטה")
                    else:
                        if s.gender == "Male":
                            self.comments.append("מתח את יד שמאל עוד למעלה")

                        else:
                            self.comments.append("מתחי את יד שמאל עוד למעלה")


            if biggerORsmaller == "bigger":
                if side is not None:
                    if side == "left":
                        if s.gender == "Male":
                            self.comments.append("מתח פחות את יד שמאל")

                        else:
                            self.comments.append("מתחי פחות את יד שמאל")

                    else:
                        if s.gender == "Male":
                            self.comments.append("הבא את יד שמאל \n  יותר לכיוון הרצפה")

                        else:
                            self.comments.append("הביאי את יד שמאל \n  יותר לכיוון הרצפה")


        if joint1 == "R_elbow" and joint2 == "R_hip" and joint3 == "L_hip":
            if biggerORsmaller == "smaller":
                if side is not None:
                    if side == "left":
                        if s.gender == "Male":
                            self.comments.append("מתח את יד ימין עוד למעלה")

                        else:
                            self.comments.append("מתחי את יד ימין עוד למעלה")

                    else:
                        if s.gender == "Male":
                            self.comments.append("מתח את יד ימין עוד למטה")

                        else:
                            self.comments.append("מתחי את יד ימין עוד למטה")

            if biggerORsmaller == "bigger":
                if side is not None:
                    if side == "left":
                        if s.gender == "Male":
                            self.comments.append("הבא את יד ימין \n יותר לכיוון הרצפה")

                        else:
                            self.comments.append("הביאי את יד ימין \n יותר לכיוון הרצפה")

                    else:
                        if s.gender == "Male":
                            self.comments.append("מתח פחות את יד ימין")

                        else:
                            self.comments.append("מתחי פחות את יד ימין")



# class GraphPage(tk.Frame):
#     def __init__(self, master, exercise, previous, **kwargs):
#         tk.Frame.__init__(self, master, **kwargs)
#         self.queue = queue.Queue()
#         self.exercise = exercise
#         self.forward_arrow_button = None
#         self.backward_arrow_button = None
#         self.label1=None
#         self.label2= None
#         image = Image.open('Pictures//background.jpg')
#         self.photo_image = ImageTk.PhotoImage(image)
#         tk.Label(self, image=self.photo_image).pack()
#
        # if previous=="ball":
        #     previous_page=ChooseBallExercisesPage
        # elif previous=="band":
        #     previous_page=ChooseRubberBandExercisesPage
        # elif previous=="stick":
        #     previous_page=ChooseStickExercisesPage
        # elif previous=="no_tool":
        #     previous_page=ChooseNoToolExercisesPage
        #
        # previous_page_button_img = Image.open("Pictures//previous.jpg")
        # previous_page_button_photo = ImageTk.PhotoImage(previous_page_button_img)
        # previous_page_category = tk.Button(self, image=previous_page_button_photo,
        #                                   command=lambda: s.screen.switch_frame(previous_page),
        #                                   width=previous_page_button_img.width,
        #                                   height=previous_page_button_img.height, bd=0,
        #                                   highlightthickness=0)
        # previous_page_category.image = previous_page_button_photo
        # previous_page_category.place(x=30, y=30)
#
#
#         sorted_folders= get_sorted_folders(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}')
#         if len(sorted_folders)==0:
#             did_before = Image.open('Pictures//patient_didnt_do.jpg')
#             self.did_before = ImageTk.PhotoImage(did_before)
#             self.did_before_label = tk.Label(self, image=self.did_before, bd=0)
#             self.did_before_label.place(x=270, y=75)
#
#         else:
#             num_of_angles= self.get_number_of_angles_in_exercise(exercise)
#             self.show_graphs(sorted_folders, 0, num_of_angles, exercise)
#
#     def show_graphs(self, sorted_folder, place, num_of_angles, exercise):
#         if place>0:
#             self.backward_arrow_button_img = Image.open("Pictures//previous_arrow.jpg")
#             self.backward_arrow_button_photo = ImageTk.PhotoImage(self.backward_arrow_button_img)
#             self.backward_arrow_button = tk.Button(self, image=self.backward_arrow_button_photo,
#                                               command=lambda: self.help_function(sorted_folder, place-1, num_of_angles, exercise),
#                                               width=self.backward_arrow_button_img.width,
#                                               height=self.backward_arrow_button_img.height, bd=0,
#                                               highlightthickness=0)
#             self.backward_arrow_button.image = self.backward_arrow_button_photo
#             self.backward_arrow_button.place(x=945, y=480)
#
#         if place<len(sorted_folder)-1:
#             self.forward_arrow_button_img = Image.open("Pictures//forward_arrow.jpg")
#             self.forward_arrow_button_photo = ImageTk.PhotoImage(self.forward_arrow_button_img)
#             self.forward_arrow_button = tk.Button(self, image=self.forward_arrow_button_photo,
#                                              command=lambda: self.help_function(sorted_folder, place+1, num_of_angles, exercise),
#                                              width=self.forward_arrow_button_img.width,
#                                              height=self.forward_arrow_button_img.height, bd=0,
#                                              highlightthickness=0)
#             self.forward_arrow_button.image = self.forward_arrow_button_photo
#             self.forward_arrow_button.place(x=20, y=480)
#
#
#         back = Image.open('Pictures//empty.JPG')
#         background_img = ImageTk.PhotoImage(back)
#
#
#         directory= f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/{sorted_folder[place]}.xlsx'
#         print(directory)
#
#         success_number = Excel.get_success_number(directory, exercise)
#         #effort_number = Excel.get_effort_number(directory, exercise)
#
#
#         self.label1 = tk.Label(self, text=f'{sorted_folder[place]}',
#                               image=background_img, compound=tk.CENTER, font=("Thaoma", 20, 'bold'), width=350,
#                               height=30)
#         self.label1.place(x=160, y=15)
#         self.label1.image = background_img
#
#         if success_number is not None:
#             self.label2 = tk.Label(self, text= "מספר חזרות מוצלחות: "+ str(success_number),
#                                   image=background_img, compound=tk.CENTER, font=("Thaoma", 11, 'bold'), width=350,
#                                   height=30)
#             self.label2.place(x=155, y=50)
#             self.label2.image = background_img
#
#         # if effort_number is not None:
#         #     self.label = tk.Label(self,
#         #                           text=str(effort_number) + "דירוג קושי התרגיל על ידי המתאמן: ",
#         #                           image=background_img, compound=tk.CENTER, font=("Thaoma", 11, 'bold'), width=350,
#         #                           height=30)
#         #     self.label.place(x=155, y=40)
#         #     self.label.image = background_img
#
#
#         if num_of_angles == 1:
#             self.one_angle_graph(exercise, sorted_folder[place])
#         if num_of_angles == 2:
#             self.two_angles_graph(exercise, sorted_folder[place])
#         if num_of_angles == 3:
#             self.three_angles_graph(exercise, sorted_folder[place])
#
#     def help_function(self, sorted_folder, place_to_put, num_of_angles, exercise):
#         if self.label1:
#             self.label1.place_forget()
#         if self.label2:
#             self.label2.place_forget()
#         if self.forward_arrow_button:
#             self.forward_arrow_button.destroy()
#             self.forward_arrow_button = None
#         if self.backward_arrow_button:
#             self.backward_arrow_button.destroy()
#             self.backward_arrow_button = None
#         self.show_graphs(sorted_folder, place_to_put, num_of_angles, exercise)
#
#     def find_image(self, directory, number):
#         # Create a regex pattern to match the files with ' ' followed by the specific number and '.jpeg'
#         pattern = re.compile(r' (\d+)\.jpeg$')
#
#         # Iterate through the files in the directory
#         for filename in os.listdir(directory):
#             match = pattern.search(filename)
#             if match and match.group(1) == str(number):
#                 return os.path.join(directory, filename)
#         return None
#
#
#
#     def get_number_of_angles_in_exercise(self, exercise):
#         try:
#             # Load the workbook
#             workbook = openpyxl.load_workbook("exercises_table.xlsx")
#
#             # Select the desired sheet
#             sheet = workbook[workbook.sheetnames[0]]
#
#             # Iterate through rows starting from the specified row
#             for row_number in range(1,sheet.max_row + 1):
#                 first_cell_value = sheet.cell(row=row_number, column=1).value
#
#                 if first_cell_value == exercise:
#                     return sheet.cell(row=row_number, column=2).value
#
#
#         except Exception as e:
#             print(f"An error occurred: {e}")
#             return False
#
#     def three_angles_graph(self, exercise, folder):
#         # Determine the resize factor
#         resize_factor = 0.42
#
#         # Load the image
#         dir1= self.find_image(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{folder}', 1)
#         graph1 = Image.open(dir1)
#         new_width1 = int(graph1.width * resize_factor)
#         new_height1 = int(graph1.height * resize_factor)
#         graph1_resized = graph1.resize((new_width1, new_height1), Image.Resampling.LANCZOS)
#         self.graph1 = ImageTk.PhotoImage(graph1_resized)
#         self.graph1_label = tk.Label(self, image=self.graph1, bd=0)
#         self.graph1_label.place(x=100, y=90)
#
#         dir2 = self.find_image(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{folder}', 2)
#         graph2 = Image.open(dir2)
#         new_width2 = int(graph2.width * resize_factor)
#         new_height2 = int(graph2.height * resize_factor)
#         graph2_resized = graph2.resize((new_width2, new_height2), Image.Resampling.LANCZOS)
#         self.graph2 = ImageTk.PhotoImage(graph2_resized)
#         self.graph2_label = tk.Label(self, image=self.graph2, bd=0)
#         self.graph2_label.place(x=100, y=325)
#
#         dir3 = self.find_image(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{folder}', 3)
#         graph3 = Image.open(dir3)
#         new_width3 = int(graph3.width * resize_factor)
#         new_height3 = int(graph3.height * resize_factor)
#         graph3_resized = graph3.resize((new_width3, new_height3), Image.Resampling.LANCZOS)
#         self.graph3 = ImageTk.PhotoImage(graph3_resized)
#         self.graph3_label = tk.Label(self, image=self.graph3, bd=0)
#         self.graph3_label.place(x=380, y=90)
#
#         dir4 = self.find_image(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{folder}', 4)
#         graph4 = Image.open(dir4)
#         new_width4 = int(graph4.width * resize_factor)
#         new_height4 = int(graph4.height * resize_factor)
#         graph4_resized = graph4.resize((new_width4, new_height4), Image.Resampling.LANCZOS)
#         self.graph4 = ImageTk.PhotoImage(graph4_resized)
#         self.graph4_label = tk.Label(self, image=self.graph4, bd=0)
#         self.graph4_label.place(x=380, y=325)
#
#         dir5 = self.find_image(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{folder}', 5)
#         graph5 = Image.open(dir5)
#         new_width5 = int(graph5.width * resize_factor)
#         new_height5 = int(graph5.height * resize_factor)
#         graph5_resized = graph5.resize((new_width5, new_height5), Image.Resampling.LANCZOS)
#         self.graph5 = ImageTk.PhotoImage(graph5_resized)
#         self.graph5_label = tk.Label(self, image=self.graph5, bd=0)
#         self.graph5_label.place(x=660, y=90)
#
#         dir6 = self.find_image(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{folder}', 6)
#         graph6 = Image.open(dir6)
#         new_width6 = int(graph6.width * resize_factor)
#         new_height6 = int(graph6.height * resize_factor)
#         graph6_resized = graph6.resize((new_width6, new_height6), Image.Resampling.LANCZOS)
#         self.graph6 = ImageTk.PhotoImage(graph6_resized)
#         self.graph6_label = tk.Label(self, image=self.graph6, bd=0)
#         self.graph6_label.place(x=660, y=325)
#
#
#     def two_angles_graph(self, exercise, folder):
#         # Determine the resize factor
#         resize_factor = 0.45
#
#         dir1 = self.find_image(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{folder}', 1)
#         graph1 = Image.open(dir1)
#         new_width1 = int(graph1.width * resize_factor)
#         new_height1 = int(graph1.height * resize_factor)
#         graph1_resized = graph1.resize((new_width1, new_height1), Image.Resampling.LANCZOS)
#         self.graph1 = ImageTk.PhotoImage(graph1_resized)
#         self.graph1_label = tk.Label(self, image=self.graph1, bd=0)
#         self.graph1_label.place(x=190, y=90)
#
#         dir2 = self.find_image(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{folder}', 2)
#         graph2 = Image.open(dir2)
#         new_width2 = int(graph2.width * resize_factor)
#         new_height2 = int(graph2.height * resize_factor)
#         graph2_resized = graph2.resize((new_width2, new_height2), Image.Resampling.LANCZOS)
#         self.graph2 = ImageTk.PhotoImage(graph2_resized)
#         self.graph2_label = tk.Label(self, image=self.graph2, bd=0)
#         self.graph2_label.place(x=190, y=325)
#
#         dir3 = self.find_image(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{folder}', 3)
#         graph3 = Image.open(dir3)
#         new_width3 = int(graph3.width * resize_factor)
#         new_height3 = int(graph3.height * resize_factor)
#         graph3_resized = graph3.resize((new_width3, new_height3), Image.Resampling.LANCZOS)
#         self.graph3 = ImageTk.PhotoImage(graph3_resized)
#         self.graph3_label = tk.Label(self, image=self.graph3, bd=0)
#         self.graph3_label.place(x=540, y=90)
#
#         dir4 = self.find_image(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{folder}', 4)
#         graph4 = Image.open(dir4)
#         new_width4 = int(graph4.width * resize_factor)
#         new_height4 = int(graph4.height * resize_factor)
#         graph4_resized = graph4.resize((new_width4, new_height4), Image.Resampling.LANCZOS)
#         self.graph4 = ImageTk.PhotoImage(graph4_resized)
#         self.graph4_label = tk.Label(self, image=self.graph4, bd=0)
#         self.graph4_label.place(x=540, y=325)
#
#
#
#     def one_angle_graph(self, exercise, folder):
#         # Determine the resize factor
#         resize_factor = 0.45
#
#         dir1 = self.find_image(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{folder}', 1)
#         graph1 = Image.open(dir1)
#         new_width1 = int(graph1.width * resize_factor)
#         new_height1 = int(graph1.height * resize_factor)
#         graph1_resized = graph1.resize((new_width1, new_height1), Image.Resampling.LANCZOS)
#         self.graph1 = ImageTk.PhotoImage(graph1_resized)
#         self.graph1_label = tk.Label(self, image=self.graph1, bd=0)
#         self.graph1_label.place(x=350, y=90)
#
#         dir2 = self.find_image(f'C:/Users/yaels/יעל פרוייקט גמר/zedcheck/Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{folder}', 2)
#         graph2 = Image.open(dir2)
#         new_width2 = int(graph2.width * resize_factor)
#         new_height2 = int(graph2.height * resize_factor)
#         graph2_resized = graph2.resize((new_width2, new_height2), Image.Resampling.LANCZOS)
#         self.graph2 = ImageTk.PhotoImage(graph2_resized)
#         self.graph2_label = tk.Label(self, image=self.graph2, bd=0)
#         self.graph2_label.place(x=350, y=325)




class TablesPage(tk.Frame):
    def __init__(self, master, exercise, previous, **kwargs):
        tk.Frame.__init__(self, master, **kwargs)
        self.queue = queue.Queue()
        self.exercise = exercise
        self.forward_arrow_button = None
        self.backward_arrow_button = None
        self.label1 = None
        self.label2 = None
        self.background_color = "#deeaf7"  # Set the background color to light blue

        # Set the background color for the Frame
        self.configure(bg=self.background_color)

        # Load background image
        image = Image.open('Pictures//background.jpg')
        self.photo_image = ImageTk.PhotoImage(image)

        # Add a label with the background image and set the background to match the window
        tk.Label(self, image=self.photo_image, bg=self.background_color).pack(fill='both', expand=True)

        # Define previous page navigation based on tool used
        if previous == "ball":
            previous_page = ChooseBallExercisesPage
        elif previous == "band":
            previous_page = ChooseRubberBandExercisesPage
        elif previous == "stick":
            previous_page = ChooseStickExercisesPage
        elif previous == "weights":
            previous_page = ChooseWeightsExercisesPage
        elif previous == "no_tool":
            previous_page = ChooseNoToolExercisesPage

        # Load the previous page button image and set background
        previous_page_button_img = Image.open("Pictures//previous.jpg")
        previous_page_button_photo = ImageTk.PhotoImage(previous_page_button_img)
        previous_page_category = tk.Button(self, image=previous_page_button_photo,
                                           command=lambda: s.screen.switch_frame(previous_page),
                                           width=previous_page_button_img.width,
                                           height=previous_page_button_img.height, bd=0,
                                           highlightthickness=0, bg=self.background_color)
        previous_page_category.image = previous_page_button_photo
        previous_page_category.place(x=30, y=30)

        # Load the previous page button image and set background
        PDF_button_img = Image.open("Pictures//see_PDF_button.jpg")
        PDF_button_photo = ImageTk.PhotoImage(PDF_button_img)
        self.PDF_button = tk.Button(self, image=PDF_button_photo,
                                           command=lambda: self.open_pdf(),
                                           width=PDF_button_img.width,
                                           height=PDF_button_img.height, bd=0,
                                           highlightthickness=0, bg=self.background_color)
        self.PDF_button.image =PDF_button_photo
        self.PDF_button.place(x=392, y=525)
        self.PDF_button.lift()  # Raises the button above all other widgets

        try:
            # Attempt to fetch sorted folders for the exercise
            sorted_folders = get_sorted_folders(f'Patients/{s.chosen_patient_ID}/Tables/{exercise}')

            if len(sorted_folders) == 0:
                # Display 'patient_didnt_do.jpg' if no folders are found
                didnt_do_before = Image.open('Pictures//patient_didnt_do.jpg')
                self.didnt_do_before = ImageTk.PhotoImage(didnt_do_before)
                self.didnt_do_before_label = tk.Label(self, image=self.didnt_do_before, bd=0, bg=self.background_color)
                self.didnt_do_before_label.place(x=270, y=75)
            else:
                # Show tables if folders are found
                num_of_angles = self.get_number_of_angles_in_exercise(exercise)
                self.show_tables(sorted_folders, 0, num_of_angles, exercise)

        except FileNotFoundError:
            # Handle case where the path does not exist
            didnt_do_before = Image.open('Pictures//patient_didnt_do.jpg')
            self.didnt_do_before = ImageTk.PhotoImage(didnt_do_before)
            self.didnt_do_before_label = tk.Label(self, image=self.didnt_do_before, bd=0, bg=self.background_color)
            self.didnt_do_before_label.place(x=270, y=75)
            print(f"Error: The path 'Patients/{s.chosen_patient_ID}/Tables/{exercise}' does not exist.")

    def open_pdf(self):
        """Opens the PDF related to the selected exercise and date."""
        try:
            # Get absolute path to the 'Patients' directory
            base_dir = os.path.abspath("Patients")

            # Construct the full PDF path
            pdf_path = os.path.join(base_dir, str(s.chosen_patient_ID), "PDF_to_Therapist_Email",
                                    f"{self.sorted_folder[self.place]}.pdf")

            # Debugging: Print the absolute path
            print(f"Checking file path: {pdf_path}")

            # Check if the file exists
            if not os.path.exists(pdf_path):
                print(f"Error: PDF file not found at {pdf_path}")
                return

            # Open PDF with the default viewer based on OS
            if sys.platform.startswith('win'):
                os.startfile(pdf_path)  # Windows
            elif sys.platform.startswith('darwin'):  # macOS
                subprocess.run(['open', pdf_path], check=True)
            elif sys.platform.startswith('linux'):
                subprocess.run(['xdg-open', pdf_path], check=True)
            else:
                webbrowser.open(pdf_path)  # Fallback (if other methods fail)

        except Exception as e:
            print(f"Error opening PDF: {e}")


    def show_tables(self, sorted_folder, place, num_of_angles, exercise):
        self.sorted_folder = sorted_folder
        self.place = place

        if place > 0:
            self.backward_arrow_button_img = Image.open("Pictures//previous_arrow.jpg")
            self.backward_arrow_button_photo = ImageTk.PhotoImage(self.backward_arrow_button_img)
            self.backward_arrow_button = tk.Button(self, image=self.backward_arrow_button_photo,
                                                   command=lambda: self.help_function(sorted_folder, place-1, num_of_angles, exercise),
                                                   width=self.backward_arrow_button_img.width,
                                                   height=self.backward_arrow_button_img.height, bd=0,
                                                   highlightthickness=0, bg=self.background_color)
            self.backward_arrow_button.image = self.backward_arrow_button_photo
            self.backward_arrow_button.place(x=945, y=480)

        if place < len(sorted_folder) - 1:
            self.forward_arrow_button_img = Image.open("Pictures//forward_arrow.jpg")
            self.forward_arrow_button_photo = ImageTk.PhotoImage(self.forward_arrow_button_img)
            self.forward_arrow_button = tk.Button(self, image=self.forward_arrow_button_photo,
                                                  command=lambda: self.help_function(sorted_folder, place+1, num_of_angles, exercise),
                                                  width=self.forward_arrow_button_img.width,
                                                  height=self.forward_arrow_button_img.height, bd=0,
                                                  highlightthickness=0, bg=self.background_color)
            self.forward_arrow_button.image = self.forward_arrow_button_photo
            self.forward_arrow_button.place(x=20, y=480)

        back = Image.open('Pictures//empty.JPG')
        background_img = ImageTk.PhotoImage(back)

        # Directory path
        directory = f'Patients/{s.chosen_patient_ID}/{sorted_folder[place]}.xlsx'
        print(directory)

        # Fetch success numbers



        self.label1 = tk.Label(self, text=f'{str(s.chosen_patient_ID)}', image=background_img, compound=tk.CENTER,
                               font=("Thaoma", 16, 'bold'), width=350, height=15, bg=self.background_color)
        self.label1.place(x=160, y=5)
        self.label1.image = background_img

        success_number = Excel.get_success_number(directory, exercise)

        if success_number is not None:
            self.label3 = tk.Label(self, text="מספר חזרות מוצלחות: " + str(success_number), image=background_img,
                                   compound=tk.CENTER, font=("Thaoma", 16, 'bold'), width=350,
                                   height=15, bg=self.background_color)
            self.label3.place(x=155, y=55)
            self.label3.image = background_img

        # Handle different number of angles
        if num_of_angles == 1:
            self.one_angle_graph(exercise, sorted_folder[place])
        if num_of_angles == 2:
            self.two_angles_graph(exercise, sorted_folder[place])
        if num_of_angles == 3:
            self.three_angles_graph(exercise, sorted_folder[place])

        formatted_text = datetime.strptime(sorted_folder[place], "%d-%m-%Y %H-%M-%S").strftime("%d/%m/%Y %H:%M:%S")

        # Display date and time with custom background color
        self.label2 = tk.Label(self, text=f'{formatted_text}', image=background_img, compound=tk.CENTER,
                               font=("Thaoma", 16, 'bold'), width=350, height=15, bg=self.background_color)
        self.label2.place(x=160, y=30)
        self.label2.image = background_img

        self.PDF_button.lift()

    def help_function(self, sorted_folder, place_to_put, num_of_angles, exercise):
        # Clean up previous labels and buttons
        if self.label1:
            self.label1.place_forget()
        if self.label2:
            self.label2.place_forget()
        if self.forward_arrow_button:
            self.forward_arrow_button.destroy()
            self.forward_arrow_button = None
        if self.backward_arrow_button:
            self.backward_arrow_button.destroy()
            self.backward_arrow_button = None
        self.show_tables(sorted_folder, place_to_put, num_of_angles, exercise)

    def find_image(self, directory, number):
        # Create a regex pattern to match files with ' ' followed by the specific number and either '.jpeg', '.jpg', or '.png'
        pattern = re.compile(r' (\d+)\.(?:jpeg|jpg|png)$')

        # Iterate through the files in the directory
        for filename in os.listdir(directory):
            match = pattern.search(filename)
            if match and match.group(1) == str(number):
                return os.path.join(directory, filename)
        return None



    def get_number_of_angles_in_exercise(self, exercise):
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook("exercises_table.xlsx")

            # Select the desired sheet
            sheet = workbook[workbook.sheetnames[0]]

            # Iterate through rows starting from the specified row
            for row_number in range(1,sheet.max_row + 1):
                first_cell_value = sheet.cell(row=row_number, column=2).value

                if first_cell_value == exercise:
                    return sheet.cell(row=row_number, column=3).value


        except Exception as e:
            print(f"An error occurred: {e}")
            return False

    def three_angles_graph(self, exercise, folder):
        try:

            new_width = 270
            new_height = 220

            dir1 = self.find_image(
                f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{folder}', 1)
            table1 = Image.open(dir1)
            dir2 = self.find_image(
                f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{folder}', 2)
            table2 = Image.open(dir2)
            dir3 = self.find_image(
                f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{folder}', 3)
            table3 = Image.open(dir3)
            dir4 = self.find_image(
                f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{folder}', 4)
            table4 = Image.open(dir4)
            dir5 = self.find_image(
                f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{folder}', 5)
            table5 = Image.open(dir5)
            dir6 = self.find_image(
                f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{folder}', 6)
            table6 = Image.open(dir6)

            # Load the image for table 1


            table1_resized = table1.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.table1 = ImageTk.PhotoImage(table1_resized)
            self.table1_label = tk.Label(self, image=self.table1, bd=0, bg='#deeaf7')  # Set background color here
            self.table1_label.place(x=95, y=90)

            # Load the image for table 2

            table2_resized = table2.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.table2 = ImageTk.PhotoImage(table2_resized)
            self.table2_label = tk.Label(self, image=self.table2, bd=0, bg='#deeaf7')  # Set background color here
            self.table2_label.place(x=95, y=325)

            # Load the image for table 3


            table3_resized = table3.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.table3 = ImageTk.PhotoImage(table3_resized)
            self.table3_label = tk.Label(self, image=self.table3, bd=0, bg='#deeaf7')  # Set background color here
            self.table3_label.place(x=375, y=90)

            # Load the image for table 4


            table4_resized = table4.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.table4 = ImageTk.PhotoImage(table4_resized)
            self.table4_label = tk.Label(self, image=self.table4, bd=0, bg='#deeaf7')  # Set background color here
            self.table4_label.place(x=375, y=325)

            # Load the image for table 5

            table5_resized = table5.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.table5 = ImageTk.PhotoImage(table5_resized)
            self.table5_label = tk.Label(self, image=self.table5, bd=0, bg='#deeaf7')  # Set background color here
            self.table5_label.place(x=655, y=90)

            # Load the image for table 6

            table6_resized = table6.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.table6 = ImageTk.PhotoImage(table6_resized)
            self.table6_label = tk.Label(self, image=self.table6, bd=0, bg='#deeaf7')  # Set background color here
            self.table6_label.place(x=655, y=325)

        except FileNotFoundError:
            # Handle case where the path does not exist
            didnt_do_before = Image.open('Pictures//patient_didnt_do.jpg')
            self.didnt_do_before = ImageTk.PhotoImage(didnt_do_before)
            self.didnt_do_before_label = tk.Label(self, image=self.didnt_do_before, bd=0, bg=self.background_color)
            self.didnt_do_before_label.place(x=270, y=75)
            print(f"Error: The path 'Patients/{s.chosen_patient_ID}/Tables/{exercise}' does not exist.")


    def two_angles_graph(self, exercise, folder):
        try:
            new_width= 350
            new_height = 220

            dir1 = self.find_image(
                f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{folder}', 1)
            table1 = Image.open(dir1)
            dir2 = self.find_image(
                f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{folder}', 2)
            table2 = Image.open(dir2)
            dir3 = self.find_image(
                f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{folder}', 3)
            table3 = Image.open(dir3)
            dir4 = self.find_image(
                f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{folder}', 4)
            table4 = Image.open(dir4)

            # Load the image for table 1


            table1_resized = table1.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.table1 = ImageTk.PhotoImage(table1_resized)
            self.table1_label = tk.Label(self, image=self.table1, bd=0, bg='#deeaf7')  # Set background color here
            self.table1_label.place(x=120, y=90)

            # Load the image for table 2


            table2_resized = table2.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.table2 = ImageTk.PhotoImage(table2_resized)
            self.table2_label = tk.Label(self, image=self.table2, bd=0, bg='#deeaf7')  # Set background color here
            self.table2_label.place(x=120, y=325)

            # Load the image for table 3

            table3_resized = table3.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.table3 = ImageTk.PhotoImage(table3_resized)
            self.table3_label = tk.Label(self, image=self.table3, bd=0, bg='#deeaf7')  # Set background color here
            self.table3_label.place(x=550, y=90)

            # Load the image for table 4


            table4_resized = table4.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.table4 = ImageTk.PhotoImage(table4_resized)
            self.table4_label = tk.Label(self, image=self.table4, bd=0, bg='#deeaf7')  # Set background color here
            self.table4_label.place(x=550, y=325)

        except FileNotFoundError:
            # Handle case where the path does not exist
            didnt_do_before = Image.open('Pictures//patient_didnt_do.jpg')
            self.didnt_do_before = ImageTk.PhotoImage(didnt_do_before)
            self.didnt_do_before_label = tk.Label(self, image=self.didnt_do_before, bd=0, bg=self.background_color)
            self.didnt_do_before_label.place(x=270, y=75)
            print(f"Error: The path 'Patients/{s.chosen_patient_ID}/Tables/{exercise}' does not exist.")




    def one_angle_graph(self, exercise, folder):
        try:

            new_width = 350
            new_height = 220

            dir1 = self.find_image(f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{folder}', 1)
            table1 = Image.open(dir1)
            dir2 = self.find_image(f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{folder}', 2)
            table2 = Image.open(dir2)
            # Load the image for table 1


            table1_resized = table1.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.table1 = ImageTk.PhotoImage(table1_resized)
            self.table1_label = tk.Label(self, image=self.table1, bd=0, bg='#deeaf7')  # Set background color here
            self.table1_label.place(x=300, y=90)

            # Load the image for table 2

            table2_resized = table2.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.table2 = ImageTk.PhotoImage(table2_resized)
            self.table2_label = tk.Label(self, image=self.table2, bd=0, bg='#deeaf7')  # Set background color here
            self.table2_label.place(x=300, y=325)

        except FileNotFoundError:
            # Handle case where the path does not exist
            didnt_do_before = Image.open('Pictures//patient_didnt_do.jpg')
            self.didnt_do_before = ImageTk.PhotoImage(didnt_do_before)
            self.didnt_do_before_label = tk.Label(self, image=self.didnt_do_before, bd=0, bg=self.background_color)
            self.didnt_do_before_label.place(x=270, y=75)
            print(f"Error: The path 'Patients/{s.chosen_patient_ID}/Tables/{exercise}' does not exist.")

def add_to_exercise_page(self, page_name):
    name_label(self, 350, 350)
    how_many_repetitions_of_exercises(self)
    rate_of_exercises(self)
    page_name_label(self,page_name)


def name_label(self, width= None, place_x= None):
    back = Image.open('Pictures//empty.JPG')
    background_img = ImageTk.PhotoImage(back)
    self.background_color = "#deeaf7"  # Set the background color to light blue


    if width is None:
        self.label1 = tk.Label(self, text=f'מספר מזהה של מטופל: {str(s.chosen_patient_ID)}', image=background_img,
                               compound=tk.CENTER,
                               font=("Ariel", 16, 'bold'), width=350, height=30, bg=self.background_color)
    else:
        self.label1 = tk.Label(self, text=f'מספר מזהה של מטופל: {str(s.chosen_patient_ID)}', image=background_img,
                               compound=tk.CENTER,
                               font=("Ariel", 16, 'bold'), width=width, height=30, bg=self.background_color)

    if place_x is None:
        self.label1.place(x=150, y=45)

    else:
        self.label1.place(x=place_x, y=45)

    self.label1.image = background_img


def page_name_label(self, page_name):
    back = Image.open('Pictures//empty.JPG')
    background_img = ImageTk.PhotoImage(back)
    self.background_color = "#deeaf7"  # Set the background color to light blue

    self.label1 = tk.Label(self, text=page_name, image=background_img,
                           compound=tk.CENTER,
                           font=("Thaoma", 26, 'bold'), width=350, height=30, bg=self.background_color)
    self.label1.place(x=350, y=12)
    self.label1.image = background_img


def how_many_repetitions_of_exercises(self):
    # List of options for the dropdown
    self.options = ["5", "6", "7", "8", "9", "10"]

    # Create a StringVar to hold the selected value
    self.selected_option_rep = tk.StringVar()

    # Get the current number of repetitions from Excel
    current_number_of_repetitions = Excel.find_value_by_colName_and_userID(
        "Patients.xlsx", "patients_details", s.chosen_patient_ID, "number of repetitions in each exercise"
    )

    # Set the default selected option based on the value from Excel
    self.selected_option_rep.set(self.options[int(current_number_of_repetitions) - 5])  # Subtract 5 to match the index

    # Create a custom style for the OptionMenu
    style = ttk.Style()
    style.theme_use('clam')  # Set theme to 'clam'

    # Configure the appearance of the OptionMenu and the dropdown items
    style.configure('Custom.TMenubutton', font=('Arial', 16, 'bold'), background='lightgray', foreground='black',
                    padding=[10, 10], anchor='center')  # Add padding and set anchor to 'center'

    # To configure the dropdown font (for the items inside the menu), modify the 'TMenu' style
    style.configure('Custom.TMenu', font=('Arial', 16))  # Increase font for dropdown items

    # Create the OptionMenu with the custom style
    self.option_menu = ttk.OptionMenu(self, self.selected_option_rep, self.selected_option_rep.get(), *self.options)
    self.option_menu['style'] = 'Custom.TMenubutton'  # Apply the custom style
    self.option_menu.config(width=4)  # Adjust the width of the dropdown
    self.option_menu.place(x=45, y=200)


    # Adjust dropdown items font
    self.option_menu['menu'].config(font=('Arial', 16))

    # Background and label setup
    self.background_color = "#deeaf7"
    background_img = ImageTk.PhotoImage(Image.open('Pictures//empty.JPG'))
    self.label1 = tk.Label(self, text=":מספר חזרות ", image=background_img, compound=tk.CENTER,
                           font=("Ariel", 16, 'bold'), width=110, height=50, bg=self.background_color)
    self.label1.place(x=35, y=140)
    self.label1.image = background_img


def rate_of_exercises(self):

    # List of options for the dropdown
    self.options_rate = ["slow", "moderate", "fast"]

    # Create a StringVar to hold the selected value
    self.selected_option_rate = tk.StringVar()

    # Get the current number of repetitions from Excel
    current_rate = Excel.find_value_by_colName_and_userID(
        "Patients.xlsx", "patients_details", s.chosen_patient_ID, "rate"
    )

    # Set the default selected option based on the value from Excel
    self.selected_option_rate.set(self.options_rate[self.options_rate.index(str(current_rate))])

    # Create a custom style for the OptionMenu
    style = ttk.Style()
    style.theme_use('clam')  # Set theme to 'clam'

    # Configure the appearance of the OptionMenu and the dropdown items
    style.configure('Custom.TMenubutton', font=('Arial', 16, 'bold'), background='lightgray', foreground='black',
                    padding=[10, 10], anchor='center')  # Add padding and set anchor to 'center'

    # To configure the dropdown font (for the items inside the menu), modify the 'TMenu' style
    style.configure('Custom.TMenu', font=('Arial', 14))  # Increase font for dropdown items


    # Create the OptionMenu with the custom style
    self.option_menu_rate = ttk.OptionMenu(self, self.selected_option_rate, self.selected_option_rate.get(), *self.options_rate)
    self.option_menu_rate['style'] = 'Custom.TMenubutton'  # Apply the custom style
    self.option_menu_rate.config(width=8)  # Adjust the width of the dropdown
    self.option_menu_rate.place(x=20, y=340)
    # Adjust dropdown items font
    self.option_menu['menu'].config(font=('Arial', 14))

    # Background and label setup
    self.background_color = "#deeaf7"
    background_img = ImageTk.PhotoImage(Image.open('Pictures//empty.JPG'))


    self.label2 = tk.Label(self, text=":קצב ", image=background_img, compound=tk.CENTER,
                           font=("Ariel", 16, 'bold'), width=50, height=50, bg=self.background_color)
    self.label2.place(x=60, y=280)
    self.label2.image = background_img




class InformationAboutTrainingPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        dates, percent_success, exertion_rate, time_in_training, pdf_paths = self.create_data()
        self.pdf_mapping = pdf_paths

        df = pd.DataFrame({
            'זמן באימון (שניות)': time_in_training,
            'דירוג מאמץ על ידי המטופל': exertion_rate,
            'אחוז הצלחה באימון': percent_success,
            'תאריך ושעה': dates
        })

        # Load the background image
        image = Image.open('Pictures//training_list.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        self.background_label = tk.Label(self, image=self.photo_image)
        self.background_label.pack(fill="both", expand=True)
        name_label(self, 350, 175)

        # Treeview setup
        self.treeview = ttk.Treeview(self, style="Treeview", show="headings")
        self.treeview["columns"] = tuple(df.columns)

        style = ttk.Style(self)
        style.configure("Treeview", font=("Thaoma", 14), rowheight=30)
        style.configure("Treeview.Heading", font=("Thaoma", 14, 'bold'), rowheight=30)


        col_width = 200
        for col in df.columns:
            self.treeview.column(col, anchor="e", width=col_width)
            self.treeview.heading(col, text=col, anchor="e")

        self.treeview.column("דירוג מאמץ על ידי המטופל", anchor="e", width=225)
        self.treeview.column("תאריך ושעה", anchor="e", width=200)

        for index, row in df.iterrows():
            values = tuple(row)
            self.treeview.insert("", index, values=values)

        # Enable row selection
        self.treeview.bind("<Double-1>", self.on_row_double_click)
        treeview_x_place = (1024- 4 * col_width - 25) /2
        self.treeview.config(height=10)
        self.treeview.place(x=treeview_x_place, y=180)

        # Dynamic placement
        tree_width = len(df.columns) * col_width + 25
        tree_height = 310
        screen_width = self.winfo_screenwidth()
        tree_x = (1024 // 2) - (tree_width // 2)
        tree_y = 180

        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.treeview.yview)
        self.treeview.configure(yscrollcommand=scrollbar.set)
        scrollbar.place(x=tree_x + tree_width + 10, y=tree_y, height=tree_height + 20)


        back_button_img = Image.open("Pictures//previous.jpg")
        back_button_photo = ImageTk.PhotoImage(back_button_img)
        back_button = tk.Button(self, image=back_button_photo, command=self.on_click_to_previous,
                                width=back_button_img.width, height=back_button_img.height, bd=0,
                                highlightthickness=0)
        back_button.image = back_button_photo
        back_button.place(x=30, y=30)

    def on_click_to_previous(self):
        s.screen.switch_frame(ChooseTrainingOrExerciseInformation)

    def on_row_double_click(self, event):
        # Get selected item
        selected_item = self.treeview.focus()
        if not selected_item:
            return
        row_index = self.treeview.index(selected_item)

        # Get the associated PDF file
        pdf_path = self.pdf_mapping[row_index]
        if pdf_path and os.path.exists(pdf_path):
            try:
                os.startfile(pdf_path)  # For Windows
            except AttributeError:
                # macOS/Linux
                import subprocess
                subprocess.call(("open", pdf_path))  # macOS
                # For Linux, use:
                # subprocess.call(("xdg-open", pdf_path))
        else:
            print(f"No PDF found for row {row_index} or file does not exist.")

    def create_data(self):
        df = pd.read_excel("Patients.xlsx", sheet_name="patients_history_of_trainings")
        df.iloc[:, 0] = df.iloc[:, 0].astype(str)
        filtered_rows = df[df.iloc[:, 0] == s.chosen_patient_ID]

        row = filtered_rows.iloc[0]
        row_values_without_id = row.iloc[1:]

        dates = []
        percent_success = []
        exertion_rate = []
        time_in_training = []
        pdf_paths = []

        base_dir = os.path.abspath("Patients")  # Get the absolute path of the 'Patients' folder

        for i in range(0, len(row_values_without_id), 4):
            date_value = row_values_without_id.iloc[i]
            if pd.notna(date_value):
                try:
                    # Parse and format the date as dd/mm/yyyy hh:mm:ss
                    formatted_date = datetime.strptime(date_value, "%d-%m-%Y %H-%M-%S").strftime("%d/%m/%Y %H:%M:%S")
                    dates.append(formatted_date)
                except ValueError:
                    # If parsing fails, append the raw value
                    dates.append(date_value)

            if i + 1 < len(row_values_without_id):
                success_value = row_values_without_id.iloc[i + 1]
                if pd.notna(success_value):
                    percent_success.append(int(success_value))

            if i + 2 < len(row_values_without_id):
                exertion_value = row_values_without_id.iloc[i + 2]
                if pd.notna(exertion_value):
                    exertion_rate.append(int(exertion_value))

            if i + 3 < len(row_values_without_id):
                training_time_value = row_values_without_id.iloc[i + 3]
                if pd.notna(training_time_value):
                    time_in_training.append(int(training_time_value))

                # Construct the absolute PDF path
                pdf_path = os.path.join(base_dir, s.chosen_patient_ID, "PDF_to_Therapist_Email", f"{date_value}.pdf")
                pdf_paths.append(pdf_path)

        # Reverse all lists to show the last training first
        dates.reverse()
        percent_success.reverse()
        exertion_rate.reverse()
        time_in_training.reverse()
        pdf_paths.reverse()

        return dates, percent_success, exertion_rate, time_in_training, pdf_paths

class ExplanationPage(tk.Frame):
    def __init__(self, master, exercise, **kwargs):
        tk.Frame.__init__(self, master, **kwargs)

        self.exercise= exercise
        image = Image.open('Pictures//background.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()

        self.label = tk.Label(self)
        self.label.place(x=300, y=85)
        video_file = f'Videos//{exercise}.mp4'
        video_path = os.path.join(os.getcwd(), video_file)
        self.cap = cv2.VideoCapture(video_path)

        skip_explanation_button_img = Image.open("Pictures//skip_explanation_button.jpg")
        skip_explanation_button_photo = ImageTk.PhotoImage(skip_explanation_button_img)
        skip_explanation_button = tk.Button(self, image=skip_explanation_button_photo, command=lambda: self.on_click_skip_explanation(),
                                width=skip_explanation_button_img.width, height=skip_explanation_button_img.height, bd=0,
                                highlightthickness=0)
        skip_explanation_button.image = skip_explanation_button_photo
        skip_explanation_button.place(x=30, y=30)

        # skip_explanation_button_img = Image.open("Pictures//skip_explanation_button.jpg")
        # skip_explanation_button_photo = ImageTk.PhotoImage(skip_explanation_button_img)
        # skip_explanation_button = tk.Button(self, image=skip_explanation_button_photo,
        #                                     command=lambda: self.on_click_skip_exercise(),
        #                                     width=skip_explanation_button_img.width,
        #                                     height=skip_explanation_button_img.height, bd=0,
        #                                     highlightthickness=0)
        # skip_explanation_button.image = skip_explanation_button_photo
        # skip_explanation_button.place(x=30, y=500)

        if not (self.cap.isOpened()):
            print("Error opening video streams or files")

        else:
            # Play videos
            play_video_explanation(self.cap, self.label, video_path)
            if exercise =="notool_right_bend_left_up_from_side" or exercise == "notool_left_bend_right_up_from_side":
                say("notool_arm_bend_arm_up_from_side", True)
                x = get_wav_duration("notool_arm_bend_arm_up_from_side")

            else:
                say(exercise, True) #True so the system will know that it is an explanation
                x = get_wav_duration(exercise)

            self.after(int(x*1000+2000), lambda: self.end_of_explanation())

    def end_of_explanation(self):
        if not s.explanation_over:
            s.explanation_over = True
        s.repeat_explanation = False
        s.suggest_repeat_explanation = False

    def on_click_skip_explanation(self):
        s.explanation_over = True
        s.repeat_explanation = False
        s.suggest_repeat_explanation = False


    def on_click_skip_exercise(self):
        s.skip = True
        s.exercises_skipped[self.exercise] = {"0"}


class Number_of_good_repetitions_page(tk.Frame):
    def __init__(self, master, **kwargs):
        tk.Frame.__init__(self, master, **kwargs)

        # Create a canvas for layering background and text with no border or highlight
        self.canvas = tk.Canvas(self, width=1024, height=576, bd=0, highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)

        # Load and display the background image
        image = Image.open("Pictures//fireworks.jpg")
        self.photo_image = ImageTk.PhotoImage(image)
        self.canvas.create_image(0, 0, image=self.photo_image, anchor="nw")

        if s.patient_repetitions_counting_in_exercise == 10:
            alignment1 = "center"
        else:
            alignment1= "e"

        if s.rep == 10:
            alignment3= "center"
        else:
            alignment3 = "w"

        # Overlay text directly on the canvas instead of using Labels
        self.canvas.create_text(960, 420, text= f'{s.patient_repetitions_counting_in_exercise}',
                                font=("Arial", 65, "bold"), fill="white", anchor=alignment1)
        self.canvas.create_text(520, 420, text= "חזרות מוצלחות מתוך",
                                font=("Arial", 65, "bold"), fill="white", anchor="center")
        self.canvas.create_text(70, 420, text= f'{s.rep}',
                                font=("Arial", 65, "bold"), fill="white", anchor=alignment3)

        exercises_left = len(s.ex_in_training) - s.num_exercises_started
        if s.stop_requested:
            self.canvas.create_text(520, 510, text=f' מספר התרגילים שלא ביצעת באימון: {exercises_left}',
                                font=("Arial", 40, "bold"), fill="white", anchor="center")

        else:
            self.canvas.create_text(520, 510, text=f' נותרו עוד {exercises_left} תרגילים לסיום האימון ',
                                font=("Arial", 40, "bold"), fill="white", anchor="center")



class ChooseTrainingOrExerciseInformation(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)

        image = Image.open('Pictures//background.jpg')
        self.photo_image = ImageTk.PhotoImage(image) #self. - for keeping the photo in memory so it will be shown
        tk.Label(self, image = self.photo_image).pack()
        name_label(self, 350, 175)

        to_previous_button_img = Image.open("Pictures//previous.jpg")
        to_previous_button_photo = ImageTk.PhotoImage(to_previous_button_img)
        to_previous_button = tk.Button(self, image=to_previous_button_photo, command=lambda: self.to_previous_button_click(),
                                   width=to_previous_button_img.width, height=to_previous_button_img.height, bd=0,
                                   highlightthickness=0)  # Set border width to 0 to remove button border
        to_previous_button.image = to_previous_button_photo  # Store reference to image to prevent garbage collection
        to_previous_button.place(x=30, y=30)

        # Load images for buttons
        training_image = Image.open("Pictures//information_training_button.jpg")
        training_photo = ImageTk.PhotoImage(training_image)
        exercise_image = Image.open("Pictures//to_information_exercises_button.jpg")
        exercise_photo = ImageTk.PhotoImage(exercise_image)

        # Store references to prevent garbage collection
        self.training_photo = training_photo
        self.patient_photo = exercise_photo

        # Create buttons with images
        enter_trainings = tk.Button(self, image=training_photo,
                                              command=self.on_click_training_chosen,
                                              width=training_image.width, height=training_image.height,
                                              bg='#50a6ad', bd=0, highlightthickness=0)
        enter_trainings.place(x=160, y=130)

        enter_exercises = tk.Button(self, image=exercise_photo,
                                            command=self.on_click_exercise_chosen,
                                            width=exercise_image.width, height=exercise_image.height,
                                            bg='#50a6ad', bd=0, highlightthickness=0)
        enter_exercises.place(x=540, y=130)

    def to_previous_button_click(self):
        s.screen.switch_frame(PatientDisplaying)

    def on_click_training_chosen(self):
        s.screen.switch_frame(InformationAboutTrainingPage)

    def on_click_exercise_chosen(self):
        s.screen.switch_frame(ChooseBallExercisesPage)



################################################# Categories Screens ##############################################
class StartOfTraining(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        s.play_song = True
        image = Image.open('Pictures//hello.jpg')
        self.photo_image = ImageTk.PhotoImage(image) #self. - for keeping the photo in memory so it will be shown
        tk.Label(self, image = self.photo_image).pack()
        s.asked_for_measurement = True
        say("welcome")
        self.after(int(get_wav_duration("welcome")*1000 +1000),lambda: s.screen.switch_frame(StartExplanationPage))



class CalibrationScreen(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master, width=1024, height=576)  # Set frame size

        # Make sure the frame does not shrink or expand
        self.pack_propagate(False)

        # Create a label for the camera feed
        self.camera_label = tk.Label(self, width=1024, height=576)  # Set label size
        self.camera_label.pack()

        # Load all overlay images at once (must be transparent PNGs)
        self.overlay_images = {
            3: Image.open(os.path.join(os.getcwd(), "Pictures", "3.png")).convert("RGBA"),
            2: Image.open(os.path.join(os.getcwd(), "Pictures", "2.png")).convert("RGBA"),
            1: Image.open(os.path.join(os.getcwd(), "Pictures", "1.png")).convert("RGBA")
        }

        # Load "TRY AGAIN" image
        self.try_again_image = Image.open(os.path.join(os.getcwd(), "Pictures", "try_again.png")).convert("RGBA")
        self.lets_go_image = Image.open(os.path.join(os.getcwd(), "Pictures", "lets_go.png")).convert("RGBA")


        say("start_calibration")
        self.after(10, self.update_camera)

        self.count = 0
        self.target_size = (1024, 576)  # Target size for the camera feed

        self.load_gif("giphy.gif")  # Or use the full path if needed

    def load_gif(self, gif_path, size=(200, 200)):
        self.gif_frames = []
        gif = Image.open(gif_path)
        for frame in ImageSequence.Iterator(gif):
            frame = frame.convert("RGBA").resize(size, Image.LANCZOS)
            self.gif_frames.append(frame)
        self.current_gif_frame = 0

    def update_camera(self):
        """Continuously updates the camera feed inside Tkinter with countdown and pauses."""
        current_time = time.time()

        # Only increment count if not in pause period
        if not hasattr(self, 'pause_until'):
            self.pause_until = time.time()  # Initialize once

        if current_time >= (self.pause_until + get_wav_duration("start_calibration")):
            self.count += 1

        frame = s.zed_camera.get_latest_frame()

        if frame is not None:
            img = Image.fromarray(frame).resize(self.target_size, Image.LANCZOS)
            overlay_image = None

            if 20 <= self.count < 35:
                if self.count == 20:
                    say("bip_sound")
                overlay_image = self.overlay_images[3]
            elif 35 <= self.count < 50:
                if self.count == 35:
                    say("bip_sound")
                overlay_image = self.overlay_images[2]
            elif 50 <= self.count <= 64:
                if self.count == 50:
                    say("bip_sound")
                overlay_image = self.overlay_images[1]

            if self.count == 64:
                say("end_counting_sound")
                s.screen_finished_counting = True

            if overlay_image and self.count < 64:
                x_offset = (1024 - overlay_image.width) // 2
                y_offset = (576 - overlay_image.height) // 2
                img.paste(overlay_image, (x_offset, y_offset), overlay_image)

            # Show GIF animation on top when count > 80
            if 65 < self.count < 100 and hasattr(self, "gif_frames") and self.gif_frames:
                gif_frame = self.gif_frames[self.current_gif_frame]
                self.current_gif_frame = (self.current_gif_frame + 1) % len(self.gif_frames)

                # Calculate vertical center
                y_offset = (img.height - gif_frame.height) -20

                # Left side (10 pixels from edge)
                left_x = 10
                img.paste(gif_frame, (left_x, 20), gif_frame)

                # Right side (10 pixels from right edge)
                right_x = img.width - gif_frame.width - 10
                img.paste(gif_frame, (right_x, y_offset), gif_frame)
            imgtk = ImageTk.PhotoImage(image=img)
            self.camera_label.imgtk = imgtk
            self.camera_label.config(image=imgtk)

        if s.asked_for_measurement and not s.finished_calibration:
            if self.count < 110:
                self.after(10, self.update_camera)
            else:
                self.check_if_distances_None()

    def check_if_distances_None(self):
        """Check if measurements are None, show 'TRY AGAIN' for 2 sec, then restart."""
        if s.len_left_arm is None or s.len_right_arm is None or s.dist_between_wrists is None or s.dist_between_shoulders is None or s.len_left_upper_arm is None or s.len_right_upper_arm is None or s.shoulder_problem_calibration:

            print("Some distances are None. Displaying 'TRY AGAIN'.")
            s.screen_finished_counting = False

            # Get latest camera frame
            frame = s.zed_camera.get_latest_frame()
            if frame is not None:
                img = Image.fromarray(frame).resize(self.target_size, Image.LANCZOS)

                # Center "TRY AGAIN" image
                x_offset = (1024 - self.try_again_image.width) // 2
                y_offset = (576 - self.try_again_image.height) // 2
                img.paste(self.try_again_image, (x_offset, y_offset), self.try_again_image)

                # Convert final image to Tkinter format
                imgtk = ImageTk.PhotoImage(image=img)
                self.camera_label.imgtk = imgtk
                self.camera_label.config(image=imgtk)

            if s.len_left_arm is None or s.len_right_arm is None or s.dist_between_wrists is None or s.dist_between_shoulders is None or s.len_left_upper_arm is None or s.len_right_upper_arm is None:
                str_to_say = "didnt_recognize_calibration"

            elif s.shoulder_problem_calibration:
                str_to_say ="shoulders_not_good_calibration"

            s.shoulder_problem_calibration = False

            say(str_to_say)
            self.after(int(get_wav_duration(str_to_say) * 1000 + 1000), self.restart_countdown)

        else:
            if s.try_again_calibration:
                say("end_not_first_time_calibration")

            else:
                say("end_calibration")

            s.try_again_calibration = False
            s.finished_calibration = True
            # Get latest camera frame
            frame = s.zed_camera.get_latest_frame()
            if frame is not None:
                img = Image.fromarray(frame).resize(self.target_size, Image.LANCZOS)

                # Center "TRY AGAIN" image
                x_offset = (1024 - self.lets_go_image.width) // 2
                y_offset = (576 - self.lets_go_image.height) // 2
                img.paste(self.lets_go_image, (x_offset, y_offset), self.lets_go_image)

                # Convert final image to Tkinter format
                imgtk = ImageTk.PhotoImage(image=img)
                self.camera_label.imgtk = imgtk
                self.camera_label.config(image=imgtk)

            s.asked_for_measurement = False


    def restart_countdown(self):
        """Restart the countdown process from 3 after displaying 'TRY AGAIN'."""
        print("🔄 Restarting countdown...")
        s.screen_finished_counting = False
        self.count = 0 # Reset count
        self.update_camera()  # Restart the countdown


class StartExplanationPage(tk.Frame):
    def __init__(self, master, **kwargs):
        tk.Frame.__init__(self, master, **kwargs)
        image = Image.open('Pictures//background.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()

        s.explanation_over = False
        self.label = tk.Label(self)
        self.label.place(x=200, y=85)
        video_file = f'Videos//start_explanation.mp4'
        video_path = os.path.join(os.getcwd(), video_file)
        self.cap = cv2.VideoCapture(video_path)

        skip_explanation_button_img = Image.open("Pictures//skip_explanation_button.jpg")
        skip_explanation_button_photo = ImageTk.PhotoImage(skip_explanation_button_img)
        skip_explanation_button = tk.Button(self, image=skip_explanation_button_photo, command=lambda: self.on_click_skip(),
                                width=skip_explanation_button_img.width, height=skip_explanation_button_img.height, bd=0,
                                highlightthickness=0)
        skip_explanation_button.image = skip_explanation_button_photo
        skip_explanation_button.place(x=30, y=30)


        if not (self.cap.isOpened()):
            print("Error opening video streams or files")

        else:
            # Play videos
            play_video_explanation(self.cap, self.label, video_path, 0.8, 0.8, 1.2)
            say("start_explanation", True)  # True so the system will know that it is an explanation
            x = get_wav_duration("start_explanation")
            self.after(int(x*1000+1000), lambda: self.end_of_explanation())

    def end_of_explanation(self):
        if not s.explanation_over:
            s.explanation_over = True
            self.after(500, lambda: s.screen.switch_frame(CalibrationScreen))

    def on_click_skip(self):
        s.explanation_over = True
        self.after(500, lambda: s.screen.switch_frame(CalibrationScreen))


def wait_until_waving():
    s.req_exercise="hello_waving"

class Ball(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//ball.jpg')
        self.photo_image = ImageTk.PhotoImage(image) #self. - for keeping the photo in memory so it will be shown
        tk.Label(self, image = self.photo_image).pack()
        say("ball")
        self.after(6000,lambda: wait_until_waving())


class Stick(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//stick.jpg')
        self.photo_image = ImageTk.PhotoImage(image) #self. - for keeping the photo in memory so it will be shown
        tk.Label(self, image = self.photo_image).pack()
        say("stick")
        self.after(6000,lambda: wait_until_waving())


class Rubber_Band(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//rubberband.jpg')
        self.photo_image = ImageTk.PhotoImage(image) #self. - for keeping the photo in memory so it will be shown
        tk.Label(self, image = self.photo_image).pack()
        say("band")
        self.after(6000,lambda: wait_until_waving())

class Weights(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//weights.jpg')
        self.photo_image = ImageTk.PhotoImage(image) #self. - for keeping the photo in memory so it will be shown
        tk.Label(self, image = self.photo_image).pack()
        say("weights")
        self.after(6000,lambda: wait_until_waving())

class NoTool(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//notool.jpg')
        self.photo_image = ImageTk.PhotoImage(image) #self. - for keeping the photo in memory so it will be shown
        tk.Label(self, image = self.photo_image).pack()
        say("notool")
        self.after(6000,lambda: wait_until_waving())


######################################### Counting Pages #############################################################
class OnePage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//1.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()


class TwoPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//2.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()


class ThreePage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//3.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()


class FourPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//4.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()


class FivePage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//5.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()


class SixPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//6.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()


class SevenPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//7.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()


class EightPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//8.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()


class NinePage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//9.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()


class TenPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//10.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()


######################################################## Goodbbye Page #################################################
class GoodbyePage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//goodbye.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()
        say('goodbye')


class ClappingPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        random_number_1 = random.randint(1, 5)
        image = Image.open(f'Pictures//clap_{random_number_1}.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        tk.Label(self, image=self.photo_image).pack()
        random_number_2 = random.randint(1, 8)
        say(f'clap {random_number_2}')
        self.after(int(get_wav_duration(f'clap {random_number_2}')*1000+500), lambda: s.screen.switch_frame(GoodbyePage))




######################################################## Effort scale Page #################################################
class EffortScale(tk.Frame):
    def __init__(self, master, **kwargs):
        s.play_song = False
        tk.Frame.__init__(self, master, **kwargs)
        image = Image.open('Pictures//scale.jpg')
        self.photo_image = ImageTk.PhotoImage(image)
        self.background_label = tk.Label(self, image=self.photo_image)
        self.background_label.pack()
        say('please_rate_effort', False, True)



        ## buttons to press for scale
        image0 = Image.open('Pictures//scale_0.jpg')
        resized_image0 = image0.resize((350, 40), Image.LANCZOS)
        self.photo_image0 = ImageTk.PhotoImage(resized_image0)
        button0= tk.Button(self,image=self.photo_image0, command=self.on_click_0)
        button0.place(height=40, width=350, x=180, y=100)

        image1 = Image.open('Pictures//scale_1.jpg')
        resized_image1 = image1.resize((350, 40), Image.LANCZOS)
        self.photo_image1 = ImageTk.PhotoImage(resized_image1)
        button1= tk.Button(self,image=self.photo_image1, command=self.on_click_1)
        button1.place(height=40, width=350, x=180, y=140)

        image2 = Image.open('Pictures//scale_2.jpg')
        resized_image2= image2.resize((350, 40), Image.LANCZOS)
        self.photo_image2 = ImageTk.PhotoImage(resized_image2)
        button2 = tk.Button(self, image=self.photo_image2, command=self.on_click_2)
        button2.place(height=40, width=350, x=180, y=180)

        image3 = Image.open('Pictures//scale_3.jpg')
        resized_image3 = image3.resize((350, 40), Image.LANCZOS)
        self.photo_image3 = ImageTk.PhotoImage(resized_image3)
        button3 = tk.Button(self, image=self.photo_image3, command=self.on_click_3)
        button3.place(height=40, width=350, x=180, y=220)

        image4 = Image.open('Pictures//scale_4.jpg')
        resized_image4 = image4.resize((350, 40), Image.LANCZOS)
        self.photo_image4 = ImageTk.PhotoImage(resized_image4)
        button4 = tk.Button(self, image=self.photo_image4, command=self.on_click_4)
        button4.place(height=40, width=350, x=180, y=260)

        image5 = Image.open('Pictures//scale_5.jpg')
        resized_image5 = image5.resize((350, 40), Image.LANCZOS)
        self.photo_image5 = ImageTk.PhotoImage(resized_image5)
        button5 = tk.Button(self, image=self.photo_image5, command=self.on_click_5)
        button5.place(height=40, width=350, x=180, y=300)

        image6 = Image.open('Pictures//scale_6.jpg')
        resized_image6 = image6.resize((350, 40), Image.LANCZOS)
        self.photo_image6 = ImageTk.PhotoImage(resized_image6)
        button6 = tk.Button(self, image=self.photo_image6, command=self.on_click_6)
        button6.place(height=40, width=350, x=180, y=340)

        image7 = Image.open('Pictures//scale_7.jpg')
        resized_image7 = image7.resize((350, 40), Image.LANCZOS)
        self.photo_image7 = ImageTk.PhotoImage(resized_image7)
        button7 = tk.Button(self, image=self.photo_image7, command=self.on_click_7)
        button7.place(height=40, width=350, x=180, y=380)

        image8 = Image.open('Pictures//scale_8.jpg')
        resized_image8 = image8.resize((350, 40), Image.LANCZOS)
        self.photo_image8 = ImageTk.PhotoImage(resized_image8)
        button8 = tk.Button(self, image=self.photo_image8, command=self.on_click_8)
        button8.place(height=40, width=350, x=180, y=420)

        image9 = Image.open('Pictures//scale_9.jpg')
        resized_image9 = image9.resize((350, 40), Image.LANCZOS)
        self.photo_image9 = ImageTk.PhotoImage(resized_image9)
        button9 = tk.Button(self, image=self.photo_image9, command=self.on_click_9)
        button9.place(height=40, width=350, x=180, y=460)

        image10 = Image.open('Pictures//scale_10.jpg')
        resized_image10 = image10.resize((350, 40), Image.LANCZOS)
        self.photo_image10 = ImageTk.PhotoImage(resized_image10)
        button10 = tk.Button(self, image=self.photo_image10, command=self.on_click_10)
        button10.place(height=40, width=350, x=180, y=500)


    def on_click_0(self):
        s.effort= 0
        s.finished_effort=True

    def on_click_1(self):
        s.effort = 1
        s.finished_effort = True

    def on_click_2(self):
        s.effort = 2
        s.finished_effort = True

    def on_click_3(self):
        s.effort = 3
        s.finished_effort = True

    def on_click_4(self):
        s.effort = 4
        s.finished_effort = True

    def on_click_5(self):
        s.effort = 5
        s.finished_effort = True

    def on_click_6(self):
        s.effort = 6
        s.finished_effort = True

    def on_click_7(self):
        s.effort = 7
        s.finished_effort = True

    def on_click_8(self):
        s.effort = 8
        s.finished_effort = True

    def on_click_9(self):
        s.effort = 9
        s.finished_effort = True
    def on_click_10(self):
        s.effort = 10
        s.finished_effort = True


class Repeat_training_or_not(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)

        image = Image.open('Pictures//background.jpg')
        self.photo_image = ImageTk.PhotoImage(image) #self. - for keeping the photo in memory so it will be shown
        tk.Label(self, image = self.photo_image).pack()




        # Load images for buttons
        finish_training_image = Image.open("Pictures//finish_training_button.jpg")
        finish_training_photo = ImageTk.PhotoImage(finish_training_image)
        repeat_training_image = Image.open("Pictures//repeat_training_button.jpg")
        repeat_training_photo = ImageTk.PhotoImage(repeat_training_image)

        # Store references to prevent garbage collection
        self.finish_training_photo = finish_training_photo
        self.repeat_training_photo = repeat_training_photo

        # Create buttons with images
        finish_training_button = tk.Button(self, image=finish_training_photo,
                                              command=self.on_click_not_repeat,
                                              width=finish_training_image.width, height=finish_training_image.height,
                                              bg='#50a6ad', bd=0, highlightthickness=0)
        finish_training_button.place(x=160, y=130)

        repeat_training_button = tk.Button(self, image=repeat_training_photo,
                                            command=self.on_click_repeat,
                                            width=repeat_training_image.width, height=repeat_training_image.height,
                                            bg='#50a6ad', bd=0, highlightthickness=0)
        repeat_training_button.place(x=540, y=130)


    def on_click_repeat(self):
        s.another_training_requested= True
        s.choose_continue_or_not= True

        s.screen.switch_frame(Not_first_round_entrance_page)

    def on_click_not_repeat(self):
        s.choose_continue_or_not= True
        s.screen.switch_frame(ClappingPage)


class Not_first_round_entrance_page(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        image = Image.open('Pictures//not_first_round_entrance.jpg')
        self.photo_image = ImageTk.PhotoImage(image) #self. - for keeping the photo in memory so it will be shown
        tk.Label(self, image = self.photo_image).pack()



class FullScreenApp(object):
    def __init__(self, master, **kwargs):
        self.master=master
        pad=3
        self._geom='200x200+0+0'
        master.geometry("{0}x{1}+0+0".format(
            master.winfo_screenwidth()-pad, master.winfo_screenheight()-pad))
        master.bind('<Escape>',self.toggle_geom)

    def toggle_geom(self,event):
        geom=self.master.winfo_geometry()
        print(geom,self._geom)
        self.master.geometry(self._geom)
        self._geom=geom


if __name__ == "__main__":
    # s.audio_path = 'audio files/Hebrew/Male/'
    s.screen = Screen()
    # s.finished_effort= False
    # s.ex_in_training=["bend_elbows_ball", "arms_up_and_down_stick"]
    # s.list_effort_each_exercise= {}
    s.chosen_patient_ID='3333'
    s.ball_exercises_number = 4
    s.band_exercises_number = 5
    s.stick_exercises_number = 5
    s.weights_exercises_number = 2
    s.no_tool_exercises_number = 6
    #s.screen.switch_frame(ExplanationPage, exercise="bend_elbows_ball")
    s.gymmy_done=False
    s.camera_done= False
    s.finish_program = False
    s.last_entry_angles = None

    s.rep=10
    s.volume = 0.3
    s.audio_path = 'audio files/Hebrew/Male/'
    s.finish_workout = False
    pygame.mixer.init()
    s.stop_song = False
    s.play_song = False
    s.reached_max_limit = False
    s.latest_keypoints = []
    s.additional_audio_playing =False
    # Start continuous audio in a separate thread
    s.continuous_audio = ContinuousAudio()
    s.continuous_audio.start()
    s.patient_repetitions_counting_in_exercise = 10
    # Initialize settings variables
    s.stop = False
    s.finish_workout = False
    s.all_rules_ok = False
    # Start the ZED camera thread
    s.zed_camera = PyZedWrapper()
    s.zed_camera.start()
    # Initialize Tkinter-based GUI system
    # s.screen.switch_frame(ExercisePageNew, exercise_type="shoulders_distance_y", real_distance=140)
    s.screen.switch_frame(ChooseTrainingOrExerciseInformation)

    # When needed, switch to the ExercisePage with the camera
    # s.screen.switch_frame(ExercisePage, camera_thread=s.camera)

    app = FullScreenApp(s.screen)
    s.screen.mainloop()