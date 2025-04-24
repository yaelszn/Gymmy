import pandas as pd
import xlsxwriter
from datetime import datetime

import ScreenNew
from Joint_zed import Joint
import openpyxl
import Settings as s
from openpyxl import load_workbook
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import os
import subprocess
import platform
import re
from datetime import timedelta



def create_and_open_folder(folder_path):
    try:
        # Create the folder if it doesn't exist
        os.makedirs(folder_path, exist_ok=True)
        print(f"Directory created or already exists: {folder_path}")

        # Check if the folder exists after creation
        if not os.path.exists(folder_path):
            raise FileNotFoundError(f"Directory still not found: {folder_path}")

    except Exception as e:
        print(f"An error occurred: {e}")




#creats a new workbook to each training
def create_workbook_for_training():
    datetime_string = datetime.now().strftime("%d-%m-%Y %H-%M-%S")
    workbook_name = f"Patients/{s.chosen_patient_ID}/{datetime_string}.xlsx"
    s.training_workbook_path = workbook_name
    s.training_workbook = openpyxl.Workbook()  # Do not pass the filename here
    s.training_workbook.save(s.training_workbook_path)  # Save the workbook after creating it




#returns a specific value by ID and name of the column
def find_value_by_colName_and_userID(workbook_path, worksheet_name, ID, target_col):
    try:
        # Load the workbook
        workbook = load_workbook(workbook_path)

        # Access the worksheet
        worksheet = workbook[worksheet_name]
        target_row=None

        # Search for the target value in the first column
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1), start=2):
            for cell in row:
                if str(cell.value) == ID:
                    target_row= row_number  # Return row number and value from specified column
                    break
            if target_row is not None:
                break

        for col in worksheet.iter_cols():
            # Check if the first cell in the column matches the specified column name
            if col[0].value == target_col:
                # Return the entire column
                return col[target_row-1].value

    except Exception as e:
        print(f"An error occurred: {e}")
        return None


#returns the value of success in a specific training and specific exercise
def get_success_number(file_path, exercise):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        # Check if the worksheet exists
        if "success_worksheet" not in workbook.sheetnames:
            print(f"Worksheet success not found in the workbook.")
            return None

        # Select the worksheet
        worksheet = workbook["success_worksheet"]

        # Iterate through the rows in the first column and search for the value
        for row in worksheet.iter_rows(values_only=True):
            if row[0] == exercise:
                # Return the value from the second column
                return row[1]

        # If the value is not found, return None
        return None

    except FileNotFoundError:
        print(f"File success not found.")
        return None

    except Exception as e:
        print(f"An error occurred: {e}")
        return None


# #returns the value of effort rate in a specific training and specific exercise
# def get_effort_number(file_path, exercise):
#     try:
#         # Load the workbook
#         workbook = openpyxl.load_workbook(file_path)
#
#         # Check if the worksheet exists
#         if "success_and_effort" not in workbook.sheetnames:
#             print(f"Worksheet success not found in the workbook.")
#             return None
#
#         # Select the worksheet
#         worksheet = workbook["success_and_effort"]
#
#         return worksheet[1,2]
#
#
#     except FileNotFoundError:
#         print(f"File success not found.")
#         return None
#

def wf_joints(ex_name, list_joints):
    # Check if "Sheet" exists and delete it
    if "Sheet" in s.training_workbook.sheetnames:
        sheet_to_delete = s.training_workbook["Sheet"]
        s.training_workbook.remove(sheet_to_delete)

    # Create a new sheet with the provided name
    worksheet1 = s.training_workbook.create_sheet(ex_name[:31])
    col = 0

    for l in range(0, len(list_joints)):
        worksheet1.cell(row=1, column=col + 1, value=col + 1)

        row = 2
        for j in list_joints[l]:
            if isinstance(j, Joint):  # Check if j is a Joint object
                j_ar = j.joint_to_array()
                for i in range(len(j_ar)):
                    worksheet1.cell(row=row, column=col + 1, value=str(j_ar[i]))
                    row += 1
            else:
                # Handle other types appropriately, e.g., just write the value to the worksheet
                worksheet1.cell(row=row, column=col + 1, value=str(j))
                row += 1

        col += 1

    # Save the workbook
    success_worksheet()
    s.training_workbook.save(s.training_workbook_path)
    create_graphs_and_tables(ex_name, list_joints)



def create_graphs_and_tables(exercise, list_joints):
    try:
        if get_number_of_angles_in_exercise(exercise) == 1:
            one_angle_graph_and_table(exercise, list_joints)
        if get_number_of_angles_in_exercise(exercise) == 2:
            two_angles_graph_and_table(exercise, list_joints)
        if get_number_of_angles_in_exercise(exercise) == 3:
            three_angles_graph_and_table(exercise, list_joints)



    except (pd.errors.ParserError, FileNotFoundError):
        # Handle the case where the sheet is not found
        pass  # Continue to the next iteration
    except ValueError as ve:
        # Handle other specific errors
        pass  # Continue to the next iteration

def get_name_by_exercise(exercise_value):
    data = pd.read_excel("exercises_table.xlsx")
    result = data.loc[data['exercise'] == exercise_value, 'name']
    # Return the exact value, including formatting (e.g., RTL or LTR)
    return result.iloc[0] if not result.empty else None


def get_number_of_angles_in_exercise(exercise_value):
    data = pd.read_excel("exercises_table.xlsx")
    result = data.loc[data['exercise'] == exercise_value, 'number of angles']
    return result.iloc[0] if not result.empty else None


def get_equipment(exercise_value):
    data = pd.read_excel("exercises_table.xlsx")
    result = data.loc[data['exercise'] == exercise_value, 'equipment']
    return result.iloc[0] if not result.empty else None


def get_repetitions_per_count(exercise_value):
    data = pd.read_excel("exercises_table.xlsx")
    result = data.loc[data['exercise'] == exercise_value, 'counts after one or two repetitions']
    return result.iloc[0] if not result.empty else None


def get_files_names_by_start_word(word):
    """
    Retrieves all file names in the directory that start with 'dont_recognize'.
    Returns:
    - List[str]: A list of matching file names without extensions.
    """
    matching_file_names = []

    if not os.path.exists(s.audio_path):
        print(f"Directory does not exist: {s.audio_path}")
        return matching_file_names

    for file_name in os.listdir(s.audio_path):
        if not file_name.startswith(word):
            continue
        name_without_extension, _ = os.path.splitext(file_name)
        matching_file_names.append(name_without_extension)

    return matching_file_names

def one_angle_graph_and_table(exercise_name, list_joints):
    if (list_joints!=[]):
        last_two_values = [entry[-2:] for entry in list_joints] #extract from each record the last 2 values (the angles)
        right_angles = [sublist[0] for sublist in last_two_values] #the right angle from each record
        left_angles = [sublist[1] for sublist in last_two_values] #the left angle from each record


        #extract the joints names and create graphs names
        first_values= list_joints[0]
        first_6_values= first_values[:6]
        joints_names = [str(sample).split()[0] for sample in first_6_values]
        first_graph_name= joints_names[0]+", "+joints_names[1]+", "+joints_names[2]+" 1"
        second_graph_name= joints_names[3]+", "+joints_names[4]+", "+joints_names[5]+" 2"

        #create a list of x values
        length= len(list_joints)
        measurement_num = list(range(1, length + 1))

        #create a data dic for graph
        data = {
        first_graph_name: {'x': measurement_num, 'y': right_angles},
        second_graph_name: {'x': measurement_num, 'y': left_angles}}

        create_and_save_graph(data, exercise_name)
        create_and_save_table_with_calculations(data, exercise_name)


def two_angles_graph_and_table(exercise_name, list_joints):
    if (list_joints!=[]):
        last_four_values = [entry[-4:] for entry in list_joints]  # extract from each record the last 4 values (the angles)
        right_angles = [sublist[0] for sublist in last_four_values]  # the right angle from each record
        left_angles = [sublist[1] for sublist in last_four_values]  # the left angle from each record
        right_angles2 = [sublist[2] for sublist in last_four_values]  # the second right angle from each record
        left_angles2 = [sublist[3] for sublist in last_four_values]  # the second left angle from each record

        # extract the joints names and create graphs names
        first_values = list_joints[0]
        first_12_values = first_values[:12]
        joints_names = [str(sample).split()[0] for sample in first_12_values]
        first_graph_name = joints_names[0] + ", " + joints_names[1] + ", " + joints_names[2]+" 1"
        second_graph_name = joints_names[3] + ", " + joints_names[4] + ", " + joints_names[5]+" 2"
        third_graph_name = joints_names[6] + ", " + joints_names[7] + ", " + joints_names[8]+" 3"
        fourth_graph_name = joints_names[9] + ", " + joints_names[10] + ", " + joints_names[11]+" 4"

        # create a list of x values
        length = len(list_joints)
        measurement_num = list(range(1, length + 1))

        # create a data dic for graph
        data = {
            first_graph_name: {'x': measurement_num, 'y': right_angles},
            second_graph_name: {'x': measurement_num, 'y': left_angles},
            third_graph_name: {'x': measurement_num, 'y': right_angles2},
            fourth_graph_name: {'x': measurement_num, 'y': left_angles2}
        }

        create_and_save_graph(data, exercise_name)
        create_and_save_table_with_calculations(data, exercise_name)


def three_angles_graph_and_table(exercise_name, list_joints):
    if (list_joints!=[]):
        last_four_values = [entry[-6:] for entry in list_joints]  # extract from each record the last 6 values (the angles)
        right_angles = [sublist[0] for sublist in last_four_values]  # the right angle from each record
        left_angles = [sublist[1] for sublist in last_four_values]  # the left angle from each record
        right_angles2 = [sublist[2] for sublist in last_four_values]  # the second right angle from each record
        left_angles2 = [sublist[3] for sublist in last_four_values]  # the second left angle from each record
        right_angles3 = [sublist[4] for sublist in last_four_values]  # the third right angle from each record
        left_angles3 = [sublist[5] for sublist in last_four_values]  # the third left angle from each record

        # extract the joints names and create graphs names
        first_values = list_joints[0]
        first_18_values = first_values[:18]
        joints_names = [str(sample).split()[0] for sample in first_18_values]
        first_graph_name = joints_names[0] + ", " + joints_names[1] + ", " + joints_names[2]+" 1"
        second_graph_name = joints_names[3] + ", " + joints_names[4] + ", " + joints_names[5]+" 2"
        third_graph_name = joints_names[6] + ", " + joints_names[7] + ", " + joints_names[8]+" 3"
        fourth_graph_name = joints_names[9] + ", " + joints_names[10] + ", " + joints_names[11]+" 4"
        fifth_graph_name = joints_names[12] + ", " + joints_names[13] + ", " + joints_names[14]+" 5"
        sixth_graph_name = joints_names[15] + ", " + joints_names[16] + ", " + joints_names[17]+" 6"

        # create a list of x values
        length = len(list_joints)
        measurement_num = list(range(1, length + 1))

        # create a data dic for graph
        data = {
            first_graph_name: {'x': measurement_num, 'y': right_angles},
            second_graph_name: {'x': measurement_num, 'y': left_angles},
            third_graph_name: {'x': measurement_num, 'y': right_angles2},
            fourth_graph_name: {'x': measurement_num, 'y': left_angles2},
            fifth_graph_name: {'x': measurement_num, 'y': right_angles3},
            sixth_graph_name: {'x': measurement_num, 'y': left_angles3}
        }

        create_and_save_graph(data, exercise_name)
        create_and_save_table_with_calculations(data, exercise_name)


# def two_joints_distance_graphs_and_table(exercise_name, list_joints):
#     if (list_joints!=[]):
#         last_two_values = [entry[-2:] for entry in list_joints] #extract from each record the last 2 values (the angles)
#         right_distance = [sublist[0] for sublist in last_two_values] #the right angle from each record
#         left_distance = [sublist[1] for sublist in last_two_values] #the left angle from each record
#
#
#         #extract the joints names and create graphs names
#         first_values= list_joints[0]
#         first_4_values= first_values[:4]
#         joints_names = [str(sample).split()[0] for sample in first_4_values] #takes from each joint only the first string which is the name
#         first_graph_name= "Distance " + joints_names[0] + " To " + joints_names[1]
#         second_graph_name= "Distance " + joints_names[2] + " To " + joints_names[3]
#
#         #create a list of x values
#         length= len(list_joints)
#         measurement_num = list(range(1, length + 1))
#
#         #create a data dic for graph
#         data = {
#         first_graph_name: {'x': measurement_num, 'y': right_distance},
#         second_graph_name: {'x': measurement_num, 'y': left_distance}}
#
#         create_and_save_graph(data, exercise_name)
#         create_and_save_table_with_calculations(data, exercise_name)


def create_and_save_graph(data, exercise):
    # Iterate over each plot data
    for plot_name, plot_data in data.items():
        # Create a new plot
        y_series = pd.Series(plot_data['y'])
        x_values = plot_data['x']
        y_values = y_series.values  # Keep NaN values

        # Check if all values in y_series are NaN or empty
        if y_series.isnull().all() or y_series.count() < 10:
            # Save a "null graph"
            timestamp = s.starts_and_ends_of_stops[0]
            start_dt = datetime.fromtimestamp(timestamp).strftime("%d-%m-%Y %H-%M-%S")
            create_and_open_folder(f'Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{start_dt}')
            plot_filename = f'Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{start_dt}/{plot_name}.jpeg'

            # Create a "null graph"
            plt.figure()
            plt.text(
                0.5, 0.5, "No Data Available", fontsize=18, color="gray", ha="center", va="center", alpha=0.7
            )
            plt.axis("off")  # Hide axes
            plt.title(plot_name[:-2], fontsize=16, weight="bold", y=0.9)
            plt.savefig(plot_filename, bbox_inches="tight", pad_inches=0, dpi=100)
            plt.close()
            continue

        # Plot the graph; matplotlib handles NaN by breaking the line
        plt.plot(x_values, y_values)

        # Highlight NaN values with red dots at y=0
        nan_indices = np.where(pd.isnull(y_values))[0]  # Find indices of NaN values
        if len(nan_indices) > 0:
            plt.scatter(
                [x_values[i] for i in nan_indices],
                [0 for _ in nan_indices],  # Placeholders at y=0 for NaN
                color='red',
                label="No Data",
                zorder=5
            )

        # Set the font size
        fontsize = 16

        plt.xlabel('מספר מדידה'[::-1], fontsize=fontsize, weight='bold')
        plt.ylabel('זווית'[::-1], fontsize=fontsize, weight='bold')
        plt.title(plot_name[:-2], fontsize=16, weight="bold", y=1)

        # Save the plot as an image file
        timestamp = s.starts_and_ends_of_stops[0]
        start_dt = datetime.fromtimestamp(timestamp).strftime("%d-%m-%Y %H-%M-%S")
        create_and_open_folder(f'Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{start_dt}')
        plot_filename = f'Patients/{s.chosen_patient_ID}/Graphs/{exercise}/{start_dt}/{plot_name}.jpeg'
        plt.savefig(plot_filename, dpi=100)
        plt.close()  # Close the plot to clear the figure


def success_worksheet():
    exercise, success_count = list(s.ex_list.items())[-1]

    # Check if the sheet exists
    if "success_worksheet" in s.training_workbook.sheetnames:
        # Get the existing worksheet
        success_sheet = s.training_workbook["success_worksheet"]
    else:
        # Create a new sheet if it doesn't exist
        success_sheet = s.training_workbook.create_sheet("success_worksheet")
        # Write headers in the first row (row=1)
        success_sheet.cell(row=1, column=1, value="exercise")
        success_sheet.cell(row=1, column=2, value="number of successful repetitions")

    # Find the first empty row in the worksheet
    row = success_sheet.max_row + 1

    # Write the exercise name and success count into the worksheet
    success_sheet.cell(row=row, column=1, value=exercise)
    success_sheet.cell(row=row, column=2, value=success_count)

    # Save the workbook after writing to the sheet
    s.training_workbook.save(s.training_workbook_path)


def find_and_change_values_exercises(new_values_dict, headers_row=1):
    # Load the workbook
    file_path = "Patients.xlsx"
    workbook = openpyxl.load_workbook(file_path)

    # Select the desired sheet
    sheet = workbook["patients_exercises"]

    # Find the column indices based on the header names
    column_indices = {}
    for header_name, new_value in new_values_dict.items():
        for cell in sheet[headers_row]:
            if cell.value == header_name:
                column_indices[header_name] = cell.column
                break

    # Iterate through the rows to find the value in the first column
    for row in sheet.iter_rows(min_row=headers_row + 1, max_row=sheet.max_row, min_col=1, max_col=1):
        cell = row[0]
        if str(cell.value) == s.chosen_patient_ID:
            # Update the values in the corresponding columns
            for header_name, column_index in column_indices.items():
                sheet.cell(row=cell.row, column=column_index, value=new_values_dict[header_name])

    # Save the changes
    workbook.save(file_path)


def calculate_training_length():
    if len(s.starts_and_ends_of_stops) % 2 != 0:
        s.starts_and_ends_of_stops.pop(-2)

    training_length = 0.0  # total in seconds

    for i in range(0, len(s.starts_and_ends_of_stops), 2):
        start_time = s.starts_and_ends_of_stops[i]
        end_time = s.starts_and_ends_of_stops[i + 1]
        print(f"Start: {start_time}, End: {end_time}")
        training_length += (end_time - start_time)

    return training_length  # float seconds


def find_and_change_values_patients(new_values_dict, headers_row=1):
    # Load the workbook
    file_path = "Patients.xlsx"
    workbook = openpyxl.load_workbook(file_path)

    # Select the desired sheet
    sheet = workbook["patients_details"]

    # Find the column indices based on the header names
    column_indices = {}
    for header_name, new_value in new_values_dict.items():
        for cell in sheet[headers_row]:
            if cell.value == header_name:
                column_indices[header_name] = cell.column
                break

    # Iterate through the rows to find the value in the first column
    for row in sheet.iter_rows(min_row=headers_row + 1, max_row=sheet.max_row, min_col=1, max_col=1):
        cell = row[0]
        if str(cell.value) == s.chosen_patient_ID:
            # Update the values in the corresponding columns
            for header_name, column_index in column_indices.items():
                sheet.cell(row=cell.row, column=column_index, value=new_values_dict[header_name])

    # Save the changes
    workbook.save(file_path)



def find_and_add_training_to_patient(headers_row=1):
    # Load the workbook
    file_path = "Patients.xlsx"
    workbook = openpyxl.load_workbook(file_path)

    # Select the desired sheet
    sheet = workbook["patients_history_of_trainings"]

    # Iterate through the rows to find the value in the first column
    for row in sheet.iter_rows(min_row=headers_row + 1, max_row=sheet.max_row, min_col=1, max_col=1):
        cell = row[0]
        if str(cell.value) == s.chosen_patient_ID:
            # Initialize next_column to 1
            next_column = 1

            # Find the next available column in the found row
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=cell.row, column=col).value is not None:
                    next_column = col + 1

            timestamp = s.starts_and_ends_of_stops[0]
            start_dt = datetime.fromtimestamp(timestamp).strftime("%d-%m-%Y %H-%M-%S")
            # Write the new value to the next available column in the found row
            sheet.cell(row=cell.row, column=next_column, value=start_dt)  # training dt, in the first place of the array there is the start time
            sheet.cell(row=cell.row, column=next_column + 1, value=(s.number_of_repetitions_in_training / s.max_repetitions_in_training))  # percent of the training that the patient managed to do
            sheet.cell(row=cell.row, column=next_column + 2, value=s.effort)  # percent of the training that the patient managed to do
            sheet.cell(row=cell.row, column=next_column + 3, value=calculate_training_length())  # percent of the training that the patient managed to do

            break  # Stop searching after finding the value

    workbook.save(file_path)


def which_welcome_record_to_say(headers_row=1):

    # Load the workbook
    file_path = "Patients.xlsx"
    workbook = openpyxl.load_workbook(file_path)

    # Select the desired sheet
    sheet = workbook["patients_history_of_trainings"]

    # Iterate through the rows to find the value in the first column
    for row in sheet.iter_rows(min_row=headers_row + 1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        if str(row[0].value) == str(s.chosen_patient_ID):  # Ensure comparison works for different types
            # Check if other cells in the row are not None or empty
            if any(cell.value for cell in row[1:]):  # Check if any other cell in the row has a value
                return "welcome"
            else:
                return "welcome_with_gymmy"

    # Return None if no matching patient is found
    return "welcome_with_gymmy"

# counts number of exercises in a training by ID by counting the true value
def count_number_of_exercises_in_training_by_ID():
    # Select the specific sheet
    workbook = openpyxl.load_workbook("Patients.xlsx")

    sheet = workbook["patients_exercises"]

    # Search for the row containing the search_value in the first column
    for row in sheet.iter_rows():
        if str(row[0].value) == s.chosen_patient_ID:
            row_number = row[0].row
            break
    else:
        # If the value is not found, return 0
        return 0

    # Get the values of the found row
    row_values = sheet[row_number]

    # Count the number of TRUE values in the row
    true_count = sum(1 for cell in row_values if cell.value is True)

    return true_count


def create_and_save_table_with_calculations(data, exercise):
    # Define the name of the file for saving the table image
    timestamp = s.starts_and_ends_of_stops[0]
    start_dt = datetime.fromtimestamp(timestamp).strftime("%d-%m-%Y %H-%M-%S")

    # Create and open the folder to save the tables
    create_and_open_folder(f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{start_dt}')

    # Set the maximum title length
    max_title_length = 32  # As specified

    # Iterate over each table data (each table if for an angle/distance)
    for table_name, table_data in data.items():
        # Perform calculations (min, max, avg, std)
        # Create a new plot
        y_series = pd.Series(table_data['y'])
        y_values = y_series.dropna().tolist()

        # Center-pad the title to 32 characters
        title_text = table_name[:-2]
        display_title = title_text.center(max_title_length)  # Pad evenly on both sides



        if len(y_values)>0:
            min_val = f"{min(y_values):.2f}"
            max_val = f"{max(y_values):.2f}"
            average = f"{(sum(y_values) / len(y_values)):.2f}"
            stdev = f"{np.std(y_values):.2f}"

        else:
            min_val = "אין נתונים"[::-1]
            max_val = "אין נתונים"[::-1]
            average = "אין נתונים"[::-1]
            stdev = "אין נתונים"[::-1]

        # Prepare data for the table
        calculation_data = {
            'ערכים'[::-1]: [min_val, max_val, average, stdev],  # Reverse Hebrew labels
            'מדדים'[::-1]: [s[::-1] for s in ['מינימום', 'מקסימום', 'ממוצע', 'סטיית תקן']]  # Reverse Hebrew labels
        }


        # Create a pandas DataFrame
        df = pd.DataFrame(calculation_data)

        # Create a new figure for the table only and set the background color
        fig, ax = plt.subplots(figsize=(2, 2))  # Adjust figure size to accommodate both table and header
        fig.patch.set_facecolor('#deeaf7')  # Set the background color of the figure

        # Hide axes completely (ensures no space around the table)
        ax.axis('off')

        # Add the table name as a header to the top of the figure
        ax.text(0.5, 0.9, display_title, ha='center', fontsize=13, weight='bold', transform=ax.transAxes)

        # Add the table to the figure with bold headers
        table = ax.table(cellText=df.values, colLabels=df.columns, cellLoc='center', loc='center')

        # Style the table
        table.auto_set_font_size(False)
        table.set_fontsize(12)
        table.scale(1.5, 1.3)  # Make the columns narrower by reducing the width scaling

        # Optionally, set the column width manually (you can remove this if not needed)
        # table.auto_set_column_width([0, 1])  # Adjust columns 0 and 1 to be narrower

        # Set the background color for the cells and the text properties
        for (i, j), cell in table.get_celld().items():
            if i == 0:  # Header row
                cell.set_text_props(weight='bold', fontsize=12)  # Set bold and increase font size
            else:
                cell.set_fontsize(12)  # Set a slightly smaller font for data rows

            cell.set_facecolor('#ffffff')  # White background for header cells

        timestamp = s.starts_and_ends_of_stops[0]
        start_dt = datetime.fromtimestamp(timestamp).strftime("%d-%m-%Y %H-%M-%S")

        # Save the table as an image with the background color and no transparency
        table_filename = f'Patients/{s.chosen_patient_ID}/Tables/{exercise}/{start_dt}/{table_name}.png'
        plt.savefig(table_filename, bbox_inches='tight', pad_inches=0, dpi=100)  # Removed transparent=True
        plt.close()  # Close the figure to clear memory



def close_workbook():
    s.training_workbook.close()


# def create_summary_workbook():
#         workbook_name = f"Patients/{s.chosen_patient_ID}/summary.xlsx"
#
#         # Create a new workbook
#         workbook = xlsxwriter.Workbook(workbook_name)
#
#         # Add a worksheet
#         workbook.add_worksheet("Sheet1")
#
#         # Close the workbook to save it
#         workbook.close()


# def add_exercise_to_summary(exercise_name, avg, sd, min_val, max_val, time_val):
#     # Load the existing workbook
#     file_path = f"Patients/{s.chosen_patient_ID}/summary.xlsx"
#     workbook = openpyxl.load_workbook(file_path)
#
#     # Check if "Sheet1" exists and rename it to the exercise name
#     if "Sheet1" in workbook.sheetnames:
#         sheet = workbook["Sheet1"]
#         sheet.title = exercise_name
#     # Check if the exercise sheet already exists, if not, create it
#     elif exercise_name not in workbook.sheetnames:
#         sheet = workbook.create_sheet(title=exercise_name)
#
#     # Add headers if it's a new or renamed sheet
#     if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:  # Check if it's an empty sheet
#         sheet.append(["Date Time", "Average", "Standard Deviation", "Minimum", "Maximum", "Time"])
#
#     # Find the next empty row (last row + 1)
#     next_row = sheet.max_row + 1
#
#     # Add data to the next available row, with the first column being the datetime
#     sheet.cell(row=next_row, column=1, value=s.start_dt)  # Datetime from s.start_dt
#     sheet.cell(row=next_row, column=2, value=avg)  # Average
#     sheet.cell(row=next_row, column=3, value=sd)  # Standard Deviation
#     sheet.cell(row=next_row, column=4, value=min_val)  # Minimum
#     sheet.cell(row=next_row, column=5, value=max_val)  # Maximum
#     sheet.cell(row=next_row, column=6, value=time_val)  # Time
#
#     # Save the workbook
#     workbook.save(file_path)









if __name__ == "__main__":

    s.chosen_patient_ID="55581599"
    print(which_welcome_record_to_say())
    # Example usage
